"""
Carga Excel -> SQL Server para el modulo de categorizacion medica.

Esta version esta ajustada al Excel actual:
    - DIM_PAIS
    - DimComponente
    - DimClasificacion
    - DimRegla
    - DimEquipo
    - DIM_REPRESENTANTE_MEDICO
    - DimEspecialidad
    - DimCentroMedico
    - DimGeografia
    - FactMedicoInput

Requiere:
    pip install openpyxl pyodbc

Ejemplo:
    python cargar_categorizacion_medica_excel_sqlserver.py ^
        --excel modelo_sqlserver_categorizacion_medica_excel.xlsx ^
        --server MI_SERVIDOR ^
        --database MI_BASE ^
        --periodo 2026-10 ^
        --archivo-origen "modelo_sqlserver_categorizacion_medica_excel.xlsx"
"""

from __future__ import annotations

import argparse
from datetime import datetime, date
from pathlib import Path
from typing import Any

import pyodbc
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


NULL_STRINGS = {"", "NULL", "None", "nan"}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value in NULL_STRINGS:
            return None
        return value
    if isinstance(value, datetime):
        return value.date()
    return value


def as_int(value: Any) -> int | None:
    value = clean(value)
    if value is None:
        return None
    return int(value)


def as_date(value: Any) -> date | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.fromisoformat(value.replace(" 00:00:00", "")).date()
    return value


def connect_sqlserver(server: str, database: str, trusted: bool, user: str | None, password: str | None):
    if trusted:
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={server};DATABASE={database};Trusted_Connection=yes;"
        )
    else:
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={server};DATABASE={database};UID={user};PWD={password};"
        )
    return pyodbc.connect(conn_str)


def read_table(excel_path: Path, sheet_name: str, table_name: str | None = None) -> list[dict[str, Any]]:
    # data_only=True lee los valores calculados de formulas como Equipo y linea_id.
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    if table_name is None:
        if not ws.tables:
            raise ValueError(f"La hoja {sheet_name} no tiene tabla definida.")
        table = next(iter(ws.tables.values()))
    else:
        table = ws.tables[table_name]

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [clean(ws.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(min_row + 1, max_row + 1):
        values = [clean(ws.cell(row_idx, col).value) for col in range(min_col, max_col + 1)]
        if all(value is None for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def create_load_batch(cursor, archivo_origen: str, periodo: str, codigo_pais_default: str | None) -> int:
    cursor.execute(
        """
        INSERT INTO cat.LoadBatch (ArchivoOrigen, Periodo, CodigoPaisDefault, UsuarioCarga)
        OUTPUT INSERTED.LoadBatchKey
        VALUES (?, ?, ?, SUSER_SNAME());
        """,
        archivo_origen,
        periodo,
        codigo_pais_default,
    )
    return int(cursor.fetchone()[0])


def pais_key(cursor, codigo_pais: str) -> int:
    cursor.execute("SELECT PaisKey FROM cat.DimPais WHERE CodigoPais = ?", codigo_pais)
    row = cursor.fetchone()
    if row is None:
        raise ValueError(f"No existe pais en cat.DimPais: {codigo_pais}")
    return int(row[0])


def componente_key(cursor, codigo_componente: str) -> int:
    cursor.execute("SELECT ComponenteKey FROM cat.DimComponenteCategoria WHERE CodigoComponente = ?", codigo_componente)
    row = cursor.fetchone()
    if row is None:
        raise ValueError(f"No existe componente en cat.DimComponenteCategoria: {codigo_componente}")
    return int(row[0])


def merge_dim_pais(cursor, rows: list[dict[str, Any]]):
    sql = """
    MERGE cat.DimPais AS tgt
    USING (SELECT ? AS PaisIdOrigen, ? AS CodigoPais, ? AS NombrePais, ? AS Moneda, ? AS ZonaHoraria, ? AS Activo) AS src
       ON tgt.CodigoPais = src.CodigoPais
    WHEN MATCHED THEN UPDATE SET
        PaisIdOrigen = src.PaisIdOrigen,
        NombrePais = src.NombrePais,
        Moneda = src.Moneda,
        ZonaHoraria = src.ZonaHoraria,
        Activo = src.Activo
    WHEN NOT MATCHED THEN INSERT (PaisIdOrigen, CodigoPais, NombrePais, Moneda, ZonaHoraria, Activo)
        VALUES (src.PaisIdOrigen, src.CodigoPais, src.NombrePais, src.Moneda, src.ZonaHoraria, src.Activo);
    """
    for r in rows:
        cursor.execute(sql, as_int(r.get("id")), r.get("codigo"), r.get("nombre"), r.get("moneda"), r.get("zona_horaria"), as_int(r.get("activo")) or 0)


def merge_dim_componente(cursor, rows: list[dict[str, Any]]):
    sql = """
    MERGE cat.DimComponenteCategoria AS tgt
    USING (SELECT ? AS CodigoComponente, ? AS NombreComponente, ? AS TipoEvaluacion,
                  ? AS PesoComponentePct, ? AS Requerido, ? AS Activo) AS src
       ON tgt.CodigoComponente = src.CodigoComponente
    WHEN MATCHED THEN UPDATE SET
        NombreComponente = src.NombreComponente,
        TipoEvaluacion = src.TipoEvaluacion,
        PesoComponentePct = src.PesoComponentePct,
        Requerido = src.Requerido,
        Activo = src.Activo
    WHEN NOT MATCHED THEN INSERT (CodigoComponente, NombreComponente, TipoEvaluacion, PesoComponentePct, Requerido, Activo)
        VALUES (src.CodigoComponente, src.NombreComponente, src.TipoEvaluacion, src.PesoComponentePct, src.Requerido, src.Activo);
    """
    for r in rows:
        cursor.execute(sql, r["CodigoComponente"], r["NombreComponente"], r["TipoEvaluacion"], r["PesoComponentePct"], as_int(r["Requerido"]), as_int(r["Activo"]))


def merge_dim_clasificacion(cursor, rows: list[dict[str, Any]]):
    sql = """
    MERGE cat.DimClasificacionMedica AS tgt
    USING (SELECT ? AS PaisKey, ? AS Clase, ? AS PuntajeMinPct, ? AS PuntajeMaxPct,
                  ? AS OrdenClase, ? AS VigenteDesde, ? AS VigenteHasta, ? AS Activo) AS src
       ON tgt.PaisKey = src.PaisKey AND tgt.Clase = src.Clase AND tgt.VigenteDesde = src.VigenteDesde
    WHEN MATCHED THEN UPDATE SET
        PuntajeMinPct = src.PuntajeMinPct,
        PuntajeMaxPct = src.PuntajeMaxPct,
        OrdenClase = src.OrdenClase,
        VigenteHasta = src.VigenteHasta,
        Activo = src.Activo
    WHEN NOT MATCHED THEN INSERT (PaisKey, Clase, PuntajeMinPct, PuntajeMaxPct, OrdenClase, VigenteDesde, VigenteHasta, Activo)
        VALUES (src.PaisKey, src.Clase, src.PuntajeMinPct, src.PuntajeMaxPct, src.OrdenClase, src.VigenteDesde, src.VigenteHasta, src.Activo);
    """
    for r in rows:
        cursor.execute(sql, pais_key(cursor, r["pais_codigo"]), r["Clase"], r["PuntajeMinPct"], r["PuntajeMaxPct"], as_int(r["OrdenClase"]), as_date(r["VigenteDesde"]), as_date(r.get("VigenteHasta")), as_int(r["Activo"]))


def merge_dim_regla(cursor, rows: list[dict[str, Any]]):
    sql = """
    MERGE cat.DimReglaCategoriaMedica AS tgt
    USING (SELECT ? AS PaisKey, ? AS ComponenteKey, ? AS CodigoRegla, ? AS Detalle,
                  ? AS ValorMinimo, ? AS ValorMaximo, ? AS ValorTexto, ? AS Criterio,
                  ? AS PesoComponentePct, ? AS PuntajePct, ? AS VigenteDesde,
                  ? AS VigenteHasta, ? AS Activo) AS src
       ON tgt.PaisKey = src.PaisKey
      AND tgt.ComponenteKey = src.ComponenteKey
      AND tgt.CodigoRegla = src.CodigoRegla
      AND tgt.VigenteDesde = src.VigenteDesde
    WHEN MATCHED THEN UPDATE SET
        Detalle = src.Detalle,
        ValorMinimo = src.ValorMinimo,
        ValorMaximo = src.ValorMaximo,
        ValorTexto = src.ValorTexto,
        Criterio = src.Criterio,
        PesoComponentePct = src.PesoComponentePct,
        PuntajePct = src.PuntajePct,
        VigenteHasta = src.VigenteHasta,
        Activo = src.Activo
    WHEN NOT MATCHED THEN INSERT (
        PaisKey, ComponenteKey, CodigoRegla, Detalle, ValorMinimo, ValorMaximo,
        ValorTexto, Criterio, PesoComponentePct, PuntajePct, VigenteDesde, VigenteHasta, Activo
    )
    VALUES (
        src.PaisKey, src.ComponenteKey, src.CodigoRegla, src.Detalle, src.ValorMinimo, src.ValorMaximo,
        src.ValorTexto, src.Criterio, src.PesoComponentePct, src.PuntajePct, src.VigenteDesde, src.VigenteHasta, src.Activo
    );
    """
    for r in rows:
        cursor.execute(
            sql,
            pais_key(cursor, r["pais_codigo"]),
            componente_key(cursor, r["CodigoComponente"]),
            r["CodigoRegla"],
            str(r["Detalle"]),
            r.get("ValorMinimo"),
            r.get("ValorMaximo"),
            r.get("ValorTexto"),
            as_int(r["Criterio"]),
            r["PesoComponentePct"],
            r["PuntajePct"],
            as_date(r["VigenteDesde"]),
            as_date(r.get("VigenteHasta")),
            as_int(r["Activo"]),
        )


def merge_dim_equipo(cursor, rows: list[dict[str, Any]]):
    sql = """
    MERGE cat.DimEquipo AS tgt
    USING (SELECT ? AS PaisKey, ? AS CodigoEquipo, ? AS NombreEquipo, ? AS Descripcion, ? AS Activo) AS src
       ON tgt.PaisKey = src.PaisKey AND tgt.CodigoEquipo = src.CodigoEquipo
    WHEN MATCHED THEN UPDATE SET NombreEquipo = src.NombreEquipo, Descripcion = src.Descripcion, Activo = src.Activo
    WHEN NOT MATCHED THEN INSERT (PaisKey, CodigoEquipo, NombreEquipo, Descripcion, Activo)
        VALUES (src.PaisKey, src.CodigoEquipo, src.NombreEquipo, src.Descripcion, src.Activo);
    """
    for r in rows:
        cursor.execute(sql, pais_key(cursor, r["pais_codigo"]), r["codigo"], r["nombre"], r.get("descripcion"), as_int(r["activo"]))


def merge_dim_representante(cursor, rows: list[dict[str, Any]]):
    sql = """
    MERGE cat.DimRepresentanteMedico AS tgt
    USING (SELECT ? AS PaisKey, ? AS RepresentanteIdOrigen, ? AS CodigoRepresentante,
                  ? AS NombreRepresentante, ? AS LineaIdOrigen, ? AS GerenteIdOrigen,
                  ? AS Email, ? AS Zona, ? AS FechaIngreso, ? AS Cedula,
                  ? AS CodigoOrigenExcel, ? AS EquipoTexto, ? AS Activo) AS src
       ON tgt.PaisKey = src.PaisKey AND tgt.CodigoRepresentante = src.CodigoRepresentante
    WHEN MATCHED THEN UPDATE SET
        RepresentanteIdOrigen = src.RepresentanteIdOrigen,
        NombreRepresentante = src.NombreRepresentante,
        LineaIdOrigen = src.LineaIdOrigen,
        GerenteIdOrigen = src.GerenteIdOrigen,
        Email = src.Email,
        Zona = src.Zona,
        FechaIngreso = src.FechaIngreso,
        Cedula = src.Cedula,
        CodigoOrigenExcel = src.CodigoOrigenExcel,
        EquipoTexto = src.EquipoTexto,
        Activo = src.Activo
    WHEN NOT MATCHED THEN INSERT (
        PaisKey, RepresentanteIdOrigen, CodigoRepresentante, NombreRepresentante, LineaIdOrigen,
        GerenteIdOrigen, Email, Zona, FechaIngreso, Cedula, CodigoOrigenExcel, EquipoTexto, Activo
    )
    VALUES (
        src.PaisKey, src.RepresentanteIdOrigen, src.CodigoRepresentante, src.NombreRepresentante, src.LineaIdOrigen,
        src.GerenteIdOrigen, src.Email, src.Zona, src.FechaIngreso, src.Cedula, src.CodigoOrigenExcel, src.EquipoTexto, src.Activo
    );
    """
    for r in rows:
        cursor.execute(
            sql,
            pais_key(cursor, r["pais_codigo"]),
            as_int(r.get("id")),
            r["codigo_id"],
            r["nombre"],
            as_int(r.get("linea_id")),
            as_int(r.get("gerente_id")),
            r.get("email"),
            r.get("zona"),
            as_date(r.get("fecha_ingreso")),
            r.get("cedula"),
            r.get("codigo_origen_excel"),
            r.get("Equipo"),
            as_int(r.get("activo")),
        )


def merge_simple_catalog(cursor, table: str, rows: list[dict[str, Any]], value_col: str, sql_col: str):
    sql = f"""
    MERGE {table} AS tgt
    USING (SELECT ? AS PaisKey, ? AS Valor) AS src
       ON tgt.PaisKey = src.PaisKey AND tgt.{sql_col} = src.Valor
    WHEN NOT MATCHED THEN INSERT (PaisKey, {sql_col}) VALUES (src.PaisKey, src.Valor);
    """
    for r in rows:
        cursor.execute(sql, pais_key(cursor, r["pais_codigo"]), r[value_col])


def merge_dim_geografia(cursor, rows: list[dict[str, Any]]):
    sql = """
    MERGE cat.DimGeografia AS tgt
    USING (SELECT ? AS PaisKey, ? AS Provincia, ? AS Municipio) AS src
       ON tgt.PaisKey = src.PaisKey AND tgt.Provincia = src.Provincia AND tgt.Municipio = src.Municipio
    WHEN NOT MATCHED THEN INSERT (PaisKey, Provincia, Municipio) VALUES (src.PaisKey, src.Provincia, src.Municipio);
    """
    for r in rows:
        cursor.execute(sql, pais_key(cursor, r["pais_codigo"]), r["Provincia"], r["Municipio"])


def insert_fact_input(cursor, load_batch_key: int, rows: list[dict[str, Any]]):
    sql = """
    INSERT INTO stg.MedicoCategoriaInput (
        LoadBatchKey, RowNumber, CodigoPais, Periodo, Equipo, LineaIdOrigen,
        CodigoRepresentante, NombreRepresentante, Medico, CentroMedico, Especialidad,
        Provincia, Municipio, PacientesSemana, CostoConsulta, RecetasSemana,
        UbicacionTerritorialCM, KOL, CategoriaExcel, Activo
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
    """
    params = []
    for idx, r in enumerate(rows, start=1):
        params.append((
            load_batch_key,
            idx,
            r["pais_codigo"],
            r["Periodo"],
            r.get("Equipo"),
            as_int(r.get("linea_id")),
            r.get("codigo_id"),
            r.get("nombre"),
            r.get("Medico"),
            r.get("CentroMedico"),
            r.get("Especialidad"),
            r.get("Provincia"),
            r.get("Municipio"),
            r.get("PacientesSemana"),
            r.get("CostoConsulta"),
            r.get("RecetasSemana"),
            r.get("UbicacionTerritorialCM"),
            r.get("KOL"),
            r.get("CategoriaExcel"),
            as_int(r.get("Activo")) or 1,
        ))
    cursor.fast_executemany = True
    cursor.executemany(sql, params)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--server", required=True)
    parser.add_argument("--database", required=True)
    parser.add_argument("--periodo", required=True)
    parser.add_argument("--archivo-origen", required=True)
    parser.add_argument("--codigo-pais-default", default="DO")
    parser.add_argument("--trusted", action="store_true", default=True)
    parser.add_argument("--user")
    parser.add_argument("--password")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    conn = connect_sqlserver(args.server, args.database, args.trusted, args.user, args.password)
    conn.autocommit = False

    try:
        cursor = conn.cursor()

        merge_dim_pais(cursor, read_table(excel_path, "DIM_PAIS", "tbl_DimPais"))
        merge_dim_componente(cursor, read_table(excel_path, "DimComponente", "tbl_DimComponente"))
        merge_dim_clasificacion(cursor, read_table(excel_path, "DimClasificacion", "tbl_DimClasificacion"))
        merge_dim_regla(cursor, read_table(excel_path, "DimRegla", "tbl_DimRegla"))
        merge_dim_equipo(cursor, read_table(excel_path, "DimEquipo"))
        merge_dim_representante(cursor, read_table(excel_path, "DIM_REPRESENTANTE_MEDICO", "tbl_DimRepresentante"))
        merge_simple_catalog(cursor, "cat.DimEspecialidad", read_table(excel_path, "DimEspecialidad", "tbl_DimEspecialidad"), "Especialidad", "Especialidad")
        merge_simple_catalog(cursor, "cat.DimCentroMedico", read_table(excel_path, "DimCentroMedico", "tbl_DimCentroMedico"), "CentroMedico", "CentroMedico")
        merge_dim_geografia(cursor, read_table(excel_path, "DimGeografia", "tbl_DimGeografia"))

        load_batch_key = create_load_batch(
            cursor,
            archivo_origen=args.archivo_origen,
            periodo=args.periodo,
            codigo_pais_default=args.codigo_pais_default,
        )

        fact_rows = read_table(excel_path, "FactMedicoInput", "tbl_FactMedicoInput")
        insert_fact_input(cursor, load_batch_key, fact_rows)
        cursor.execute("EXEC cat.sp_CalcularCategoriaMedica @LoadBatchKey = ?", load_batch_key)

        conn.commit()
        print(f"Carga y calculo completados. LoadBatchKey={load_batch_key}. Filas fact: {len(fact_rows)}")

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
