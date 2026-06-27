"""
SCGCPR — Router: Importación de Dimensiones (DIMs)
====================================================
Permite cargar masivamente las tablas de dimensiones desde un archivo Excel
(.xlsx) que contiene múltiples hojas, una por cada DIM.

Endpoints:
  POST /dims/preview  → Sube el Excel y retorna las hojas detectadas con
                        su mapeo propuesto a tabla del sistema
  POST /dims/importar → Sube el Excel + lista de hojas seleccionadas
                        y carga los datos en BD
"""
import io
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.deps import get_db, get_current_active_user, require_roles
from app.models.usuario import Rol
from app.models.dimensiones import (
    Pais, Linea, Gerente, RepresentanteMedico,
    Indicador, IndicadorTabla, Ciclo, Mes,
)

router = APIRouter(prefix="/dims", tags=["Importación DIMs"])

# Solo ADMIN y GERENTE_PRODUCTIVIDAD pueden importar
AdminOGerProd = Depends(require_roles(Rol.ADMIN, Rol.GERENTE_PRODUCTIVIDAD))

# ── Mapeo de nombres de hojas a información descriptiva ────────────────
HOJAS_CONOCIDAS = {
    "DIM_PAIS":            {"tabla": "DIM_Pais",           "label": "Países",                   "orden": 1},
    "DIM_LINEA":           {"tabla": "DIM_Linea",          "label": "Líneas de Productos",       "orden": 2},
    "DIM_GERENTE":         {"tabla": "DIM_Gerente",        "label": "Gerentes",                  "orden": 3},
    "DIM_RM":              {"tabla": "DIM_RM",             "label": "Representantes Médicos",     "orden": 4},
    "DIM_INDICADOR":       {"tabla": "DIM_Indicador",      "label": "Indicadores de Desempeño",  "orden": 5},
    "DIM_INDICADOR_TABLA": {"tabla": "DIM_IndicadorTabla", "label": "Rangos de Puntuación",      "orden": 6},
    "DIM_CICLO":           {"tabla": "DIM_Ciclo",          "label": "Ciclos de Trabajo",         "orden": 7},
    "DIM_MES":             {"tabla": "DIM_Mes",            "label": "Meses del Año",             "orden": 8},
}


# ── Schemas ────────────────────────────────────────────────────────────

class HojaInfo(BaseModel):
    """Información de una hoja detectada en el Excel."""
    nombre_hoja: str        # Nombre exacto de la hoja en el Excel
    tabla_sistema: str      # Nombre de la tabla en BD
    label: str              # Nombre amigable para mostrar al usuario
    filas: int              # Número de filas de datos (sin header)
    columnas: List[str]     # Columnas detectadas
    reconocida: bool        # True si es una DIM conocida del sistema
    orden: int              # Orden sugerido de importación

class PreviewResponse(BaseModel):
    """Respuesta del endpoint preview con las hojas detectadas."""
    hojas: List[HojaInfo]
    total_hojas: int
    hojas_reconocidas: int

class ResultadoHoja(BaseModel):
    """Resultado de importar una hoja específica."""
    nombre_hoja: str
    label: str
    exitoso: bool
    insertados: int
    omitidos: int          # Registros que ya existían
    errores: int
    mensaje: str

class ImportarResponse(BaseModel):
    """Respuesta completa de la importación."""
    resultados: List[ResultadoHoja]
    total_insertados: int
    total_omitidos: int
    total_errores: int


# ── Helpers de conversión ──────────────────────────────────────────────

def _leer_hoja(ws) -> list[dict]:
    """Lee una hoja de openpyxl y retorna lista de dicts {COLUMNA: valor}."""
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip().upper() if cell.value is not None else "")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows

def _s(v) -> Optional[str]:
    return str(v).strip() if v is not None else None

def _i(v) -> Optional[int]:
    try: return int(v)
    except: return None

def _d(v) -> Optional[Decimal]:
    try: return Decimal(str(v))
    except: return None

def _fecha(v) -> Optional[date]:
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except: return None


# ── Endpoint 1: Preview ────────────────────────────────────────────────

@router.post("/preview", response_model=PreviewResponse,
             summary="Leer Excel y detectar hojas disponibles")
async def preview_dims(
    file: UploadFile = File(..., description="Archivo Excel (.xlsx) con hojas DIM_*"),
    _=AdminOGerProd,
):
    """
    Sube un archivo Excel y retorna la lista de hojas detectadas.
    Para cada hoja indica: nombre, tabla del sistema, filas, columnas
    y si es una DIM reconocida.
    No modifica la base de datos.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    contenido = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

    hojas = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        nombre_upper = nombre_hoja.strip().upper()
        info_conocida = HOJAS_CONOCIDAS.get(nombre_upper, {})

        # Leer headers (fila 1)
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers.append(str(cell.value).strip())

        # Contar filas de datos
        filas = max(0, ws.max_row - 1)

        hojas.append(HojaInfo(
            nombre_hoja=nombre_hoja,
            tabla_sistema=info_conocida.get("tabla", nombre_hoja),
            label=info_conocida.get("label", nombre_hoja),
            filas=filas,
            columnas=headers,
            reconocida=bool(info_conocida),
            orden=info_conocida.get("orden", 99),
        ))

    wb.close()

    # Ordenar: reconocidas primero, luego por orden
    hojas.sort(key=lambda h: (0 if h.reconocida else 1, h.orden))

    return PreviewResponse(
        hojas=hojas,
        total_hojas=len(hojas),
        hojas_reconocidas=sum(1 for h in hojas if h.reconocida),
    )


# ── Endpoint 2: Importar ───────────────────────────────────────────────

@router.post("/importar", response_model=ImportarResponse,
             summary="Importar hojas seleccionadas del Excel a la BD")
async def importar_dims(
    file: UploadFile = File(..., description="Archivo Excel (.xlsx) con hojas DIM_*"),
    hojas: str = Form(..., description="JSON con lista de nombres de hojas a importar"),
    db: Session = Depends(get_db),
    _=AdminOGerProd,
):
    """
    Importa las hojas seleccionadas del Excel a la base de datos.

    - Respeta el orden correcto de carga (Pais → Linea → Gerente → RM → ...)
    - Registros ya existentes se omiten (no se duplican)
    - Retorna un resumen por hoja: insertados, omitidos, errores
    """
    import json
    try:
        hojas_seleccionadas: List[str] = json.loads(hojas)
    except Exception:
        raise HTTPException(400, "El campo 'hojas' debe ser un JSON array de strings")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    contenido = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)

    resultados = []
    total_ins = total_om = total_err = 0

    try:

        # Ordenar las hojas seleccionadas según el orden correcto de FK
        def orden_hoja(nombre):
            info = HOJAS_CONOCIDAS.get(nombre.upper(), {})
            return info.get("orden", 99)

        hojas_ordenadas = sorted(hojas_seleccionadas, key=orden_hoja)

        # Construir mapas de lookups para resolver FKs
        pais_map = {p.codigo: p.id for p in db.query(Pais).filter(Pais.activo == True).all()}
        linea_map = {}  # (pais_codigo, linea_id_original) → id
        gerente_map = {}  # (pais_codigo, gerente_id_original) → id

        # Pre-cargar mapas si ya hay datos en BD
        for lin in db.query(Linea).all():
            pais_codigo = next((k for k, v in pais_map.items() if v == lin.pais_id), None)
            if pais_codigo:
                linea_map[(pais_codigo, lin.codigo)] = lin.id

        for ger in db.query(Gerente).all():
            pais_codigo = next((k for k, v in pais_map.items() if v == ger.pais_id), None)
            if pais_codigo:
                gerente_map[(pais_codigo, ger.codigo)] = ger.id

        indicador_map = {}
        for ind in db.query(Indicador).all():
            pais_codigo = next((k for k, v in pais_map.items() if v == ind.pais_id), None)
            if pais_codigo:
                indicador_map[(pais_codigo, ind.codigo)] = ind.id

        for nombre_hoja in hojas_ordenadas:
            # Buscar la hoja (case-insensitive)
            ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.strip().upper() == nombre_hoja.strip().upper():
                    ws = wb[sheet_name]
                    nombre_hoja = sheet_name
                    break

            info = HOJAS_CONOCIDAS.get(nombre_hoja.strip().upper(), {})
            label = info.get("label", nombre_hoja)

            if ws is None:
                resultados.append(ResultadoHoja(
                    nombre_hoja=nombre_hoja, label=label,
                    exitoso=False, insertados=0, omitidos=0, errores=1,
                    mensaje=f"Hoja '{nombre_hoja}' no encontrada en el archivo",
                ))
                total_err += 1
                continue

            try:
                ins, om, err, msg = _importar_hoja(
                    db, nombre_hoja.upper(), ws,
                    pais_map, linea_map, gerente_map, indicador_map
                )
                resultados.append(ResultadoHoja(
                    nombre_hoja=nombre_hoja, label=label,
                    exitoso=err == 0, insertados=ins, omitidos=om, errores=err,
                    mensaje=msg,
                ))
                total_ins += ins; total_om += om; total_err += err
            except Exception as e:
                db.rollback()
                resultados.append(ResultadoHoja(
                    nombre_hoja=nombre_hoja, label=label,
                    exitoso=False, insertados=0, omitidos=0, errores=1,
                    mensaje=str(e)[:300],
                ))
                total_err += 1

    finally:
        wb.close()

    return ImportarResponse(
        resultados=resultados,
        total_insertados=total_ins,
        total_omitidos=total_om,
        total_errores=total_err,
    )


# ── Función interna: importar cada hoja ───────────────────────────────

def _importar_hoja(db, nombre_upper, ws, pais_map, linea_map, gerente_map, indicador_map):
    """
    Importa una hoja específica. Retorna (insertados, omitidos, errores, mensaje).
    Actualiza los mapas de lookup con los nuevos IDs insertados.
    """
    rows = _leer_hoja(ws)
    ins = om = err = 0

    if nombre_upper == "DIM_PAIS":
        for row in rows:
            codigo = _s(row.get("PAIS_CODIGO"))
            if not codigo: continue
            if db.query(Pais).filter(Pais.codigo == codigo).first():
                om += 1; continue
            obj = Pais(
                codigo=codigo,
                nombre=_s(row.get("NOMBRE")) or codigo,
                moneda=_s(row.get("MONEDA")),
                zona_horaria=_s(row.get("ZONA_HORARIA")),
                activo=True,
            )
            db.add(obj); db.flush()
            pais_map[codigo] = obj.id
            ins += 1
        db.commit()

    elif nombre_upper == "DIM_LINEA":
        for row in rows:
            pais_c = _s(row.get("PAIS_CODIGO"))
            codigo = _s(row.get("CODIGO"))
            linea_id_orig = _i(row.get("LINEA_ID"))
            pais_id = pais_map.get(pais_c)
            if not pais_id or not codigo: err += 1; continue
            if db.query(Linea).filter(Linea.pais_id == pais_id, Linea.codigo == codigo).first():
                linea_map[(pais_c, linea_id_orig)] = db.query(Linea).filter(Linea.pais_id == pais_id, Linea.codigo == codigo).first().id
                om += 1; continue
            obj = Linea(pais_id=pais_id, codigo=codigo, nombre=_s(row.get("NOMBRE")) or codigo, activo=True)
            db.add(obj); db.flush()
            linea_map[(pais_c, linea_id_orig)] = obj.id
            if codigo: linea_map[(pais_c, codigo)] = obj.id
            ins += 1
        db.commit()

    elif nombre_upper == "DIM_GERENTE":
        for row in rows:
            codigo = _s(row.get("CODIGO"))
            pais_c = _s(row.get("PAIS_CODIGO"))
            ger_id_orig = _i(row.get("GERENTE_ID"))
            pais_id = pais_map.get(pais_c)
            if not pais_id or not codigo: err += 1; continue
            existing = db.query(Gerente).filter(Gerente.codigo == codigo).first()
            if existing:
                gerente_map[(pais_c, ger_id_orig)] = existing.id
                gerente_map[(pais_c, codigo)] = existing.id
                om += 1; continue
            obj = Gerente(
                pais_id=pais_id, codigo=codigo,
                nombre=_s(row.get("NOMBRE")) or codigo,
                tipo=_s(row.get("TIPO")) or "DISTRITO", activo=True,
            )
            db.add(obj); db.flush()
            gerente_map[(pais_c, ger_id_orig)] = obj.id
            gerente_map[(pais_c, codigo)] = obj.id
            ins += 1
        db.commit()

    elif nombre_upper == "DIM_RM":
        for row in rows:
            rm_codigo = _s(row.get("RM_CODIGO"))
            pais_c = _s(row.get("PAIS_CODIGO"))
            linea_id_orig = _i(row.get("LINEA_ID"))
            ger_id_orig = _i(row.get("GERENTE_ID"))
            pais_id = pais_map.get(pais_c)
            linea_id = linea_map.get((pais_c, linea_id_orig))
            gerente_id = gerente_map.get((pais_c, ger_id_orig))
            if not pais_id or not linea_id or not rm_codigo: err += 1; continue
            if db.query(RepresentanteMedico).filter(RepresentanteMedico.codigo == str(rm_codigo)).first():
                om += 1; continue
            obj = RepresentanteMedico(
                pais_id=pais_id, linea_id=linea_id, gerente_id=gerente_id,
                codigo=str(rm_codigo),
                nombre=_s(row.get("NOMBRE")) or str(rm_codigo),
                cedula=_s(row.get("CEDULA")),
                email=_s(row.get("EMAIL")),
                fecha_ingreso=_fecha(row.get("FECHA_INGRESO")),
                activo=True,
            )
            db.add(obj); ins += 1
        db.commit()

    elif nombre_upper == "DIM_INDICADOR":
        for row in rows:
            pais_c = _s(row.get("PAIS_CODIGO"))
            codigo = _s(row.get("CODIGO"))
            pais_id = pais_map.get(pais_c)
            if not pais_id or not codigo: err += 1; continue
            if db.query(Indicador).filter(Indicador.pais_id == pais_id, Indicador.codigo == codigo).first():
                indicador_map[(pais_c, codigo)] = db.query(Indicador).filter(Indicador.pais_id == pais_id, Indicador.codigo == codigo).first().id
                om += 1; continue
            pond = _i(row.get("PONDERACION_PCT")) or 0
            obj = Indicador(
                pais_id=pais_id, codigo=codigo,
                nombre=_s(row.get("NOMBRE")) or codigo,
                rol=_s(row.get("ROL")) or "RM",
                modulo=_s(row.get("MODULO")) or "GESTION",
                tipo_periodo=_s(row.get("TIPO_PERIODO")) or "CICLO",
                ponderacion_pct=pond,
                escala=_i(row.get("ESCALA")) or 1,
                valor_min=_d(row.get("VALOR_MIN")),
                valor_max=_d(row.get("VALOR_MAX")),
                peso_iup=Decimal(str(pond)) / Decimal("100"),
                activo=True,
            )
            db.add(obj); db.flush()
            indicador_map[(pais_c, codigo)] = obj.id
            ins += 1
        db.commit()

    elif nombre_upper == "DIM_INDICADOR_TABLA":
        for row in rows:
            ind_codigo = _s(row.get("INDICADOR_CODIGO"))
            pais_c = _s(row.get("PAIS_CODIGO"))
            pais_id = pais_map.get(pais_c)
            ind_id = indicador_map.get((pais_c, ind_codigo))
            if not pais_id or not ind_id: err += 1; continue
            desde = _d(row.get("VALOR_DESDE"))
            hasta = _d(row.get("VALOR_HASTA"))
            if db.query(IndicadorTabla).filter(
                IndicadorTabla.indicador_id == ind_id,
                IndicadorTabla.pais_id == pais_id,
                IndicadorTabla.rango_desde == desde,
            ).first():
                om += 1; continue
            db.add(IndicadorTabla(
                indicador_id=ind_id, pais_id=pais_id,
                rango_desde=desde, rango_hasta=hasta,
                puntos=_d(row.get("PUNTOS")), activo=True,
            ))
            ins += 1
        db.commit()

    elif nombre_upper == "DIM_CICLO":
        trimestre_map = {1:1,2:1,3:1,4:2,5:2,6:2,7:3,8:3,9:3,10:4,11:4,12:4}
        semestre_map  = {1:1,2:1,3:1,4:1,5:1,6:1,7:2,8:2,9:2,10:2,11:2,12:2}
        for row in rows:
            numero = _i(row.get("CICLO_ID"))
            nombre = _s(row.get("CICLO_NOMBRE_CORTO"))
            canon  = _s(row.get("NOMBRE_CANONICO"))
            if not numero: continue
            for pais_c, pais_id in pais_map.items():
                if db.query(Ciclo).filter(Ciclo.pais_id == pais_id, Ciclo.numero == numero, Ciclo.anio == 2026).first():
                    om += 1; continue
                db.add(Ciclo(
                    pais_id=pais_id, anio=2026, numero=numero,
                    nombre=nombre, nombre_canonico=canon,
                    fecha_inicio=date(2026, 1, 1), fecha_fin=date(2026, 12, 31),
                    dias_laborables=21, cerrado=False, activo=True,
                ))
                ins += 1
        db.commit()

    elif nombre_upper == "DIM_MES":
        trimestre_map = {1:1,2:1,3:1,4:2,5:2,6:2,7:3,8:3,9:3,10:4,11:4,12:4}
        semestre_map  = {1:1,2:1,3:1,4:1,5:1,6:1,7:2,8:2,9:2,10:2,11:2,12:2}
        for row in rows:
            mes_num = _i(row.get("MESID"))
            if not mes_num: continue
            if db.query(Mes).filter(Mes.mes == mes_num, Mes.anio == 2026).first():
                om += 1; continue
            db.add(Mes(
                anio=2026, mes=mes_num,
                nombre=_s(row.get("MES")),
                abrev=_s(row.get("MESABREV")),
                ciclo_mes=_i(row.get("CICLOMES")),
                trimestre=trimestre_map.get(mes_num, 1),
                semestre=semestre_map.get(mes_num, 1),
            ))
            ins += 1
        db.commit()

    else:
        return 0, 0, 0, f"Hoja '{nombre_upper}' no reconocida — omitida"

    return ins, om, err, f"{ins} insertados, {om} ya existían, {err} errores"
