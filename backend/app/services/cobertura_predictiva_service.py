"""
SCGCPR — Motor de Cobertura Predictiva y Ritmo de Ejecución (4DX)
==================================================================
Replica exacta del motor de cálculo de la hoja `Motor_Formulas` del Excel
de referencia `Modulo_Cobertura_Predictiva_Ejemplo_VM_v3.xlsx` (adjuntado
por el usuario como especificación funcional/técnica del módulo).

Sustituye en el dashboard de GD las métricas de Comercial / Ventas, EVO IR
y cumplimiento de cuota (lag measures — resultados finales) por métricas
predictivas de cobertura médica y ritmo de visitas (lead measures —
indicadores que predicen el resultado y son accionables durante el ciclo),
aplicando la metodología 4DX (4 Disciplinas de la Ejecución).

Convenciones del Excel de referencia (confirmadas en hoja `Parametros`):
  - COBERTURA = médicos ÚNICOS visitados. Las visitas repetidas al mismo
    médico NO incrementan cobertura.
  - PSP / CONTACTOS = conteo TOTAL de visitas. Las visitas repetidas SÍ
    cuentan como contacto.

Mapeo de columnas del Motor_Formulas (letra de columna del Excel → cálculo
aquí, verificado extrayendo las fórmulas reales, no solo los valores
cacheados, de las filas VM01/VM02/VM03 del Excel adjunto):

    H  meta_cobertura            ParametroCobertura (cascada país/línea/ciclo)
    I  meta_contactos_ciclo      = J
    J  medicos_programados       = COUNT(TargetMedico) WHERE rm,ciclo,programado=True
    K  medicos_requeridos_meta   = CEILING(J × H, 1)
    L  medicos_unicos_visitados  = COUNT(DISTINCT Visita.medico_codigo) al corte
    M  contactos_realizados      = COUNT(Visita) al corte (cuenta repetidos)
    N  dias_habiles_ciclo         = Ciclo.dias_laborables, o NETWORKDAYS si no está configurado
    O  dias_habiles_transcurridos = NETWORKDAYS(inicio, MIN(corte, fin))
    P  dias_habiles_restantes     = MAX(0, N - O)
    Q  objetivo_medicos_diario    = IFERROR(K / N, 0)
    R  medicos_esperados_corte    = Q × O
    S  medicos_dejados_de_ver     = MAX(0, R - L)
    T  aceleracion_requerida_raw  = S / P  (si P>0, sino 0)
    U  medicos_diarios_requeridos = CEILING(Q + T, 1)  (si P>0, sino 0)
    V  objetivo_contactos_diario  = IFERROR(I / N, 0)
    W  contactos_esperados_corte  = V × O
    X  contactos_dejados_de_hacer = MAX(0, W - M)
    Y  contactos_diarios_requeridos = CEILING(V + X/P, 1)  (si P>0, sino 0)
    Z  cobertura_actual           = IFERROR(L / J, 0)
    AA cobertura_proyectada       = MIN(1, IFERROR((L/O × N) / J, 0))   ← run-rate
    AB estado_cobertura           = Verde si AA≥H; Amarillo si AA≥0.85 (piso fijo); si no Rojo
    AC estado_psp                 = Verde si proy≥I; Amarillo si proy≥I×0.9 (piso fijo); si no Rojo
    AD lectura_gd                 = mensaje accionable en español (ver _generar_lectura_gd)

NOTA IMPORTANTE sobre la columna L: en el Excel de ejemplo, L aparece como
un valor LITERAL fijo (70/90/85) en lugar de una fórmula COUNTIFS, porque
Excel nativo no puede expresar "conteo de valores distintos" con COUNTIFS.
Cruzando esto con la nota de negocio de la hoja `Parametros` ("la cobertura
usa médicos únicos... las visitas repetidas no incrementan cobertura"), la
regla de negocio REAL es `COUNT(DISTINCT medico_codigo)`, que es exactamente
lo que se implementa aquí contra DW.FACT_Visita — no se replica el atajo
de valor fijo del Excel.

Reglas de redondeo: todas las fórmulas CEILING(x, 1) de Excel redondean
hacia arriba al entero. Aquí se implementan con Decimal cuantizado para
evitar que el ruido de punto flotante (p. ej. 2.0000000000003) genere un
entero de más — ver `_ceiling1`.

Cálculo 100% en vivo (sin tabla de resultados materializada): a diferencia
del motor IUP/Ranking (que corre una vez por ciclo vía ETL y se guarda en
FACT_RankingRM), este módulo se recalcula en cada request del dashboard,
igual que el resto de los endpoints de `dashboard.py` — el "corte" es
siempre relativo a `fecha_corte` (por defecto HOY), por lo que no tiene
sentido materializarlo.
"""
from datetime import date, timedelta
from decimal import Decimal, ROUND_CEILING, ROUND_HALF_UP
from typing import Optional

from sqlalchemy import func
from sqlalchemy.orm import Session
from loguru import logger

from app.models.dimensiones import (
    RepresentanteMedico,
    Ciclo,
    TargetMedico,
    Feriado,
    ParametroCobertura,
)
from app.models.hechos import Visita


# Pisos fijos de semáforo (no parametrizables, igual que en el Excel de referencia)
_PISO_AMARILLO_COBERTURA = Decimal("0.85")
_PISO_AMARILLO_PSP = Decimal("0.90")
_META_COBERTURA_DEFECTO = Decimal("0.90")


# ─────────────────────────────────────────────────────────────────────────
# Helpers numéricos (replican el comportamiento de Excel)
# ─────────────────────────────────────────────────────────────────────────

def _safe_div(numerador: Decimal, denominador: Decimal) -> Decimal:
    """Replica IFERROR(numerador / denominador, 0)."""
    if denominador == 0:
        return Decimal("0")
    return numerador / denominador


def _ceiling1(x: Decimal) -> int:
    """
    Replica CEILING(x, 1) de Excel (redondeo hacia arriba al entero).
    Cuantiza antes de redondear para absorber ruido de punto flotante
    (evita que 2.000000001 se redondee a 3).
    """
    if x <= 0:
        return 0
    x = x.quantize(Decimal("0.000000001"), rounding=ROUND_HALF_UP)
    return int(x.to_integral_value(rounding=ROUND_CEILING))


def _networkdays(fecha_inicio: date, fecha_fin: date, feriados: set) -> int:
    """
    Replica NETWORKDAYS(fecha_inicio, fecha_fin, feriados) de Excel:
    cuenta días lunes-viernes entre ambas fechas (inclusive), excluyendo
    los feriados que caigan en día hábil.
    """
    if fecha_inicio > fecha_fin:
        return 0
    dias = 0
    actual = fecha_inicio
    un_dia = timedelta(days=1)
    while actual <= fecha_fin:
        if actual.weekday() < 5 and actual not in feriados:  # 0=lunes ... 6=domingo
            dias += 1
        actual += un_dia
    return dias


# ─────────────────────────────────────────────────────────────────────────
# Parametrización (meta de cobertura, feriados, días hábiles)
# ─────────────────────────────────────────────────────────────────────────

def obtener_meta_cobertura(
    db: Session,
    pais_codigo: str,
    linea_id: Optional[int] = None,
    ciclo_id: Optional[int] = None,
) -> Decimal:
    """
    Resuelve la meta de cobertura (columna H del Motor_Formulas) en cascada,
    de la más específica a la más general:
      1. (pais_codigo, linea_id, ciclo_id)
      2. (pais_codigo, linea_id, NULL)
      3. (pais_codigo, NULL, NULL)
      4. valor por defecto (0.90), si no hay ningún ParametroCobertura configurado
    """
    if linea_id is not None and ciclo_id is not None:
        p = db.query(ParametroCobertura).filter(
            ParametroCobertura.pais_codigo == pais_codigo,
            ParametroCobertura.linea_id == linea_id,
            ParametroCobertura.ciclo_id == ciclo_id,
            ParametroCobertura.activo == True,
        ).first()
        if p:
            return p.meta_cobertura

    if linea_id is not None:
        p = db.query(ParametroCobertura).filter(
            ParametroCobertura.pais_codigo == pais_codigo,
            ParametroCobertura.linea_id == linea_id,
            ParametroCobertura.ciclo_id.is_(None),
            ParametroCobertura.activo == True,
        ).first()
        if p:
            return p.meta_cobertura

    p = db.query(ParametroCobertura).filter(
        ParametroCobertura.pais_codigo == pais_codigo,
        ParametroCobertura.linea_id.is_(None),
        ParametroCobertura.ciclo_id.is_(None),
        ParametroCobertura.activo == True,
    ).first()
    if p:
        return p.meta_cobertura

    logger.debug(
        f"Sin ParametroCobertura configurado para país={pais_codigo}, línea={linea_id}, "
        f"ciclo={ciclo_id} — usando meta por defecto {_META_COBERTURA_DEFECTO}"
    )
    return _META_COBERTURA_DEFECTO


def _feriados_pais(db: Session, pais_codigo: str, fecha_inicio: date, fecha_fin: date) -> set:
    """Feriados activos del país dentro del rango del ciclo (Config.DIM_Feriado)."""
    filas = db.query(Feriado.fecha).filter(
        Feriado.pais_codigo == pais_codigo,
        Feriado.activo == True,
        Feriado.fecha >= fecha_inicio,
        Feriado.fecha <= fecha_fin,
    ).all()
    return {f[0] for f in filas}


def obtener_dias_habiles_ciclo(db: Session, ciclo: Ciclo) -> int:
    """
    Columna N del Motor_Formulas. Usa `Ciclo.dias_laborables` si está
    configurado (>0); si no, lo calcula con NETWORKDAYS usando
    Config.DIM_Feriado (reemplaza el rango estático
    Parametros!$B$17:$B$25 del Excel de referencia).
    """
    if ciclo.dias_laborables and ciclo.dias_laborables > 0:
        return ciclo.dias_laborables
    feriados = _feriados_pais(db, ciclo.pais_codigo, ciclo.fecha_inicio, ciclo.fecha_fin)
    return _networkdays(ciclo.fecha_inicio, ciclo.fecha_fin, feriados)


# ─────────────────────────────────────────────────────────────────────────
# Lectura GD (columna AD — mensaje accionable)
# ─────────────────────────────────────────────────────────────────────────

def _generar_lectura_gd(Y: int, V: Decimal, X: Decimal, P: int, U: int, aceleracion: int) -> str:
    """
    Replica la fórmula de texto concatenado de la columna AD del
    Motor_Formulas (hoja Dashboard GD, columna "Lectura GD" según la hoja
    Diccionario). Convierte el KPI numérico en una instrucción operativa
    para el GD/VM.
    """
    return (
        f"Su número de contactos diario requerido es {Y}. Este número es la suma del "
        f"objetivo diario base ({float(V):.1f}) + los contactos dejados de hacer "
        f"({float(X):.1f}) repartidos en {P} días hábiles restantes. Para cobertura "
        f"debe ver {U} médicos diarios; aceleración requerida: +{aceleracion} "
        f"médicos adicionales/día."
    )


# ─────────────────────────────────────────────────────────────────────────
# Cálculo principal por RM (una fila del Motor_Formulas)
# ─────────────────────────────────────────────────────────────────────────

def calcular_cobertura_rm(
    db: Session,
    rm_id: int,
    ciclo_id: int,
    fecha_corte: Optional[date] = None,
) -> dict:
    """
    Calcula, EN VIVO, todas las columnas del Motor_Formulas (I..AD) para un
    RM/VM en un ciclo dado, al corte indicado (por defecto HOY).

    Lanza ValueError si el RM o el ciclo no existen.
    """
    rm = db.query(RepresentanteMedico).filter(RepresentanteMedico.id == rm_id).first()
    if not rm:
        raise ValueError(f"RM ID={rm_id} no encontrado")

    ciclo = db.query(Ciclo).filter(Ciclo.id == ciclo_id).first()
    if not ciclo:
        raise ValueError(f"Ciclo ID={ciclo_id} no encontrado")

    if fecha_corte is None:
        fecha_corte = date.today()

    meta_cobertura = obtener_meta_cobertura(db, rm.pais_codigo, rm.linea_id, ciclo_id)  # H
    feriados = _feriados_pais(db, rm.pais_codigo, ciclo.fecha_inicio, ciclo.fecha_fin)

    # N — días hábiles totales del ciclo
    if ciclo.dias_laborables and ciclo.dias_laborables > 0:
        N = ciclo.dias_laborables
    else:
        N = _networkdays(ciclo.fecha_inicio, ciclo.fecha_fin, feriados)

    # O — días hábiles transcurridos hasta fecha_corte (acotado a [inicio, fin] del ciclo)
    if fecha_corte < ciclo.fecha_inicio:
        O = 0
    else:
        fin_parcial = min(fecha_corte, ciclo.fecha_fin)
        O = _networkdays(ciclo.fecha_inicio, fin_parcial, feriados)

    P = max(0, N - O)  # días hábiles restantes

    # J — médicos programados (universo target del RM en el ciclo)
    J = db.query(func.count(TargetMedico.id)).filter(
        TargetMedico.rm_id == rm_id,
        TargetMedico.ciclo_id == ciclo_id,
        TargetMedico.programado == True,
        TargetMedico.activo == True,
    ).scalar() or 0

    I = J  # meta de contactos del ciclo = médicos programados

    K = _ceiling1(Decimal(J) * meta_cobertura)  # médicos requeridos para cumplir meta

    # L — médicos ÚNICOS visitados al corte (COUNT DISTINCT real, no el atajo del Excel)
    L = db.query(func.count(func.distinct(Visita.medico_codigo))).filter(
        Visita.rm_id == rm_id,
        Visita.ciclo_id == ciclo_id,
        Visita.estado_visita == "Realizada",
        Visita.fecha_visita <= fecha_corte,
    ).scalar() or 0

    # M — contactos TOTALES realizados al corte (cuenta repetidos)
    M = db.query(func.count(Visita.id)).filter(
        Visita.rm_id == rm_id,
        Visita.ciclo_id == ciclo_id,
        Visita.estado_visita == "Realizada",
        Visita.fecha_visita <= fecha_corte,
    ).scalar() or 0

    Q = _safe_div(Decimal(K), Decimal(N))             # objetivo médicos/día (ritmo de cobertura)
    R = Q * Decimal(O)                                 # médicos esperados al corte
    S = max(Decimal("0"), R - Decimal(L))              # médicos dejados de ver
    T = (S / Decimal(P)) if P > 0 else Decimal("0")    # aceleración requerida (médicos/día extra)
    U = _ceiling1(Q + T) if P > 0 else 0               # médicos diarios requeridos (base + recuperación)

    V = _safe_div(Decimal(I), Decimal(N))              # objetivo contactos/día base
    W = V * Decimal(O)                                 # contactos esperados al corte
    X = max(Decimal("0"), W - Decimal(M))              # contactos dejados de hacer
    Y = _ceiling1(V + _safe_div(X, Decimal(P))) if P > 0 else 0  # contactos diarios requeridos

    Z = _safe_div(Decimal(L), Decimal(J))              # cobertura actual

    if O > 0:
        AA = min(Decimal("1"), _safe_div(_safe_div(Decimal(L), Decimal(O)) * Decimal(N), Decimal(J)))
        psp_proyectado = _safe_div(Decimal(M), Decimal(O)) * Decimal(N)
    else:
        AA = Decimal("0")
        psp_proyectado = Decimal("0")

    if AA >= meta_cobertura:
        AB = "Verde"
    elif AA >= _PISO_AMARILLO_COBERTURA:
        AB = "Amarillo"
    else:
        AB = "Rojo"

    if psp_proyectado >= Decimal(I):
        AC = "Verde"
    elif psp_proyectado >= Decimal(I) * _PISO_AMARILLO_PSP:
        AC = "Amarillo"
    else:
        AC = "Rojo"

    aceleracion_redondeada = _ceiling1(T)  # solo para la narrativa de la columna AD
    lectura_gd = _generar_lectura_gd(Y=Y, V=V, X=X, P=P, U=U, aceleracion=aceleracion_redondeada)

    logger.debug(
        f"Cobertura predictiva RM={rm.codigo} ciclo={ciclo_id} corte={fecha_corte}: "
        f"cobertura_actual={Z:.2%} proyectada={AA:.2%} estado={AB}/{AC}"
    )

    return {
        "rm_id": rm_id,
        "rm_codigo": rm.codigo,
        "rm_nombre": rm.nombre,
        "ciclo_id": ciclo_id,
        "ciclo_nombre": ciclo.nombre,
        "fecha_corte": fecha_corte.isoformat(),
        "meta_cobertura": float(meta_cobertura),
        "medicos_programados": J,
        "medicos_requeridos_meta": K,
        "medicos_unicos_visitados": L,
        "contactos_realizados": M,
        "dias_habiles_ciclo": N,
        "dias_habiles_transcurridos": O,
        "dias_habiles_restantes": P,
        "objetivo_medicos_diario": float(Q),
        "medicos_esperados_corte": float(R),
        "medicos_dejados_de_ver": float(S),
        "aceleracion_requerida": float(T),
        "aceleracion_requerida_redondeada": aceleracion_redondeada,
        "medicos_diarios_requeridos": U,
        "meta_contactos_ciclo": I,
        "objetivo_contactos_diario": float(V),
        "contactos_esperados_corte": float(W),
        "contactos_dejados_de_hacer": float(X),
        "contactos_diarios_requeridos": Y,
        "cobertura_actual_pct": float(Z * 100),
        "cobertura_proyectada_pct": float(AA * 100),
        "estado_cobertura": AB,
        "estado_psp": AC,
        "lectura_gd": lectura_gd,
    }


# ─────────────────────────────────────────────────────────────────────────
# Resumen de equipo (dashboard GD/Ejecutivo) — sustituye a dashboard_comercial
# ─────────────────────────────────────────────────────────────────────────

def calcular_cobertura_equipo(
    db: Session,
    ciclo_id: int,
    pais_codigo: Optional[str] = None,
    gerente_id: Optional[int] = None,
    linea_id: Optional[int] = None,
    fecha_corte: Optional[date] = None,
) -> dict:
    """
    Agrega `calcular_cobertura_rm` sobre todos los RMs en alcance (con
    universo Target_Medicos cargado para el ciclo), filtrando opcionalmente
    por país/gerente(GD)/línea. Reemplaza el contenido de
    `GET /dashboard/comercial` (ventas/EVO IR/cuota) con las métricas
    predictivas de cobertura y ritmo de ejecución.
    """
    if fecha_corte is None:
        fecha_corte = date.today()

    q = db.query(RepresentanteMedico.id).join(
        TargetMedico, TargetMedico.rm_id == RepresentanteMedico.id
    ).filter(
        TargetMedico.ciclo_id == ciclo_id,
        RepresentanteMedico.activo == True,
    )
    if pais_codigo:
        q = q.filter(RepresentanteMedico.pais_codigo == pais_codigo)
    if gerente_id:
        q = q.filter(RepresentanteMedico.gerente_id == gerente_id)
    if linea_id:
        q = q.filter(RepresentanteMedico.linea_id == linea_id)

    rm_ids = [r[0] for r in q.distinct().all()]
    detalles = [calcular_cobertura_rm(db, rm_id, ciclo_id, fecha_corte) for rm_id in rm_ids]

    total = len(detalles)
    divisor = total or 1

    verdes = sum(1 for d in detalles if d["estado_cobertura"] == "Verde")
    amarillos = sum(1 for d in detalles if d["estado_cobertura"] == "Amarillo")
    rojos = sum(1 for d in detalles if d["estado_cobertura"] == "Rojo")

    logger.info(
        f"Resumen cobertura predictiva: ciclo={ciclo_id} país={pais_codigo} gerente={gerente_id} "
        f"corte={fecha_corte} → {total} RMs (Verde={verdes}, Amarillo={amarillos}, Rojo={rojos})"
    )

    return {
        "ciclo_id": ciclo_id,
        "fecha_corte": fecha_corte.isoformat(),
        "total_rms": total,
        "cobertura_actual_promedio_pct": round(sum(d["cobertura_actual_pct"] for d in detalles) / divisor, 2),
        "cobertura_proyectada_promedio_pct": round(sum(d["cobertura_proyectada_pct"] for d in detalles) / divisor, 2),
        "medicos_programados_total": sum(d["medicos_programados"] for d in detalles),
        "medicos_unicos_visitados_total": sum(d["medicos_unicos_visitados"] for d in detalles),
        "contactos_realizados_total": sum(d["contactos_realizados"] for d in detalles),
        "meta_contactos_ciclo_total": sum(d["meta_contactos_ciclo"] for d in detalles),
        "rms_en_verde": verdes,
        "rms_en_amarillo": amarillos,
        "rms_en_rojo": rojos,
        "pct_rms_en_riesgo": round((amarillos + rojos) / divisor * 100, 2),
        "detalle_rms": detalles,
        "filtros": {
            "pais_codigo": pais_codigo,
            "ciclo_id": ciclo_id,
            "gerente_id": gerente_id,
            "linea_id": linea_id,
        },
    }


# ═══════════════════════════════════════════════════════════════════════════════
# CAPA cat.* — Funciones que usan el esquema cat.* (SP + vista + DimCiclo)
# ═══════════════════════════════════════════════════════════════════════════════

from sqlalchemy import text as _text
from app.core.config import settings as _settings
from app.models.cat_models import (
    CatDimPais, CatDimCiclo, CatDimCalendario,
    CatFactTargetMedicoCiclo, CatFactVisitaMedica,
    CatKpiCoberturaPredictiva,
    CatDimRepresentante as CatDimRepresentanteMedico,
    CatDimMedico,
    CatDimEspecialidad,
)


def get_ciclos_cat(db: Session, pais_codigo: Optional[str] = None) -> list:
    """Lista ciclos desde cat.DimCiclo para los selectores del frontend."""
    q = db.query(
        CatDimCiclo.ciclo_key,
        CatDimCiclo.codigo_ciclo,
        CatDimCiclo.linea,
        CatDimCiclo.fecha_inicio,
        CatDimCiclo.fecha_fin,
        CatDimCiclo.dias_habiles_ciclo,
        CatDimCiclo.meta_cobertura_pct,
        CatDimCiclo.meta_contactos_ciclo,
        CatDimCiclo.activo,
        CatDimPais.codigo_pais,
        CatDimPais.nombre_pais,
    ).join(CatDimPais, CatDimPais.pais_key == CatDimCiclo.pais_key).filter(
        CatDimCiclo.activo == True
    )
    if pais_codigo:
        q = q.filter(CatDimPais.codigo_pais == pais_codigo.upper())
    rows = q.order_by(CatDimCiclo.fecha_inicio.desc()).all()
    return [
        {
            "ciclo_key": r.ciclo_key,
            "codigo_ciclo": r.codigo_ciclo,
            "linea": r.linea,
            "fecha_inicio": r.fecha_inicio.isoformat() if r.fecha_inicio else None,
            "fecha_fin": r.fecha_fin.isoformat() if r.fecha_fin else None,
            "dias_habiles_ciclo": r.dias_habiles_ciclo,
            "meta_cobertura_pct": float(r.meta_cobertura_pct) * 100,
            "meta_contactos_ciclo": r.meta_contactos_ciclo,
            "activo": r.activo,
            "pais_codigo": r.codigo_pais,
            "pais_nombre": r.nombre_pais,
        }
        for r in rows
    ]


def _dias_habiles(f_ini: date, f_fin: date) -> int:
    """Cuenta días L-V (Mon-Fri) en [f_ini, f_fin] inclusive — fallback sin DimCalendario."""
    n, d = 0, f_ini
    while d <= f_fin:
        if d.weekday() < 5:
            n += 1
        d += timedelta(days=1)
    return n


def calcular_cobertura_py(db: Session, ciclo_codigo: str, pais_codigo: str,
                          fecha_corte: date, representante_key=None, linea=None) -> dict:
    """Equivalente Python de cat.sp_CalcularCoberturaPredictiva. Calcula los KPIs de
    cobertura por representante y los inserta en cat.KpiCoberturaPredictiva."""
    import math
    from datetime import datetime, timezone

    pais_key = db.execute(_text('SELECT "PaisKey" FROM "cat"."DimPais" WHERE "CodigoPais"=:p'),
                          {"p": pais_codigo}).scalar()
    if pais_key is None:
        raise ValueError(f"País no encontrado: {pais_codigo}")
    c = db.execute(_text(
        'SELECT "CicloKey", "FechaInicio", "FechaFin", "MetaCoberturaPct", COALESCE("MetaContactosCiclo",0) AS "MetaContactosCiclo" '
        'FROM "cat"."DimCiclo" WHERE "CodigoCiclo"=:cc AND "PaisKey"=:pk AND "Activo"=1 '
        'AND (:linea IS NULL OR "Linea"=:linea) ORDER BY "CicloKey" OFFSET 0 ROWS FETCH FIRST 1 ROWS ONLY'),
        {"cc": ciclo_codigo, "pk": pais_key, "linea": linea}).mappings().first()
    if c is None:
        raise ValueError(f"Ciclo no encontrado: {ciclo_codigo} / País: {pais_codigo}")
    ciclo_key, f_ini, f_fin = c["CicloKey"], c["FechaInicio"], c["FechaFin"]
    meta = Decimal(str(c["MetaCoberturaPct"]))
    fce = f_fin if fecha_corte > f_fin else fecha_corte

    # Días hábiles (DimCalendario, o fallback L-V)
    total = db.execute(_text('SELECT COUNT(*) FROM "cat"."DimCalendario" WHERE "PaisKey"=:pk AND "CicloKey"=:ck AND "EsHabil"=1'),
                       {"pk": pais_key, "ck": ciclo_key}).scalar() or 0
    if total == 0:
        total = _dias_habiles(f_ini, f_fin)
    transc = db.execute(_text('SELECT COUNT(*) FROM "cat"."DimCalendario" WHERE "PaisKey"=:pk AND "CicloKey"=:ck '
                              'AND "EsHabil"=1 AND "Fecha"<=:fce AND "Fecha">=:fi'),
                        {"pk": pais_key, "ck": ciclo_key, "fce": fce, "fi": f_ini}).scalar() or 0
    if transc == 0 and fce >= f_ini:
        transc = _dias_habiles(f_ini, fce)
        if fce <= f_fin and fce.weekday() < 5:
            transc += 1
    restantes = total - transc if total > transc else 0

    db.execute(_text('DELETE FROM "cat"."KpiCoberturaPredictiva" WHERE "CicloKey"=:ck AND "FechaCorte"=:fc '
                     'AND (:rk IS NULL OR "RepresentanteKey"=:rk)'),
               {"ck": ciclo_key, "fc": fecha_corte, "rk": representante_key})

    targets = {r["RepresentanteKey"]: r["MedicosProg"] for r in db.execute(_text(
        'SELECT "RepresentanteKey", COUNT(*) AS "MedicosProg" FROM "cat"."FactTargetMedicoCiclo" '
        'WHERE "CicloKey"=:ck AND "PaisKey"=:pk AND "ProgramadoFlag"=1 AND (:rk IS NULL OR "RepresentanteKey"=:rk) '
        'GROUP BY "RepresentanteKey"'), {"ck": ciclo_key, "pk": pais_key, "rk": representante_key}).mappings().all()}
    visitas = {r["RepresentanteKey"]: (r["MedicosUnicos"], r["ContactosTotales"]) for r in db.execute(_text(
        'SELECT "RepresentanteKey", COUNT(DISTINCT "CodigoMedicoOrigen") AS "MedicosUnicos", COUNT(*) AS "ContactosTotales" '
        'FROM "cat"."FactVisitaMedica" WHERE "CicloKey"=:ck AND "PaisKey"=:pk AND "EstadoVisita"=\'Realizada\' '
        'AND "FechaVisita"<=:fc AND (:rk IS NULL OR "RepresentanteKey"=:rk) GROUP BY "RepresentanteKey"'),
        {"ck": ciclo_key, "pk": pais_key, "fc": fecha_corte, "rk": representante_key}).mappings().all()}
    reps = db.execute(_text('SELECT "RepresentanteKey", "EquipoTexto", "NombreRepresentante" FROM "cat"."DimRepresentanteMedico" '
                            'WHERE "Activo"=1 AND (:linea IS NULL OR "EquipoTexto"=:linea) ORDER BY "RepresentanteKey"'),
                      {"linea": linea}).mappings().all()

    ahora = datetime.now(timezone.utc)
    filas = []
    for r in reps:
        rk = r["RepresentanteKey"]
        prog = targets.get(rk)
        if prog is None:  # INNER JOIN: solo reps con target
            continue
        unicos, contactos = visitas.get(rk, (0, 0))
        prog_d, unicos_d, cont_d = Decimal(prog), Decimal(unicos), Decimal(contactos)

        cob_actual = (unicos_d / prog_d) if prog > 0 else Decimal(0)
        cob_esperada = (meta * Decimal(transc) / total) if total > 0 else Decimal(0)
        if prog == 0 or transc == 0:
            proj = Decimal(0)
        else:
            raw = (unicos_d / transc * total) / prog_d
            proj = Decimal(1) if raw > 1 else raw
        brecha_actual = (cob_actual - cob_esperada) if (prog > 0 and total > 0) else Decimal(0)
        brecha_proj = proj - meta
        req = math.ceil(prog_d * meta) if prog > 0 else 0
        pend = max(0, req - unicos) if prog > 0 else 0
        med_diarios = (math.ceil(Decimal(req - unicos) / restantes) if (restantes and prog > 0 and req > unicos) else 0)
        cumpl_cont = (cont_d / prog_d) if prog > 0 else Decimal(0)
        cont_proj = (cont_d / transc * total) if transc > 0 else Decimal(0)
        cont_pend = (prog_d - cont_d) if prog > contactos else Decimal(0)
        cont_diarios = (math.ceil(prog_d / total + (cont_pend / restantes if restantes else Decimal(0)))
                        if (restantes and total) else 0)

        def _estado_cob():
            if prog == 0 or transc == 0:
                return "Rojo"
            if proj >= meta:
                return "Verde"
            if proj >= Decimal("0.85"):
                return "Amarillo"
            return "Rojo"

        def _estado_ritmo():
            if prog == 0 or total == 0:
                return "Rojo"
            umbral = meta * Decimal(transc) / total
            if cob_actual >= umbral:
                return "Verde"
            if cob_actual >= umbral - Decimal("0.05"):
                return "Amarillo"
            return "Rojo"

        def _estado_psp():
            if prog == 0 or transc == 0:
                return "Rojo"
            if cont_proj >= prog_d:
                return "Verde"
            if cont_proj >= prog_d * Decimal("0.9"):
                return "Amarillo"
            return "Rojo"

        lectura = (f"Médicos diarios requeridos: {med_diarios}. Contactos/día requeridos: {cont_diarios}. "
                   f"Días hábiles restantes: {restantes}. Médicos pendientes: {pend} de {prog} programados.")
        filas.append({
            "FechaCorte": fecha_corte, "CicloKey": ciclo_key, "PaisKey": pais_key, "Linea": r["EquipoTexto"],
            "GD": None, "RepresentanteKey": rk, "NombreVM": r["NombreRepresentante"],
            "MedicosProgramados": prog, "MedicosVisitadosUnicos": unicos,
            "CoberturaActualPct": cob_actual, "CoberturaEsperadaPct": cob_esperada,
            "CoberturaProyectadaPct": proj, "MetaCoberturaPct": meta,
            "BrechaActualVsEsperada": brecha_actual, "BrechaProyectadaVsMeta": brecha_proj,
            "MedicosRequeridosMeta": req, "MedicosPendientesMeta": pend, "MedicosDiariosRequeridos": med_diarios,
            "ContactosMetaCiclo": prog, "ContactosRealizados": contactos, "CumplimientoContactosPct": cumpl_cont,
            "ContactosProyectados": cont_proj, "ContactosPendientes": cont_pend, "ContactosDiariosRequeridos": cont_diarios,
            "DiasHabilesTotales": total, "DiasHabilesTranscurridos": transc, "DiasHabilesRestantes": restantes,
            "EstadoCobertura": _estado_cob(), "EstadoRitmo": _estado_ritmo(), "EstadoPSP": _estado_psp(),
            "LecturaAccionable": lectura, "FechaCargaUtc": ahora})

    if filas:
        db.execute(_text("""INSERT INTO "cat"."KpiCoberturaPredictiva"
            ("FechaCorte","CicloKey","PaisKey","Linea","GD","RepresentanteKey","NombreVM","MedicosProgramados","MedicosVisitadosUnicos",
             "CoberturaActualPct","CoberturaEsperadaPct","CoberturaProyectadaPct","MetaCoberturaPct","BrechaActualVsEsperada","BrechaProyectadaVsMeta",
             "MedicosRequeridosMeta","MedicosPendientesMeta","MedicosDiariosRequeridos","ContactosMetaCiclo","ContactosRealizados","CumplimientoContactosPct",
             "ContactosProyectados","ContactosPendientes","ContactosDiariosRequeridos","DiasHabilesTotales","DiasHabilesTranscurridos","DiasHabilesRestantes",
             "EstadoCobertura","EstadoRitmo","EstadoPSP","LecturaAccionable","FechaCargaUtc")
            VALUES (:FechaCorte,:CicloKey,:PaisKey,:Linea,:GD,:RepresentanteKey,:NombreVM,:MedicosProgramados,:MedicosVisitadosUnicos,
             :CoberturaActualPct,:CoberturaEsperadaPct,:CoberturaProyectadaPct,:MetaCoberturaPct,:BrechaActualVsEsperada,:BrechaProyectadaVsMeta,
             :MedicosRequeridosMeta,:MedicosPendientesMeta,:MedicosDiariosRequeridos,:ContactosMetaCiclo,:ContactosRealizados,:CumplimientoContactosPct,
             :ContactosProyectados,:ContactosPendientes,:ContactosDiariosRequeridos,:DiasHabilesTotales,:DiasHabilesTranscurridos,:DiasHabilesRestantes,
             :EstadoCobertura,:EstadoRitmo,:EstadoPSP,:LecturaAccionable,:FechaCargaUtc)"""), filas)
    return {"FilasInsertadas": len(filas), "CicloKey": ciclo_key, "FechaCorte": fecha_corte,
            "DiasHabilesTotales": total, "DiasHabilesTranscurridos": transc, "DiasHabilesRestantes": restantes}


def calcular_cobertura_sp(
    db: Session,
    ciclo_codigo: str,
    pais_codigo: str,
    fecha_corte: Optional[date] = None,
    representante_key: Optional[int] = None,
    linea: Optional[str] = None,
) -> dict:
    """(Re)calcula los KPIs de cobertura predictiva en cat.KpiCoberturaPredictiva.

    jul-2026: el cálculo se movió de cat.sp_CalcularCoberturaPredictiva (T-SQL) a
    Python (calcular_cobertura_py) para dejar el core agnóstico de BD."""
    if fecha_corte is None:
        fecha_corte = date.today()
    try:
        result = calcular_cobertura_py(db, ciclo_codigo, pais_codigo.upper(), fecha_corte,
                                       representante_key, linea)
        db.commit()
        return {
            "ok": True,
            "filas_insertadas": result["FilasInsertadas"],
            "ciclo_key": result["CicloKey"],
            "fecha_corte": str(fecha_corte),
            "dias_habiles_totales": result["DiasHabilesTotales"],
            "dias_habiles_transcurridos": result["DiasHabilesTranscurridos"],
            "dias_habiles_restantes": result["DiasHabilesRestantes"],
        }
    except Exception as exc:
        db.rollback()
        logger.error(f"[calcular_cobertura_py] Error: {exc}")
        raise


def get_cobertura_por_categoria(
    db: Session,
    ciclo_codigo: str,
    pais_codigo: str,
    representante_codigo: Optional[str] = None,
    gd: Optional[str] = None,
) -> list:
    """
    Devuelve el avance de visitas por categoría de médico (A/B/C/D).
    Cruza FactTargetMedicoCiclo (programados) con FactVisitaMedica (realizadas).
    Resultado: lista ordenada A→D con total_programados, visitados, pendientes, pct.
    """
    from sqlalchemy import text as _t

    # Resolver claves
    ciclo_row = db.execute(
        _t("""
            SELECT c."CicloKey", p."PaisKey"
            FROM "cat"."DimCiclo" c
            JOIN "cat"."DimPais" p ON p."PaisKey" = c."PaisKey"
            WHERE c."CodigoCiclo" = :cc AND p."CodigoPais" = :pc
            ORDER BY c."CicloKey" DESC
        """),
        {"cc": ciclo_codigo, "pc": pais_codigo},
    ).mappings().first()

    if not ciclo_row:
        return []

    ciclo_key = ciclo_row["CicloKey"]
    pais_key  = ciclo_row["PaisKey"]

    # Filtro opcional por representante
    rep_filter = ""
    params: dict = {"ciclo_key": ciclo_key, "pais_key": pais_key}

    if representante_codigo:
        rep_filter += ' AND r."CodigoRepresentante" = :rep_cod'
        params["rep_cod"] = representante_codigo.upper()
    if gd:
        rep_filter += ' AND r."EquipoTexto" = :gd'
        params["gd"] = gd

    sql = f"""
        SELECT
            t."Potencial"                                       AS "Categoria",
            COUNT(DISTINCT t."CodigoMedicoOrigen")             AS "TotalProgramados",
            COUNT(DISTINCT v."CodigoMedicoOrigen")             AS "Visitados",
            COUNT(DISTINCT t."CodigoMedicoOrigen")
              - COUNT(DISTINCT v."CodigoMedicoOrigen")         AS "Pendientes"
        FROM "cat"."FactTargetMedicoCiclo" t
        JOIN "cat"."DimRepresentanteMedico" r
          ON r."RepresentanteKey" = t."RepresentanteKey"
        LEFT JOIN "cat"."FactVisitaMedica" v
          ON  v."CodigoMedicoOrigen"  = t."CodigoMedicoOrigen"
          AND v."CicloKey"            = t."CicloKey"
          AND v."PaisKey"             = t."PaisKey"
          AND v."RepresentanteKey"    = t."RepresentanteKey"
          AND v."EstadoVisita"        = 'Realizada'
        WHERE t."CicloKey"       = :ciclo_key
          AND t."PaisKey"        = :pais_key
          AND t."ProgramadoFlag" = 1
          AND t."Potencial"      IS NOT NULL
          AND t."Potencial"      <> ''
          {rep_filter}
        GROUP BY t."Potencial"
        ORDER BY t."Potencial"
    """

    rows = db.execute(_t(sql), params).mappings().all()
    result = []
    for r in rows:
        total = r["TotalProgramados"] or 0
        visitados = r["Visitados"] or 0
        result.append({
            "categoria": r["Categoria"],
            "total_programados": total,
            "visitados": visitados,
            "pendientes": r["Pendientes"] or 0,
            "pct_visita": round(visitados / total * 100, 1) if total > 0 else 0.0,
        })
    return result


def get_dashboard_cat(
    db: Session,
    ciclo_codigo: str,
    pais_codigo: str,
    fecha_corte: Optional[date] = None,
    linea: Optional[str] = None,
    gd: Optional[str] = None,
    representante: Optional[str] = None,
) -> dict:
    """
    Lee de cat.vwDashboardCoberturaPredictivaGD y devuelve el resumen
    del equipo + detalle por VM. Compatible con el contrato del frontend.
    """
    if fecha_corte is None:
        fecha_corte = date.today()

    # Resolver ciclo_key + pais_key más reciente para este ciclo_codigo
    ciclo_row = db.execute(
        _text("""
            SELECT c."CicloKey", p."PaisKey", c."CodigoCiclo",
                         c."FechaInicio", c."FechaFin", c."MetaCoberturaPct",
                         c."DiasHabilesCiclo"
            FROM "cat"."DimCiclo" c
            INNER JOIN "cat"."DimPais" p ON p."PaisKey" = c."PaisKey"
            WHERE c."CodigoCiclo" = :ciclo_codigo AND p."CodigoPais" = :pais_codigo
              AND c."Activo" = 1
            ORDER BY c."CicloKey"
        """),
        {"ciclo_codigo": ciclo_codigo, "pais_codigo": pais_codigo.upper()},
    ).mappings().first()

    if not ciclo_row:
        return {
            "ok": False,
            "mensaje": f"Ciclo '{ciclo_codigo}' no encontrado para país '{pais_codigo}'",
            "total_rms": 0,
            "detalle_rms": [],
        }

    # Buscar la fecha de corte más próxima (el SP puede haberse corrido en otra fecha).
    # La diferencia de días es dialect-específica: SQL Server usa DATEDIFF(DAY, a, b);
    # Postgres resta directamente dos DATE y devuelve un entero de días.
    _diff_dias = (
        'ABS("FechaCorte"::date - CAST(:fecha_corte AS date))'
        if _settings.DB_ENGINE == "postgres"
        else 'ABS(DATEDIFF(DAY, "FechaCorte", :fecha_corte))'
    )
    kpi_fecha = db.execute(
        _text(f"""
            SELECT "FechaCorte"
            FROM "cat"."KpiCoberturaPredictiva"
            WHERE "CicloKey" = :ciclo_key
            ORDER BY {_diff_dias}
            OFFSET 0 ROWS FETCH FIRST 1 ROWS ONLY
        """),
        {"ciclo_key": ciclo_row["CicloKey"], "fecha_corte": fecha_corte},
    ).scalar()

    if not kpi_fecha:
        return {
            "ok": False,
            "mensaje": "Sin datos calculados. Ejecute 'Calcular KPIs' primero.",
            "total_rms": 0,
            "detalle_rms": [],
            "ciclo": {
                "codigo": ciclo_row["CodigoCiclo"],
                "fecha_inicio": str(ciclo_row["FechaInicio"]),
                "fecha_fin": str(ciclo_row["FechaFin"]),
            },
        }

    # Leer desde la vista
    filters = [
        '"CodigoCiclo" = :ciclo_codigo',
        '"FechaCorte" = :fecha_corte_ef',
    ]
    params: dict = {"ciclo_codigo": ciclo_row["CodigoCiclo"], "fecha_corte_ef": kpi_fecha}

    if linea:
        filters.append('"Linea" = :linea')
        params["linea"] = linea
    if gd:
        filters.append('"GD" = :gd')
        params["gd"] = gd
    if representante:
        filters.append('("NombreRepresentante" LIKE :rep OR "CodigoRepresentante" = :rep)')
        params["rep"] = f"%{representante}%"

    where = " AND ".join(filters)

    rows = db.execute(
        _text(f"""
            SELECT * FROM "cat"."vwDashboardCoberturaPredictivaGD"
            WHERE {where}
            ORDER BY "EstadoCobertura", "CoberturaActualPct"
        """),
        params,
    ).mappings().all()

    detalle = []
    for r in rows:
        detalle.append({
            "kpi_key": r["KpiKey"],
            "codigo_representante": r["CodigoRepresentante"],
            "nombre_vm": r["NombreRepresentante"],
            "linea": r["Linea"],
            "gd": r["GD"],
            "medicos_programados": r["MedicosProgramados"],
            "medicos_visitados_unicos": r["MedicosVisitadosUnicos"],
            "cobertura_actual_pct": float(r["CoberturaActualPct"] or 0),
            "cobertura_esperada_pct": float(r["CoberturaEsperadaPct"] or 0),
            "cobertura_proyectada_pct": float(r["CoberturaProyectadaPct"] or 0),
            "meta_cobertura_pct": float(r["MetaCoberturaPct"] or 0),
            "brecha_actual_vs_esperada_pct": float(r["BrechaActualVsEsperadaPct"] or 0),
            "brecha_proyectada_vs_meta_pct": float(r["BrechaProyectadaVsMetaPct"] or 0),
            "medicos_requeridos_meta": r["MedicosRequeridosMeta"],
            "medicos_pendientes_meta": r["MedicosPendientesMeta"],
            "medicos_diarios_requeridos": r["MedicosDiariosRequeridos"],
            "contactos_meta_ciclo": r["ContactosMetaCiclo"],
            "contactos_realizados": r["ContactosRealizados"],
            "cumplimiento_contactos_pct": float(r["CumplimientoContactosPct"] or 0),
            "contactos_proyectados": float(r["ContactosProyectados"] or 0),
            "contactos_pendientes": float(r["ContactosPendientes"] or 0),
            "contactos_diarios_requeridos": r["ContactosDiariosRequeridos"],
            "dias_habiles_totales": r["DiasHabilesTotales"],
            "dias_habiles_transcurridos": r["DiasHabilesTranscurridos"],
            "dias_habiles_restantes": r["DiasHabilesRestantes"],
            "estado_cobertura": r["EstadoCobertura"],
            "estado_ritmo": r["EstadoRitmo"],
            "estado_psp": r["EstadoPSP"],
            "lectura_accionable": r["LecturaAccionable"],
        })

    total = len(detalle)
    divisor = total or 1

    def _avg(field: str) -> float:
        return round(sum(d[field] for d in detalle) / divisor, 2) if detalle else 0.0

    verdes = sum(1 for d in detalle if d["estado_cobertura"] == "Verde")
    amarillos = sum(1 for d in detalle if d["estado_cobertura"] == "Amarillo")
    rojos = sum(1 for d in detalle if d["estado_cobertura"] == "Rojo")

    verdes_psp = sum(1 for d in detalle if d["estado_psp"] == "Verde")
    rojos_psp = sum(1 for d in detalle if d["estado_psp"] == "Rojo")

    return {
        "ok": True,
        "fuente": "cat.KpiCoberturaPredictiva",
        "ciclo": {
            "codigo": ciclo_row["CodigoCiclo"],
            "fecha_inicio": str(ciclo_row["FechaInicio"]),
            "fecha_fin": str(ciclo_row["FechaFin"]),
            "meta_cobertura_pct": float(ciclo_row["MetaCoberturaPct"]) * 100,
        },
        "fecha_corte": str(kpi_fecha),
        "pais_codigo": pais_codigo.upper(),
        "total_rms": total,
        "cobertura_actual_promedio_pct": _avg("cobertura_actual_pct"),
        "cobertura_esperada_promedio_pct": _avg("cobertura_esperada_pct"),
        "cobertura_proyectada_promedio_pct": _avg("cobertura_proyectada_pct"),
        "meta_cobertura_pct": _avg("meta_cobertura_pct"),
        "medicos_programados_total": sum(d["medicos_programados"] for d in detalle),
        "medicos_unicos_visitados_total": sum(d["medicos_visitados_unicos"] for d in detalle),
        "contactos_realizados_total": sum(d["contactos_realizados"] for d in detalle),
        "contactos_meta_total": sum(d["contactos_meta_ciclo"] for d in detalle),
        "medicos_diarios_requeridos_max": max((d["medicos_diarios_requeridos"] for d in detalle), default=0),
        "rms_en_verde": verdes,
        "rms_en_amarillo": amarillos,
        "rms_en_rojo": rojos,
        "rms_psp_verde": verdes_psp,
        "rms_psp_rojo": rojos_psp,
        "pct_rms_en_riesgo": round((amarillos + rojos) / divisor * 100, 2),
        "dias_habiles_totales": detalle[0]["dias_habiles_totales"] if detalle else 0,
        "dias_habiles_transcurridos": detalle[0]["dias_habiles_transcurridos"] if detalle else 0,
        "dias_habiles_restantes": detalle[0]["dias_habiles_restantes"] if detalle else 0,
        "detalle_rms": detalle,
        "filtros": {
            "ciclo_codigo": ciclo_codigo,
            "pais_codigo": pais_codigo,
            "fecha_corte": str(fecha_corte),
            "linea": linea,
            "gd": gd,
            "representante": representante,
        },
    }


def cargar_excel_cobertura(
    db: Session,
    excel_bytes: bytes,
    pais_codigo: str,
    ciclo_codigo_default: Optional[str] = None,
    usuario: str = "sistema",
) -> dict:
    """
    Carga el Excel de Cobertura Predictiva.
    Acepta dos formatos de hojas:
      • Formato estándar: COB_DIM_CICLO / COB_DIM_CALENDARIO / COB_TARGET_MEDICO_CICLO / COB_FACT_VISITAS
      • Formato modelo VM: Parametros / Target_Medicos / Fact_Visitas

    Auto-crea representantes y ciclos si no existen.
    """
    import openpyxl
    from io import BytesIO
    from datetime import datetime as _dt, date as _date
    from decimal import Decimal

    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True, read_only=False)
    sheet_names_upper = {s.upper(): s for s in wb.sheetnames}

    # Alias: nombre alternativo → nombre estándar interno
    _ALIASES = {
        "TARGET_MEDICOS": "COB_TARGET_MEDICO_CICLO",
        "FACT_VISITAS": "COB_FACT_VISITAS",
        "PARAMETROS": "COB_PARAMETROS",
    }

    def _read_sheet(*keys: str):
        """Busca la hoja por varios nombres clave (incluye alias). Devuelve lista de dicts."""
        for key in keys:
            k = key.upper()
            k = _ALIASES.get(k, k)
            # buscar el alias inverso también
            real = sheet_names_upper.get(k)
            if not real:
                # buscar el alias original
                real = sheet_names_upper.get(key.upper())
            if not real:
                continue
            ws = wb[real]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return []
            # Encontrar la fila header: primera con ≥2 celdas no nulas
            header_idx = 0
            for i, r in enumerate(rows):
                non_null = sum(1 for c in r if c is not None and str(c).strip())
                if non_null >= 2:
                    header_idx = i
                    break
            header = [str(c).strip().upper() if c is not None else f"_COL{i}"
                      for i, c in enumerate(rows[header_idx])]
            result = []
            for row in rows[header_idx + 1:]:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                result.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
            return result
        return []

    def _s(v): return str(v).strip() if v is not None else None

    def _f(v):
        """Convierte a date. Acepta datetime, date, string ISO, serial Excel (int)."""
        if v is None:
            return None
        if isinstance(v, _dt):
            return v.date()
        if isinstance(v, _date):
            return v
        # Serial de Excel: días desde 1899-12-30
        try:
            n = int(float(str(v)))
            if 30000 < n < 60000:  # rango razonable 1982–2064
                from datetime import timedelta
                return (_date(1899, 12, 30) + timedelta(days=n))
        except (ValueError, TypeError):
            pass
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return _dt.strptime(str(v).strip(), fmt).date()
            except Exception:
                pass
        return None

    def _i(v):
        try:
            return int(float(str(v))) if v is not None else None
        except Exception:
            return None

    def _dec(v):
        try:
            return Decimal(str(v)) if v is not None else None
        except Exception:
            return None

    def _now():
        from datetime import timezone
        return _dt.now(timezone.utc).replace(tzinfo=None)

    stats: dict = {"ciclos": 0, "representantes": 0, "calendario": 0,
                   "target": 0, "visitas": 0, "errores": [], "advertencias": []}

    # ── Resolver PaisKey ─────────────────────────────────────────────────────
    pais_row = db.query(CatDimPais).filter(
        CatDimPais.codigo_pais == pais_codigo.upper()
    ).first()
    if not pais_row:
        raise ValueError(f"País '{pais_codigo}' no encontrado en cat.DimPais")
    pais_key = pais_row.pais_key

    # ── CICLO: hoja estándar o extraído de Parametros ────────────────────────
    ciclos_map: dict = {}  # codigo_ciclo → CicloKey

    import re as _re

    def _normalizar_ciclo(code: str) -> str:
        """Normaliza cualquier formato de ciclo al estándar del sistema: C06-2026.
        Acepta: '2026-06' → 'C06-2026', 'C06-2026' → 'C06-2026' (ya correcto)."""
        if not code:
            return code
        # Ya está en formato correcto: C06-2026
        m = _re.match(r'^[Cc](\d{2})-(\d{4})$', code.strip())
        if m:
            return f"C{m.group(1)}-{m.group(2)}"
        # Formato YYYY-MM → C MM-YYYY
        m2 = _re.match(r'^(\d{4})-(\d{2})$', code.strip())
        if m2:
            return f"C{m2.group(2)}-{m2.group(1)}"
        return code.strip()

    def _upsert_ciclo(codigo, fi, ff, linea=None, meta=None, dias_hab=None):
        nonlocal ciclos_map
        if not codigo or not fi or not ff:
            return
        # Normalizar al formato canónico C06-2026
        codigo = _normalizar_ciclo(codigo)
        if meta is not None and meta > 1:
            meta = meta / Decimal("100")
        filt = [CatDimCiclo.pais_key == pais_key, CatDimCiclo.codigo_ciclo == codigo]
        if linea:
            filt.append(CatDimCiclo.linea == linea)
        else:
            filt.append(CatDimCiclo.linea.is_(None))
        existing = db.query(CatDimCiclo).filter(*filt).first()
        if not existing:
            obj = CatDimCiclo(
                pais_key=pais_key,
                codigo_ciclo=codigo,
                linea=linea,
                fecha_inicio=fi,
                fecha_fin=ff,
                dias_habiles_ciclo=dias_hab,
                meta_cobertura_pct=meta or Decimal("0.90"),
                activo=True,
                fecha_carga_utc=_now(),
            )
            db.add(obj)
            db.flush()
            ciclos_map[codigo] = obj.ciclo_key
            stats["ciclos"] += 1
        else:
            existing.fecha_inicio = fi
            existing.fecha_fin = ff
            if meta:
                existing.meta_cobertura_pct = meta
            if dias_hab:
                existing.dias_habiles_ciclo = dias_hab
            existing.activo = True
            db.flush()
            ciclos_map[codigo] = existing.ciclo_key

    # Intentar hoja estándar primero
    for row in _read_sheet("COB_DIM_CICLO"):
        codigo = _s(row.get("CICLO_CODIGO") or row.get("CODIGO_CICLO") or row.get("CICLO"))
        _upsert_ciclo(
            codigo=codigo,
            fi=_f(row.get("FECHA_INICIO")),
            ff=_f(row.get("FECHA_FIN")),
            linea=_s(row.get("LINEA") or row.get("LINEA_ID")),
            meta=_dec(row.get("META_COBERTURA_PCT") or row.get("META_COBERTURA")),
            dias_hab=_i(row.get("DIAS_HABILES_CICLO")),
        )

    # Si no había hoja estándar, extraer de Parametros
    if not ciclos_map:
        params_dict: dict = {}
        for row in _read_sheet("COB_PARAMETROS", "PARAMETROS"):
            k = _s(row.get("PARÁMETRO") or row.get("PARAMETRO") or list(row.values())[0] if row else None)
            v = list(row.values())[1] if len(row) > 1 else None
            if k:
                params_dict[k.upper()] = v
        codigo_p = _s(params_dict.get("CICLO"))
        fi_p = _f(params_dict.get("FECHA INICIO CICLO") or params_dict.get("FECHA INICIO DEL CICLO"))
        ff_p = _f(params_dict.get("FECHA FIN CICLO") or params_dict.get("FECHA FIN DEL CICLO"))
        meta_p = _dec(params_dict.get("META COBERTURA"))
        dias_p = _i(params_dict.get("DÍAS HÁBILES DEL CICLO") or params_dict.get("DIAS HABILES DEL CICLO")
                    or params_dict.get("DÍAS HÁBILES CICLO"))
        if codigo_p and fi_p and ff_p:
            _upsert_ciclo(codigo_p, fi_p, ff_p, meta=meta_p, dias_hab=dias_p)
        else:
            # Extraer ciclo desde Target_Medicos como último recurso
            for row in _read_sheet("COB_TARGET_MEDICO_CICLO", "TARGET_MEDICOS"):
                c = _s(row.get("CICLO") or row.get("CICLO_CODIGO"))
                if c and c not in ciclos_map:
                    stats["advertencias"].append(
                        f"Ciclo '{c}' sin fechas en Parametros — se requiere hoja COB_DIM_CICLO o Parametros con fechas"
                    )
                break

    # ── Fallback: usar ciclo_codigo_default si aún no hay ciclos ────────────
    if not ciclos_map and ciclo_codigo_default:
        codigo_norm = _normalizar_ciclo(ciclo_codigo_default)
        existing = db.query(CatDimCiclo).filter(
            CatDimCiclo.pais_key == pais_key,
            CatDimCiclo.codigo_ciclo == codigo_norm,
        ).first()
        if existing:
            ciclos_map[codigo_norm] = existing.ciclo_key
            stats["advertencias"].append(
                f"Ciclo '{codigo_norm}' tomado del parámetro de ciclo (no venía en el Excel)."
            )
        else:
            stats["advertencias"].append(
                f"Ciclo '{codigo_norm}' especificado pero no encontrado en cat.DimCiclo. "
                "Carga primero el ciclo con la hoja COB_DIM_CICLO."
            )

    db.flush()

    # ── Representantes: precargar y auto-crear si no existen ─────────────────
    rep_rows = db.query(CatDimRepresentanteMedico).filter(
        CatDimRepresentanteMedico.activo == True
    ).all()
    rep_map: dict = {(r.codigo_representante or "").upper(): r.representante_key for r in rep_rows}

    def _get_or_create_rep(vm_cod: str, nombre_vm: str = "", equipo: str = "") -> Optional[int]:
        key = vm_cod.upper()
        if key in rep_map:
            return rep_map[key]
        # Auto-crear
        obj = CatDimRepresentanteMedico(
            pais_key=pais_key,
            codigo_representante=vm_cod,
            nombre_representante=nombre_vm or vm_cod,
            equipo_texto=equipo or None,
            activo=True,
        )
        db.add(obj)
        db.flush()
        rep_map[key] = obj.representante_key
        stats["representantes"] += 1
        return obj.representante_key

    # Pre-scan Target_Medicos para crear representantes antes de cargar target
    # Nuevo formato: codigo_id = código del VM; fallback al formato anterior
    for row in _read_sheet("COB_TARGET_MEDICO_CICLO", "TARGET_MEDICOS"):
        vm_cod = _s(
            row.get("CODIGO_ID") or row.get("VM_CODIGO_ID")
            or row.get("VM_CODIGO") or row.get("VM_ID")
        )
        if vm_cod:
            nombre_vm = _s(row.get("NOMBRE_VM") or row.get("NOMBRE_REPRESENTANTE") or "")
            equipo = _s(row.get("LINEA") or row.get("EQUIPO") or "")
            _get_or_create_rep(vm_cod, nombre_vm or "", equipo or "")

    db.flush()

    # ── Caches de médicos y especialidades ──────────────────────────────────
    # medico_map: (pais_key, medico_id_negocio) → medico_key
    med_rows = db.query(CatDimMedico).filter(CatDimMedico.pais_key == pais_key).all()
    medico_map_by_id: dict = {m.medico_id: m.medico_key for m in med_rows if m.medico_id}
    medico_map_by_cod: dict = {(m.codigo_medico or "").upper(): m.medico_key
                                for m in med_rows if m.codigo_medico}

    esp_rows = db.query(CatDimEspecialidad).filter(CatDimEspecialidad.pais_key == pais_key).all()
    especialidad_map: dict = {e.especialidad.upper(): e.especialidad_key for e in esp_rows}

    def _get_or_create_especialidad(nombre: Optional[str]) -> Optional[int]:
        if not nombre:
            return None
        key = nombre.upper()
        if key in especialidad_map:
            return especialidad_map[key]
        obj = CatDimEspecialidad(pais_key=pais_key, especialidad=nombre, activo=True)
        db.add(obj)
        db.flush()
        especialidad_map[key] = obj.especialidad_key
        return obj.especialidad_key

    # medico_map_by_name: (nombre_upper, especialidad_key) → medico_key  (dedup por nombre)
    medico_map_by_name: dict = {
        (m.nombre_medico.upper(), m.especialidad_key): m.medico_key
        for m in med_rows if m.nombre_medico
    }

    def _get_or_create_medico(
        med_id: Optional[int],
        med_cod: Optional[str],
        nombre: Optional[str],
        especialidad_key: Optional[int] = None,
    ) -> Optional[int]:
        """Resuelve medico_key desde DimMedico, creando si no existe.
        Prioridad: ID numérico → código de texto → nombre+especialidad (evita UNIQUE violation)."""
        if med_id and med_id in medico_map_by_id:
            return medico_map_by_id[med_id]
        cod_upper = (med_cod or "").upper()
        if cod_upper and cod_upper in medico_map_by_cod:
            existing_key = medico_map_by_cod[cod_upper]
            if med_id and existing_key:
                obj = db.query(CatDimMedico).filter(
                    CatDimMedico.medico_key == existing_key
                ).first()
                if obj and not obj.medico_id:
                    obj.medico_id = med_id
                    db.flush()
                    medico_map_by_id[med_id] = existing_key
            return existing_key
        # Fallback: dedup por nombre + especialidad (el mismo médico puede tener distintos códigos)
        if nombre:
            name_key = (nombre.upper(), especialidad_key)
            if name_key in medico_map_by_name:
                existing_key = medico_map_by_name[name_key]
                if cod_upper:
                    medico_map_by_cod[cod_upper] = existing_key
                if med_id:
                    medico_map_by_id[med_id] = existing_key
                return existing_key
        if not nombre:
            return None
        # Crear nuevo registro en DimMedico
        obj = CatDimMedico(
            pais_key=pais_key,
            medico_id=med_id,
            codigo_medico=med_cod,
            nombre_medico=nombre,
            especialidad_key=especialidad_key,
            activo=True,
        )
        db.add(obj)
        db.flush()
        if med_id:
            medico_map_by_id[med_id] = obj.medico_key
        if med_cod:
            medico_map_by_cod[cod_upper] = obj.medico_key
        medico_map_by_name[(nombre.upper(), especialidad_key)] = obj.medico_key
        return obj.medico_key

    # ── COB_DIM_CALENDARIO (opcional) ────────────────────────────────────────
    for row in _read_sheet("COB_DIM_CALENDARIO"):
        fecha = _f(row.get("FECHA"))
        ciclo_codigo_cal = _s(row.get("CICLO_CODIGO") or row.get("CICLO"))
        ciclo_key_cal = ciclos_map.get(ciclo_codigo_cal) if ciclo_codigo_cal else None
        if not fecha:
            continue
        existing = db.query(CatDimCalendario).filter(
            CatDimCalendario.pais_key == pais_key,
            CatDimCalendario.fecha == fecha,
        ).first()
        if not existing:
            es_hab = row.get("ES_HABIL")
            es_fer = row.get("ES_FERIADO")
            obj = CatDimCalendario(
                pais_key=pais_key,
                fecha=fecha,
                ciclo_key=ciclo_key_cal,
                es_habil=bool(_i(es_hab) if es_hab is not None else 1),
                es_feriado=bool(_i(es_fer) if es_fer is not None else 0),
                semana=_i(row.get("SEMANA")),
                mes=_i(row.get("MES")),
                anio=_i(row.get("ANIO")),
                nota=_s(row.get("NOTA")),
                fecha_carga_utc=_now(),
            )
            db.add(obj)
            stats["calendario"] += 1

    db.flush()

    # ── Target médicos ───────────────────────────────────────────────────────
    # Nuevo formato de columnas Excel:
    #   codigo_id = código del VM/Representante
    #   medico_id = ID numérico del médico (llave hacia DimMedico)
    #   nombre    = nombre del médico
    #   Especialidad, potencial, programado_flag, gerente_id, linea_id, Ciclo
    for row in _read_sheet("COB_TARGET_MEDICO_CICLO", "TARGET_MEDICOS"):
        ciclo_codigo_t = _normalizar_ciclo(_s(row.get("CICLO") or row.get("CICLO_CODIGO") or row.get("CICLO_ID")) or "")
        ciclo_key_t = ciclos_map.get(ciclo_codigo_t) if ciclo_codigo_t else None
        if not ciclo_key_t:
            if ciclo_codigo_t:
                stats["errores"].append(f"Target: ciclo '{ciclo_codigo_t}' no encontrado")
            continue

        # VM: nuevo formato → CODIGO_ID; fallback formato antiguo
        vm_cod = _s(
            row.get("CODIGO_ID") or row.get("VM_CODIGO_ID")
            or row.get("VM_CODIGO") or row.get("VM_ID")
        )
        # Médico: MEDICO_ID puede ser numérico o string (ej. 'VM01_MED_001')
        med_id_raw = _i(row.get("MEDICO_ID"))
        med_cod = _s(row.get("MEDICO_CODIGO_ID") or row.get("MEDICO_CODIGO"))
        # Si MEDICO_ID no era numérico, usarlo como código de texto
        if not med_id_raw and not med_cod:
            med_cod = _s(row.get("MEDICO_ID"))
        nombre_med = _s(row.get("NOMBRE") or row.get("NOMBRE_MEDICO") or row.get("MEDICO_NOMBRE"))
        # Sin nombre → usar el código como nombre para que pueda crearse en DimMedico
        if not nombre_med and med_cod:
            nombre_med = med_cod

        if not vm_cod:
            stats["errores"].append("Target: falta código de VM (CODIGO_ID/VM_CODIGO_ID)")
            continue
        if not med_id_raw and not med_cod and not nombre_med:
            stats["errores"].append("Target: falta identificador de médico")
            continue

        rep_key = rep_map.get(vm_cod.upper())
        if not rep_key:
            stats["errores"].append(f"Target: VM '{vm_cod}' no encontrado")
            continue

        # Especialidad → clave en DimEspecialidad (auto-crea si no existe)
        esp_nombre = _s(row.get("ESPECIALIDAD") or row.get("ESPECIALIDAD_MEDICO"))
        esp_key = _get_or_create_especialidad(esp_nombre)

        # Resolver/crear médico en DimMedico
        medico_key_t = _get_or_create_medico(med_id_raw, med_cod, nombre_med, esp_key)

        # código de origen uniforme: str(medico_id) si existe, si no el código string
        cod_origen = str(med_id_raw) if med_id_raw else (med_cod or nombre_med or "")
        potencial_val = _s(row.get("POTENCIAL") or row.get("CATEGORIA_MEDICA"))
        prog_raw = row.get("PROGRAMADO_FLAG") or row.get("PROGRAMADO")

        existing = db.query(CatFactTargetMedicoCiclo).filter(
            CatFactTargetMedicoCiclo.ciclo_key == ciclo_key_t,
            CatFactTargetMedicoCiclo.representante_key == rep_key,
            CatFactTargetMedicoCiclo.codigo_medico_origen == cod_origen,
        ).first()
        if not existing:
            obj = CatFactTargetMedicoCiclo(
                ciclo_key=ciclo_key_t,
                pais_key=pais_key,
                representante_key=rep_key,
                medico_key=medico_key_t,
                codigo_medico_origen=cod_origen,
                nombre_medico=nombre_med,
                especialidad_key=esp_key,
                potencial=potencial_val,
                territorio=_s(row.get("TERRITORIO") or row.get("GD")),
                frecuencia_objetivo=_i(row.get("FRECUENCIA_OBJETIVO")),
                programado_flag=bool(_i(prog_raw) if prog_raw is not None else 1),
                categoria_medica=potencial_val,
                fuente="Excel",
                fecha_carga_utc=_now(),
            )
            db.add(obj)
            stats["target"] += 1

    db.flush()

    # ── Visitas ──────────────────────────────────────────────────────────────
    # Nuevo formato de columnas Excel:
    #   visita_id   = ID de la visita
    #   fecha_visita = fecha de la visita
    #   Ciclo        = código del ciclo
    #   codigo_id    = código del VM/Representante
    #   medico_id    = ID numérico del médico
    #   tipo_contacto, estado_visita, producto_foco
    for row in _read_sheet("COB_FACT_VISITAS", "FACT_VISITAS"):
        ciclo_codigo_v = _normalizar_ciclo(_s(row.get("CICLO") or row.get("CICLO_CODIGO")) or "")
        ciclo_key_v = ciclos_map.get(ciclo_codigo_v) if ciclo_codigo_v else None
        if not ciclo_key_v:
            continue

        # VM: nuevo formato → CODIGO_ID; fallback formato antiguo
        vm_cod = _s(
            row.get("CODIGO_ID") or row.get("VM_CODIGO_ID")
            or row.get("VM_ID") or row.get("VM_CODIGO")
        )
        # Médico: MEDICO_ID puede ser numérico o string (ej. 'VM01_MED_001')
        med_id_raw_v = _i(row.get("MEDICO_ID"))
        med_cod_v = _s(row.get("MEDICO_CODIGO_ID") or row.get("MEDICO_CODIGO"))
        # Si MEDICO_ID no era numérico, usarlo como código de texto
        if not med_id_raw_v and not med_cod_v:
            med_cod_v = _s(row.get("MEDICO_ID"))
        fecha_v = _f(row.get("FECHA_VISITA") or row.get("FECHA"))

        if not vm_cod or not fecha_v:
            continue
        if not med_id_raw_v and not med_cod_v:
            continue

        rep_key = rep_map.get(vm_cod.upper())
        if not rep_key:
            continue

        # código de origen uniforme para joins con FactTargetMedicoCiclo
        cod_origen_v = str(med_id_raw_v) if med_id_raw_v else (med_cod_v or "")

        # Resolver medico_key (solo lookup, sin auto-crear para visitas)
        medico_key_v = (
            medico_map_by_id.get(med_id_raw_v) if med_id_raw_v
            else medico_map_by_cod.get((med_cod_v or "").upper())
        )

        estado = _s(row.get("ESTADO_VISITA") or row.get("ESTADO")) or "Pendiente"
        obj = CatFactVisitaMedica(
            visita_id_origen=_s(row.get("VISITA_ID")),
            fecha_visita=fecha_v,
            ciclo_key=ciclo_key_v,
            pais_key=pais_key,
            representante_key=rep_key,
            medico_key=medico_key_v,
            codigo_medico_origen=cod_origen_v,
            tipo_contacto=_s(row.get("TIPO_CONTACTO") or row.get("TIPO")),
            estado_visita=estado,
            producto_foco=_s(row.get("PRODUCTO_FOCO") or row.get("PRODUCTO")),
            fuente="Excel",
            fecha_carga_utc=_now(),
        )
        db.add(obj)
        stats["visitas"] += 1

    db.commit()
    logger.info(
        f"[cargar_excel_cobertura] país={pais_codigo} "
        f"ciclos={stats['ciclos']} reps={stats['representantes']} "
        f"cal={stats['calendario']} target={stats['target']} visitas={stats['visitas']}"
    )
    return stats
