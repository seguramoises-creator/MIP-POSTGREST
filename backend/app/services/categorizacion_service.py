"""
Servicio de Categorización Médica — nueva versión (jun-2026).

Loader de Categorización Médica desde Excel con SQLAlchemy / psycopg2 (PostgreSQL).

Flujo:
1. Leer hojas del Excel con openpyxl (data_only=True para fórmulas calculadas).
2. UPSERT de catálogos cat.Dim* (SQL portable).
3. Crear registro en cat.LoadBatch.
4. INSERT en stg.MedicoCategoriaInput.
5. Calcular categorías con el motor 100% Python (calcular_categorias_py).
6. Retornar resumen del lote.
"""

from __future__ import annotations

import io
from datetime import datetime, date, timezone
from decimal import Decimal
from typing import Any, Optional

from loguru import logger
from sqlalchemy import text
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

NULL_STRINGS = {"", "null", "none", "nan"}


# ── Utilidades ─────────────────────────────────────────────────────────────────

def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v.lower() in NULL_STRINGS:
            return None
        return v
    if isinstance(value, datetime):
        return value.date()
    return value


def _as_int(value):
    v = _clean(value)
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _as_decimal(value):
    v = _clean(value)
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def _as_bool(value, default=True):
    """Convierte 1/0/'1'/'true'/'si'… a bool de Python.

    Las columnas de estado (Activo/Requerido) son BOOLEAN en Postgres; psycopg2
    NO castea int→boolean, así que hay que pasar bool nativo. `default=True`
    replica el idiom heredado `_as_int(...) or 1` (activo por defecto).
    """
    v = _clean(value)
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        return v.strip().lower() in ("1", "true", "t", "si", "sí", "y", "yes")
    return default


def _as_date(value):
    v = _clean(value)
    if v is None:
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v.replace(" 00:00:00", "")).date()
        except ValueError:
            return None
    return None


def _read_table(wb, sheet_name, table_name=None):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Hoja no encontrada: '{sheet_name}'")
    ws = wb[sheet_name]
    if not ws.tables:
        # Hoja sin tabla Excel definida — leer fila de encabezado + datos directamente
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []
        headers = [_clean(v) for v in all_rows[0]]
        rows = []
        for raw in all_rows[1:]:
            vals = [_clean(v) for v in raw]
            if all(v is None for v in vals):
                continue
            rows.append(dict(zip(headers, vals)))
        return rows
    table = ws.tables[table_name] if table_name and table_name in ws.tables else next(iter(ws.tables.values()))
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [_clean(ws.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
    rows = []
    for ri in range(min_row + 1, max_row + 1):
        vals = [_clean(ws.cell(ri, col).value) for col in range(min_col, max_col + 1)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


# ── Catálogos ──────────────────────────────────────────────────────────────────

def _get_pais_key(db, codigo_pais):
    row = db.execute(text('SELECT "PaisKey" FROM "cat"."DimPais" WHERE "CodigoPais" = :c'), {"c": codigo_pais}).fetchone()
    if row is None:
        raise ValueError(f"País no encontrado en cat.DimPais: '{codigo_pais}'")
    return int(row[0])


def _get_componente_key(db, codigo):
    row = db.execute(text('SELECT "ComponenteKey" FROM "cat"."DimComponenteCategoria" WHERE "CodigoComponente" = :c'), {"c": codigo}).fetchone()
    if row is None:
        raise ValueError(f"Componente no encontrado: '{codigo}'")
    return int(row[0])


def _portable_upsert(db, tabla, claves, valores, solo_insert=None):
    """UPSERT portable a PostgreSQL: SELECT de existencia + INSERT o UPDATE como
    dos statements separados (SELECT de existencia y luego INSERT/UPDATE)
    (procedural, inválido en Postgres) sin requerir ``ON CONFLICT`` ni constraints
    únicos sobre la clave natural.

    - ``claves``      : columnas de la clave natural (van al WHERE y al INSERT)
    - ``valores``     : columnas que se insertan **y** se actualizan
    - ``solo_insert`` : columnas que solo se fijan al insertar (p.ej. FechaCargaUtc)

    Los nombres de columna se usan como nombres de parámetro (no llevan espacios).
    """
    solo_insert = solo_insert or {}
    where = " AND ".join(f'"{c}" = :{c}' for c in claves)
    existe = db.execute(text(f'SELECT 1 FROM {tabla} WHERE {where}'), claves).first()
    if existe:
        if valores:
            set_cols = ", ".join(f'"{c}" = :{c}' for c in valores)
            db.execute(text(f'UPDATE {tabla} SET {set_cols} WHERE {where}'), {**claves, **valores})
    else:
        todas = {**claves, **valores, **solo_insert}
        cols = ", ".join(f'"{c}"' for c in todas)
        binds = ", ".join(f":{c}" for c in todas)
        db.execute(text(f'INSERT INTO {tabla} ({cols}) VALUES ({binds})'), todas)


def _upsert_pais(db, rows):
    for r in rows:
        _portable_upsert(
            db, '"cat"."DimPais"',
            claves={"CodigoPais": r.get("codigo")},
            valores={"PaisIdOrigen": _as_int(r.get("id")), "NombrePais": r.get("nombre"),
                     "Moneda": r.get("moneda"), "ZonaHoraria": r.get("zona_horaria"),
                     "Activo": _as_bool(r.get("activo"))},
            solo_insert={"FechaCargaUtc": datetime.now(timezone.utc)},
        )


def _upsert_componente(db, rows):
    for r in rows:
        _portable_upsert(
            db, '"cat"."DimComponenteCategoria"',
            claves={"CodigoComponente": r["CodigoComponente"]},
            valores={"NombreComponente": r["NombreComponente"], "TipoEvaluacion": r["TipoEvaluacion"],
                     "PesoComponentePct": _as_decimal(r["PesoComponentePct"]),
                     "Requerido": _as_bool(r["Requerido"]), "Activo": _as_bool(r["Activo"])},
            solo_insert={"FechaCargaUtc": datetime.now(timezone.utc)},
        )


def _upsert_clasificacion(db, rows):
    for r in rows:
        pk = _get_pais_key(db, r["pais_codigo"])
        _portable_upsert(
            db, '"cat"."DimClasificacionMedica"',
            claves={"PaisKey": pk, "Clase": r["Clase"], "VigenteDesde": _as_date(r["VigenteDesde"])},
            valores={"PuntajeMinPct": _as_decimal(r["PuntajeMinPct"]), "PuntajeMaxPct": _as_decimal(r["PuntajeMaxPct"]),
                     "OrdenClase": _as_int(r["OrdenClase"]) or 0, "VigenteHasta": _as_date(r.get("VigenteHasta")),
                     "Activo": _as_bool(r["Activo"])},
            solo_insert={"FechaCargaUtc": datetime.now(timezone.utc)},
        )


def _upsert_regla(db, rows):
    for r in rows:
        pk = _get_pais_key(db, r["pais_codigo"])
        ck = _get_componente_key(db, r["CodigoComponente"])
        _portable_upsert(
            db, '"cat"."DimReglaCategoriaMedica"',
            claves={"PaisKey": pk, "ComponenteKey": ck, "CodigoRegla": r["CodigoRegla"],
                    "VigenteDesde": _as_date(r["VigenteDesde"])},
            valores={"Detalle": str(r["Detalle"]), "ValorMinimo": _as_decimal(r.get("ValorMinimo")),
                     "ValorMaximo": _as_decimal(r.get("ValorMaximo")), "ValorTexto": r.get("ValorTexto"),
                     "Criterio": _as_int(r["Criterio"]), "PesoComponentePct": _as_decimal(r["PesoComponentePct"]),
                     "PuntajePct": _as_decimal(r["PuntajePct"]), "VigenteHasta": _as_date(r.get("VigenteHasta")),
                     "Activo": _as_bool(r["Activo"])},
            solo_insert={"FechaCargaUtc": datetime.now(timezone.utc)},
        )


def _upsert_equipo(db, rows):
    for r in rows:
        pk = _get_pais_key(db, r["pais_codigo"])
        _portable_upsert(
            db, '"cat"."DimEquipo"',
            claves={"PaisKey": pk, "CodigoEquipo": r["codigo"]},
            valores={"NombreEquipo": r["nombre"], "Descripcion": r.get("descripcion"),
                     "Activo": _as_bool(r.get("activo"))},
        )


def _upsert_representante(db, rows):
    for r in rows:
        pk = _get_pais_key(db, r["pais_codigo"])
        _portable_upsert(
            db, '"cat"."DimRepresentanteMedico"',
            claves={"PaisKey": pk, "CodigoRepresentante": r["codigo_id"]},
            valores={"RepresentanteIdOrigen": _as_int(r.get("id")), "NombreRepresentante": r["nombre"],
                     "LineaIdOrigen": _as_int(r.get("linea_id")), "GerenteIdOrigen": _as_int(r.get("gerente_id")),
                     "Email": r.get("email"), "Zona": r.get("zona"), "FechaIngreso": _as_date(r.get("fecha_ingreso")),
                     "Cedula": r.get("cedula"), "CodigoOrigenExcel": r.get("codigo_origen_excel"),
                     "EquipoTexto": r.get("Equipo"), "Activo": _as_bool(r.get("activo"))},
        )


def _insert_staging(db, load_batch_key, rows, periodo=None):
    for idx, r in enumerate(rows, start=1):
        db.execute(text("""
            INSERT INTO "stg"."MedicoCategoriaInput"
                ("LoadBatchKey","RowNumber","CodigoPais","Periodo","Equipo","LineaIdOrigen","CodigoRepresentante","NombreRepresentante",
                 "Medico","CentroMedico","Especialidad","Provincia","Municipio","PacientesSemana","CostoConsulta","RecetasSemana",
                 "UbicacionTerritorialCM","KOL","CategoriaExcel","Activo")
            VALUES (:batch,:row,:pais,:periodo,:equipo,:linea,:cod_rep,:nom_rep,
                    :medico,:centro,:esp,:provincia,:municipio,:pac,:costo,:recetas,
                    :ubic,:kol,:cat_excel,:activo)
        """), {
            "batch": load_batch_key, "row": idx, "pais": r.get("pais_codigo"),
            "periodo": r.get("Periodo") or periodo,  # usa el período del usuario si el Excel no lo trae
            "equipo": r.get("Equipo"), "linea": _as_int(r.get("linea_id")),
            "cod_rep": r.get("codigo_id"), "nom_rep": r.get("nombre"), "medico": r.get("Medico"),
            "centro": r.get("CentroMedico"), "esp": r.get("Especialidad"),
            "provincia": r.get("Provincia"), "municipio": r.get("Municipio"),
            "pac": _as_decimal(r.get("PacientesSemana")), "costo": _as_decimal(r.get("CostoConsulta")),
            "recetas": r.get("RecetasSemana"), "ubic": r.get("UbicacionTerritorialCM"),
            "kol": r.get("KOL"), "cat_excel": r.get("CategoriaExcel"),
            "activo": _as_bool(r.get("Activo")),  # stg.Activo es BOOLEAN (psycopg2 no castea int→bool)
        })


# ── Punto de entrada ───────────────────────────────────────────────────────────

def _limpiar_periodo(db: Session, periodo: str) -> int:
    """Elimina todos los datos del período indicado antes de una recarga.
    Respeta el orden de FK: Detalle → Snapshot → Staging → LoadBatch.
    Retorna la cantidad de lotes eliminados.
    """
    lotes = db.execute(
        text('SELECT "LoadBatchKey" FROM "cat"."LoadBatch" WHERE "Periodo" = :p'),
        {"p": periodo},
    ).fetchall()
    if not lotes:
        return 0
    keys = [r[0] for r in lotes]
    keys_str = ",".join(str(k) for k in keys)
    db.execute(text(f"""
        DELETE FROM "cat"."FactMedicoCategoriaDetalle"
        WHERE "MedicoCategoriaKey" IN (
            SELECT "MedicoCategoriaKey" FROM "cat"."FactMedicoCategoriaSnapshot"
            WHERE "LoadBatchKey" IN ({keys_str})
        )
    """))
    db.execute(text(f'DELETE FROM "cat"."FactMedicoCategoriaSnapshot" WHERE "LoadBatchKey" IN ({keys_str})'))
    db.execute(text(f'DELETE FROM "stg"."MedicoCategoriaInput" WHERE "LoadBatchKey" IN ({keys_str})'))
    db.execute(text(f'DELETE FROM "cat"."LoadBatch" WHERE "LoadBatchKey" IN ({keys_str})'))
    logger.info(f"[CAT] Período '{periodo}' limpiado — {len(keys)} lote(s) eliminados: {keys_str}")
    return len(keys)


# Mapeo de componentes → columna de staging (numérico vs texto), como en el SP.
_COMP_NUM = {"PACIENTES_SEMANA": "PacientesSemana", "PODER_ADQUISITIVO": "CostoConsulta"}
_COMP_TXT = {"POTENCIAL_PRESCRIPCION": "RecetasSemana",
             "UBICACION_TERRITORIAL_CM": "UbicacionTerritorialCM", "KOL": "KOL"}


def _mejor_regla(matches):
    """rn=1 del SP: ORDER BY Criterio DESC, ReglaKey ASC. Devuelve la regla top o None."""
    if not matches:
        return None
    orden = sorted(matches, key=lambda r: r["ReglaKey"])
    orden.sort(key=lambda r: (r["Criterio"] or ""), reverse=True)  # NULL/'' al final (como SQL NULL en DESC)
    return orden[0]


def _regla_aplica(regla, valor_num, valor_txt):
    if valor_num is not None:
        if (regla["ValorMinimo"] is None or valor_num >= regla["ValorMinimo"]) and \
           (regla["ValorMaximo"] is None or valor_num <= regla["ValorMaximo"]):
            return True
    if valor_txt is not None and regla["ValorTexto"] is not None:
        if valor_txt.strip().upper() == str(regla["ValorTexto"]).strip().upper():
            return True
    return False


# SQL de sincronización de dimensiones maestras: cat.* (cargado del Excel) → Config.DIM_*
# (fuente única del sistema, usada por Panel Médico / Productos). Aditivo (insert-if-not-exists,
# case-insensitive). El país se toma de cat.DimPais.CodigoPais y se
# valida contra Config.DIM_Pais para no romper la FK.
_SYNC_ESPECIALIDAD = """
INSERT INTO "Config"."DIM_Especialidad" (nombre, activo)
SELECT DISTINCT TRIM(e."Especialidad"), TRUE FROM "cat"."DimEspecialidad" e
WHERE e."Especialidad" IS NOT NULL AND TRIM(e."Especialidad") <> ''
  AND NOT EXISTS (SELECT 1 FROM "Config"."DIM_Especialidad" d WHERE LOWER(d.nombre)=LOWER(TRIM(e."Especialidad")))
"""
_SYNC_CENTRO = """
INSERT INTO "Config"."DIM_CentroMedico" (pais_codigo, nombre, activo)
SELECT DISTINCT pa."CodigoPais", TRIM(c."CentroMedico"), TRUE
FROM "cat"."DimCentroMedico" c JOIN "cat"."DimPais" pa ON pa."PaisKey"=c."PaisKey"
WHERE c."CentroMedico" IS NOT NULL AND TRIM(c."CentroMedico") <> ''
  AND EXISTS (SELECT 1 FROM "Config"."DIM_Pais" cp WHERE cp.codigo=pa."CodigoPais")
  AND NOT EXISTS (SELECT 1 FROM "Config"."DIM_CentroMedico" d WHERE d.pais_codigo=pa."CodigoPais" AND d.nombre=TRIM(c."CentroMedico"))
"""
_SYNC_PROVINCIA = """
INSERT INTO "Config"."DIM_Provincia" (pais_codigo, nombre, activo)
SELECT DISTINCT pa."CodigoPais", TRIM(g."Provincia"), TRUE
FROM "cat"."DimGeografia" g JOIN "cat"."DimPais" pa ON pa."PaisKey"=g."PaisKey"
WHERE g."Provincia" IS NOT NULL AND TRIM(g."Provincia") <> ''
  AND EXISTS (SELECT 1 FROM "Config"."DIM_Pais" cp WHERE cp.codigo=pa."CodigoPais")
  AND NOT EXISTS (SELECT 1 FROM "Config"."DIM_Provincia" d WHERE d.pais_codigo=pa."CodigoPais" AND LOWER(d.nombre)=LOWER(TRIM(g."Provincia")))
"""
_SYNC_MUNICIPIO = """
INSERT INTO "Config"."DIM_Municipio" (provincia_id, nombre, activo)
SELECT DISTINCT pr.id, TRIM(g."Municipio"), TRUE
FROM "cat"."DimGeografia" g JOIN "cat"."DimPais" pa ON pa."PaisKey"=g."PaisKey"
JOIN "Config"."DIM_Provincia" pr ON pr.pais_codigo=pa."CodigoPais" AND LOWER(pr.nombre)=LOWER(TRIM(g."Provincia"))
WHERE g."Municipio" IS NOT NULL AND TRIM(g."Municipio") <> ''
  AND NOT EXISTS (SELECT 1 FROM "Config"."DIM_Municipio" d WHERE d.provincia_id=pr.id AND LOWER(d.nombre)=LOWER(TRIM(g."Municipio")))
"""


def sincronizar_dims_maestras(db: Session) -> dict:
    """Puebla las dimensiones compartidas Config.DIM_* con los valores reales cargados
    en cat.* (Especialidad, Centro Médico, Provincia, Municipio). Aditivo e idempotente:
    solo inserta lo que falta. Es la fuente única usada por todo el sistema."""
    r = {}
    r["especialidades"] = db.execute(text(_SYNC_ESPECIALIDAD)).rowcount
    r["provincias"] = db.execute(text(_SYNC_PROVINCIA)).rowcount
    r["municipios"] = db.execute(text(_SYNC_MUNICIPIO)).rowcount   # tras provincias (dependen de ellas)
    r["centros"] = db.execute(text(_SYNC_CENTRO)).rowcount
    db.flush()
    logger.info(f"Dims maestras sincronizadas desde cat.*: {r}")
    return r


def calcular_categorias_py(db: Session, load_batch_key: int) -> None:
    """Equivalente Python de cat.sp_CalcularCategoriaMedica (ETL dimensional + scoring).
    Conforma dimensiones, calcula puntaje por reglas y escribe Snapshot + Detalle."""
    from datetime import datetime, timezone
    from collections import defaultdict
    b = {"b": load_batch_key}
    fc = {"fc": datetime.now(timezone.utc).date()}

    # 1-5: conformar dimensiones (SQL portable: TRIM/COALESCE)
    # 'Activo' es NOT NULL en estas dimensiones; su default es client-side del ORM
    # y NO aplica a INSERT...SELECT por text() crudo, así que se fija TRUE explícito.
    db.execute(text("""INSERT INTO "cat"."DimEspecialidad" ("PaisKey", "Especialidad", "Activo")
        SELECT DISTINCT p."PaisKey", TRIM(s."Especialidad"), TRUE FROM "stg"."MedicoCategoriaInput" s
        JOIN "cat"."DimPais" p ON p."CodigoPais"=s."CodigoPais"
        WHERE s."LoadBatchKey"=:b AND s."Especialidad" IS NOT NULL
          AND NOT EXISTS (SELECT 1 FROM "cat"."DimEspecialidad" d WHERE d."PaisKey"=p."PaisKey" AND d."Especialidad"=TRIM(s."Especialidad"))"""), b)
    db.execute(text("""INSERT INTO "cat"."DimCentroMedico" ("PaisKey", "CentroMedico", "Activo")
        SELECT DISTINCT p."PaisKey", TRIM(s."CentroMedico"), TRUE FROM "stg"."MedicoCategoriaInput" s
        JOIN "cat"."DimPais" p ON p."CodigoPais"=s."CodigoPais"
        WHERE s."LoadBatchKey"=:b AND s."CentroMedico" IS NOT NULL
          AND NOT EXISTS (SELECT 1 FROM "cat"."DimCentroMedico" d WHERE d."PaisKey"=p."PaisKey" AND d."CentroMedico"=TRIM(s."CentroMedico"))"""), b)
    db.execute(text("""INSERT INTO "cat"."DimGeografia" ("PaisKey", "Provincia", "Municipio", "Activo")
        SELECT DISTINCT p."PaisKey", TRIM(s."Provincia"), TRIM(s."Municipio"), TRUE FROM "stg"."MedicoCategoriaInput" s
        JOIN "cat"."DimPais" p ON p."CodigoPais"=s."CodigoPais"
        WHERE s."LoadBatchKey"=:b AND s."Provincia" IS NOT NULL AND s."Municipio" IS NOT NULL
          AND NOT EXISTS (SELECT 1 FROM "cat"."DimGeografia" d WHERE d."PaisKey"=p."PaisKey" AND d."Provincia"=TRIM(s."Provincia") AND d."Municipio"=TRIM(s."Municipio"))"""), b)
    db.execute(text("""INSERT INTO "cat"."DimRepresentanteMedico" ("PaisKey", "CodigoRepresentante", "NombreRepresentante", "LineaIdOrigen", "EquipoTexto", "Activo")
        SELECT DISTINCT p."PaisKey", s."CodigoRepresentante", TRIM(s."NombreRepresentante"), s."LineaIdOrigen", TRIM(s."Equipo"), TRUE
        FROM "stg"."MedicoCategoriaInput" s JOIN "cat"."DimPais" p ON p."CodigoPais"=s."CodigoPais"
        WHERE s."LoadBatchKey"=:b AND s."CodigoRepresentante" IS NOT NULL
          AND NOT EXISTS (SELECT 1 FROM "cat"."DimRepresentanteMedico" d WHERE d."PaisKey"=p."PaisKey" AND d."CodigoRepresentante"=s."CodigoRepresentante")"""), b)
    db.execute(text("""INSERT INTO "cat"."DimMedico" ("PaisKey", "CodigoMedico", "NombreMedico", "EspecialidadKey", "Activo")
        SELECT DISTINCT p."PaisKey", NULL, TRIM(s."Medico"), esp."EspecialidadKey", TRUE
        FROM "stg"."MedicoCategoriaInput" s JOIN "cat"."DimPais" p ON p."CodigoPais"=s."CodigoPais"
        LEFT JOIN "cat"."DimEspecialidad" esp ON esp."PaisKey"=p."PaisKey" AND esp."Especialidad"=TRIM(s."Especialidad")
        WHERE s."LoadBatchKey"=:b
          AND NOT EXISTS (SELECT 1 FROM "cat"."DimMedico" d WHERE d."PaisKey"=p."PaisKey" AND d."NombreMedico"=TRIM(s."Medico")
            AND COALESCE(d."EspecialidadKey",-1)=COALESCE(esp."EspecialidadKey",-1))"""), b)
    db.flush()

    # Unificación: las dimensiones compartidas del sistema viven en Config.DIM_* (las
    # usan Panel Médico, Productos, etc.). Al cargar Categorización, poblamos ahí las
    # especialidades/provincias/municipios/centros reales del Excel — fuente única.
    sincronizar_dims_maestras(db)

    # 6-7: filas enriquecidas con las keys de dimensión
    rows = db.execute(text("""
        SELECT s."RowNumber", p."PaisKey", med."MedicoKey", cm."CentroMedicoKey", geo."GeografiaKey", rep."RepresentanteKey",
               s."Periodo", s."Equipo", s."LineaIdOrigen", s."PacientesSemana", s."CostoConsulta", s."RecetasSemana",
               s."UbicacionTerritorialCM", s."KOL", TRIM(s."Provincia") AS "Provincia", TRIM(s."Municipio") AS "Municipio", s."CategoriaExcel"
        FROM "stg"."MedicoCategoriaInput" s
        JOIN "cat"."DimPais" p ON p."CodigoPais"=s."CodigoPais"
        LEFT JOIN "cat"."DimEspecialidad" esp ON esp."PaisKey"=p."PaisKey" AND esp."Especialidad"=TRIM(s."Especialidad")
        JOIN "cat"."DimMedico" med ON med."PaisKey"=p."PaisKey" AND med."NombreMedico"=TRIM(s."Medico")
             AND COALESCE(med."EspecialidadKey",-1)=COALESCE(esp."EspecialidadKey",-1)
        LEFT JOIN "cat"."DimCentroMedico" cm ON cm."PaisKey"=p."PaisKey" AND cm."CentroMedico"=TRIM(s."CentroMedico")
        LEFT JOIN "cat"."DimGeografia" geo ON geo."PaisKey"=p."PaisKey" AND geo."Provincia"=TRIM(s."Provincia") AND geo."Municipio"=TRIM(s."Municipio")
        LEFT JOIN "cat"."DimRepresentanteMedico" rep ON rep."PaisKey"=p."PaisKey" AND rep."CodigoRepresentante"=s."CodigoRepresentante"
        WHERE s."LoadBatchKey"=:b"""), b).mappings().all()

    comps = db.execute(text('SELECT "ComponenteKey", "CodigoComponente", "Requerido" FROM "cat"."DimComponenteCategoria" WHERE "Activo"=TRUE')).mappings().all()
    reglas = db.execute(text("""SELECT "ReglaKey", "PaisKey", "ComponenteKey", "ValorMinimo", "ValorMaximo", "ValorTexto", "Criterio", "PuntajePct"
        FROM "cat"."DimReglaCategoriaMedica" WHERE "Activo"=TRUE AND "VigenteDesde"<=:fc AND ("VigenteHasta" IS NULL OR "VigenteHasta">=:fc)"""), fc).mappings().all()
    clases = db.execute(text("""SELECT "ClasificacionKey", "PaisKey", "Clase", "PuntajeMinPct", "PuntajeMaxPct"
        FROM "cat"."DimClasificacionMedica" WHERE "Activo"=TRUE AND "VigenteDesde"<=:fc AND ("VigenteHasta" IS NULL OR "VigenteHasta">=:fc)"""), fc).mappings().all()

    reglas_idx = defaultdict(list)
    for r in reglas:
        reglas_idx[(r["PaisKey"], r["ComponenteKey"])].append(r)
    clases_pais = defaultdict(list)
    for c in clases:
        clases_pais[c["PaisKey"]].append(c)
    n_req = sum(1 for c in comps if c["Requerido"])

    def _valor(row, cod):
        if cod in _COMP_NUM:
            return row[_COMP_NUM[cod]], None
        if cod in _COMP_TXT:
            v = row[_COMP_TXT[cod]]
            return None, (None if v is None else str(v))
        return None, None

    snap_params, detalle_por_row = [], {}
    for row in rows:
        total = 0
        calculados = 0
        detalles = []
        for c in comps:
            vnum, vtxt = _valor(row, c["CodigoComponente"])
            candidatas = [g for g in reglas_idx.get((row["PaisKey"], c["ComponenteKey"]), []) if _regla_aplica(g, vnum, vtxt)]
            mejor = _mejor_regla(candidatas)
            if c["Requerido"]:
                if mejor is not None:
                    total += mejor["PuntajePct"]
                    calculados += 1
            detalles.append({"ComponenteKey": c["ComponenteKey"], "ReglaKey": mejor["ReglaKey"] if mejor else None,
                             "ValorEntradaTexto": vtxt, "ValorEntradaNumero": vnum,
                             "Criterio": mejor["Criterio"] if mejor else None,
                             "PuntajePct": mejor["PuntajePct"] if mejor else None,
                             "EstadoComponente": "OK" if mejor else "SIN_REGLA"})
        clase = next((cl for cl in clases_pais.get(row["PaisKey"], []) if cl["PuntajeMinPct"] <= total <= cl["PuntajeMaxPct"]), None)
        cat_excel = row["CategoriaExcel"]
        if cat_excel is None or str(cat_excel).strip() == "":
            concil = "SIN_CATEGORIA_EXCEL"
        elif clase is not None and str(clase["Clase"]).strip().upper() == str(cat_excel).strip().upper():
            concil = "COINCIDE"
        else:
            concil = "DISCREPANCIA"
        estado = "CALCULADO" if calculados == n_req else "PENDIENTE"
        mensaje = None if calculados == n_req else f"Componentes calculados {calculados} de {n_req}"
        snap_params.append({
            "LoadBatchKey": load_batch_key, "RowNumber": row["RowNumber"], "Periodo": row["Periodo"],
            "PaisKey": row["PaisKey"], "MedicoKey": row["MedicoKey"], "CentroMedicoKey": row["CentroMedicoKey"],
            "GeografiaKey": row["GeografiaKey"], "RepresentanteKey": row["RepresentanteKey"], "Equipo": row["Equipo"],
            "Provincia": row["Provincia"], "Municipio": row["Municipio"], "LineaIdOrigen": row["LineaIdOrigen"],
            "PacientesSemana": row["PacientesSemana"], "CostoConsulta": row["CostoConsulta"], "RecetasSemana": row["RecetasSemana"],
            "UbicacionTerritorialCM": row["UbicacionTerritorialCM"], "KOL": row["KOL"], "PuntajeTotalPct": total,
            "ClasificacionKey": clase["ClasificacionKey"] if clase else None,
            "CategoriaCalculada": clase["Clase"] if clase else None, "CategoriaExcel": cat_excel,
            "EstadoConciliacion": concil, "EstadoCalculo": estado, "MensajeCalculo": mensaje})
        detalle_por_row[row["RowNumber"]] = detalles

    if snap_params:
        db.execute(text("""INSERT INTO "cat"."FactMedicoCategoriaSnapshot"
            ("LoadBatchKey","RowNumber","Periodo","PaisKey","MedicoKey","CentroMedicoKey","GeografiaKey","RepresentanteKey","Equipo",
             "Provincia","Municipio","LineaIdOrigen","PacientesSemana","CostoConsulta","RecetasSemana","UbicacionTerritorialCM","KOL",
             "PuntajeTotalPct","ClasificacionKey","CategoriaCalculada","CategoriaExcel","EstadoConciliacion","EstadoCalculo","MensajeCalculo")
            VALUES (:LoadBatchKey,:RowNumber,:Periodo,:PaisKey,:MedicoKey,:CentroMedicoKey,:GeografiaKey,:RepresentanteKey,:Equipo,
             :Provincia,:Municipio,:LineaIdOrigen,:PacientesSemana,:CostoConsulta,:RecetasSemana,:UbicacionTerritorialCM,:KOL,
             :PuntajeTotalPct,:ClasificacionKey,:CategoriaCalculada,:CategoriaExcel,:EstadoConciliacion,:EstadoCalculo,:MensajeCalculo)"""),
            snap_params)
        db.flush()
        # mapear RowNumber -> MedicoCategoriaKey generado
        keymap = dict(db.execute(text('SELECT "RowNumber", "MedicoCategoriaKey" FROM "cat"."FactMedicoCategoriaSnapshot" WHERE "LoadBatchKey"=:b'), b).all())
        det_params = []
        for rn, dets in detalle_por_row.items():
            mck = keymap.get(rn)
            for d in dets:
                det_params.append({**d, "MedicoCategoriaKey": mck})
        if det_params:
            db.execute(text("""INSERT INTO "cat"."FactMedicoCategoriaDetalle"
                ("MedicoCategoriaKey","ComponenteKey","ReglaKey","ValorEntradaTexto","ValorEntradaNumero","Criterio","PuntajePct","EstadoComponente")
                VALUES (:MedicoCategoriaKey,:ComponenteKey,:ReglaKey,:ValorEntradaTexto,:ValorEntradaNumero,:Criterio,:PuntajePct,:EstadoComponente)"""),
                det_params)

    db.execute(text("""UPDATE "cat"."LoadBatch" SET "Estado"='CALCULADO', "Mensaje"='Proceso de categorizacion ejecutado.' WHERE "LoadBatchKey"=:b"""), b)
    db.flush()
    logger.info(f"[CAT] calcular_categorias_py batch={load_batch_key}: {len(snap_params)} snapshots")


# ---------------------------------------------------------------------------
# Categorización de UN médico (alta desde el Panel) — Bloque B
# ---------------------------------------------------------------------------
# El alta de un médico en el Panel captura los mismos 5 criterios que el Excel, pero
# de a un médico. Para que un médico dé LA MISMA categoría venga del Excel o del Panel,
# aquí se reutilizan las MISMAS reglas (cat.DimReglaCategoriaMedica) y los MISMOS
# helpers de matcheo (_regla_aplica / _mejor_regla) que usa calcular_categorias_py.
# `_puntuar` es pura (sin BD) para poder probarla directamente.

# Los 5 criterios, con el nombre de campo que espera _COMP_NUM/_COMP_TXT.
CAMPOS_CRITERIOS = ("PacientesSemana", "CostoConsulta", "RecetasSemana",
                    "UbicacionTerritorialCM", "KOL")


def _puntuar(comps, reglas_idx, clases, valores: dict) -> dict:
    """Suma PuntajePct de la mejor regla por componente requerido y ubica la clase.
    Misma lógica que el bucle de calcular_categorias_py, sin tocar la BD.

    OJO CON LA ESCALA: pese al nombre `PuntajePct`, los valores son una FRACCIÓN 0-1
    (0.30 = 30%), y las bandas de DimClasificacionMedica están en la misma escala
    (p. ej. DO: A=0.86-1.00, B=0.66-0.85, C=0.46-0.65, D=0.00-0.45). El tope suma
    1.00 con los pesos 30/20/10/30/10. Para mostrarlo al usuario: multiplicar por 100."""
    total = 0
    calculados = 0
    n_req = sum(1 for c in comps if c["Requerido"])
    detalle = []
    for c in comps:
        cod = c["CodigoComponente"]
        if cod in _COMP_NUM:
            vnum, vtxt = valores.get(_COMP_NUM[cod]), None
        elif cod in _COMP_TXT:
            v = valores.get(_COMP_TXT[cod])
            vnum, vtxt = None, (None if v is None else str(v))
        else:
            vnum = vtxt = None
        candidatas = [g for g in reglas_idx.get(c["ComponenteKey"], []) if _regla_aplica(g, vnum, vtxt)]
        mejor = _mejor_regla(candidatas)
        if c["Requerido"] and mejor is not None:
            total += mejor["PuntajePct"]
            calculados += 1
        detalle.append({"componente": cod, "valor_num": vnum, "valor_txt": vtxt,
                        "puntaje_pct": mejor["PuntajePct"] if mejor else None,
                        "estado": "OK" if mejor else "SIN_REGLA"})
    clase = next((cl for cl in clases if cl["PuntajeMinPct"] <= total <= cl["PuntajeMaxPct"]), None)
    return {
        "puntaje_pct": total,
        "categoria": (clase["Clase"] if clase else None),
        "clasificacion_key": (clase["ClasificacionKey"] if clase else None),
        # PENDIENTE = falta alguna regla para un componente requerido (config incompleta),
        # NO se inventa categoría: el GD lo verá como no clasificable.
        "estado": "CALCULADO" if calculados == n_req else "PENDIENTE",
        "detalle": detalle,
    }


def calcular_categoria_de_valores(db: Session, pais_codigo: str, valores: dict) -> dict:
    """Categoría A/B/C/D de UN médico a partir de sus 5 valores crudos.

    `valores` usa las claves de CAMPOS_CRITERIOS. Devuelve
    {puntaje_pct, categoria, clasificacion_key, estado, detalle}. `categoria` es None
    si la configuración del país no cubre el caso (estado PENDIENTE)."""
    from collections import defaultdict
    from datetime import datetime, timezone

    fila = db.execute(text('SELECT "PaisKey" FROM "cat"."DimPais" WHERE "CodigoPais"=:p'),
                      {"p": pais_codigo}).first()
    if not fila:
        return {"puntaje_pct": 0, "categoria": None, "clasificacion_key": None,
                "estado": "PENDIENTE", "detalle": [], "motivo": f"País {pais_codigo} sin configuración de categorización"}
    pais_key = fila[0]
    fc = {"fc": datetime.now(timezone.utc).date(), "pk": pais_key}

    comps = db.execute(text('SELECT "ComponenteKey", "CodigoComponente", "Requerido" '
                            'FROM "cat"."DimComponenteCategoria" WHERE "Activo"=TRUE')).mappings().all()
    reglas = db.execute(text("""SELECT "ReglaKey", "ComponenteKey", "ValorMinimo", "ValorMaximo", "ValorTexto",
                                       "Criterio", "PuntajePct"
        FROM "cat"."DimReglaCategoriaMedica"
        WHERE "Activo"=TRUE AND "PaisKey"=:pk AND "VigenteDesde"<=:fc
          AND ("VigenteHasta" IS NULL OR "VigenteHasta">=:fc)"""), fc).mappings().all()
    clases = db.execute(text("""SELECT "ClasificacionKey", "Clase", "PuntajeMinPct", "PuntajeMaxPct"
        FROM "cat"."DimClasificacionMedica"
        WHERE "Activo"=TRUE AND "PaisKey"=:pk AND "VigenteDesde"<=:fc
          AND ("VigenteHasta" IS NULL OR "VigenteHasta">=:fc)"""), fc).mappings().all()

    reglas_idx = defaultdict(list)
    for r in reglas:
        reglas_idx[r["ComponenteKey"]].append(r)
    return _puntuar(comps, reglas_idx, clases, valores)


# Etiquetas legibles de los 5 criterios (el código es el de cat.DimComponenteCategoria).
_ETIQUETAS_CRITERIO = {
    "PACIENTES_SEMANA": "Pacientes por semana",
    "PODER_ADQUISITIVO": "Poder adquisitivo (costo de consulta)",
    "POTENCIAL_PRESCRIPCION": "Potencial de prescripción (recetas/semana)",
    "UBICACION_TERRITORIAL_CM": "Ubicación territorial",
    "KOL": "KOL (líder de opinión)",
}


def opciones_plantilla(db: Session, pais_codigo: str) -> list[dict]:
    """Opciones válidas de cada criterio para el formulario de alta de un médico.

    Se derivan de las REGLAS del país, porque cada criterio tiene un vocabulario cerrado
    y NO intuitivo (p. ej. POTENCIAL_PRESCRIPCION usa '10 o Mas'/'6 a 9'/…, KOL usa
    'Charlista Internacional'/'Ninguno'/…). Si el formulario dejara escribir libre, el
    valor no matchearía ninguna regla y el médico quedaría sin clasificar.

    NUNCA se devuelve el puntaje de cada opción: la categoría debe permanecer oculta
    hasta que el Gerente de Distrito aprueba (mismo criterio que `score_oculto` en LSII;
    si el rep viera los puntos podría capturar apuntando a la categoría que quiere)."""
    from datetime import datetime, timezone

    fila = db.execute(text('SELECT "PaisKey" FROM "cat"."DimPais" WHERE "CodigoPais"=:p'),
                      {"p": pais_codigo}).first()
    if not fila:
        return []
    fc = {"pk": fila[0], "fc": datetime.now(timezone.utc).date()}

    filas = db.execute(text("""
        SELECT c."CodigoComponente", c."Requerido", r."ValorMinimo", r."ValorMaximo",
               r."ValorTexto", r."PuntajePct"
        FROM "cat"."DimReglaCategoriaMedica" r
        JOIN "cat"."DimComponenteCategoria" c ON c."ComponenteKey" = r."ComponenteKey"
        WHERE r."Activo" = TRUE AND c."Activo" = TRUE AND r."PaisKey" = :pk
          AND r."VigenteDesde" <= :fc AND (r."VigenteHasta" IS NULL OR r."VigenteHasta" >= :fc)
        ORDER BY c."CodigoComponente", r."PuntajePct\""""), fc).all()

    por_codigo: dict[str, dict] = {}
    for codigo, requerido, vmin, vmax, vtxt, _pct in filas:
        campo = _COMP_NUM.get(codigo) or _COMP_TXT.get(codigo)
        if not campo:
            continue
        d = por_codigo.setdefault(codigo, {
            "codigo": codigo, "campo": campo,
            "etiqueta": _ETIQUETAS_CRITERIO.get(codigo, codigo),
            "tipo": "NUMERICO" if codigo in _COMP_NUM else "TEXTO",
            "requerido": bool(requerido), "opciones": [], "minimo": None, "maximo": None,
        })
        if d["tipo"] == "TEXTO":
            if vtxt and vtxt not in d["opciones"]:
                d["opciones"].append(vtxt)          # sin puntaje: solo el vocabulario válido
        else:
            if vmin is not None:
                d["minimo"] = vmin if d["minimo"] is None else min(d["minimo"], vmin)
            if vmax is not None:
                d["maximo"] = vmax if d["maximo"] is None else max(d["maximo"], vmax)
    return list(por_codigo.values())


def cargar_excel_categorizacion(db: Session, excel_bytes: bytes, nombre_archivo: str, periodo: str, usuario: str) -> dict:
    logger.info(f"[CAT] Iniciando carga '{nombre_archivo}' periodo={periodo}")
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True, read_only=False)
    try:
        _upsert_pais(db, _read_table(wb, "DIM_PAIS", "tbl_DimPais"))
        _upsert_componente(db, _read_table(wb, "DimComponente", "tbl_DimComponente"))
        _upsert_clasificacion(db, _read_table(wb, "DimClasificacion", "tbl_DimClasificacion"))
        _upsert_regla(db, _read_table(wb, "DimRegla", "tbl_DimRegla"))
        _upsert_equipo(db, _read_table(wb, "DimEquipo"))
        _upsert_representante(db, _read_table(wb, "DIM_REPRESENTANTE_MEDICO", "tbl_DimRepresentante"))

        # Limpiar datos previos del período antes de insertar (idempotente)
        lotes_eliminados = _limpiar_periodo(db, periodo)
        if lotes_eliminados:
            logger.info(f"[CAT] Reemplazando {lotes_eliminados} lote(s) existente(s) para período '{periodo}'")

        # Recupera la clave recién insertada con RETURNING (PostgreSQL).
        # Se hace el INSERT sin cláusula de salida y luego se relee la clave por MAX
        # sobre la clave natural (archivo+periodo+usuario). Es seguro porque
        # _limpiar_periodo() ya borró cualquier lote previo de este período.
        # 'Estado' es NOT NULL; el default del ORM (default="RECIBIDO") es client-side
        # y NO aplica a un INSERT por text() crudo, así que se pasa explícito.
        db.execute(text("""
            INSERT INTO "cat"."LoadBatch" ("ArchivoOrigen","Periodo","CodigoPaisDefault","UsuarioCarga","FechaCargaUtc","Estado")
            VALUES (:archivo,:periodo,'DO',:usuario,:ahora,'RECIBIDO')
        """), {"archivo": nombre_archivo, "periodo": periodo, "usuario": usuario,
               "ahora": datetime.now(timezone.utc)})
        db.flush()
        row = db.execute(text("""
            SELECT MAX("LoadBatchKey") FROM "cat"."LoadBatch"
            WHERE "ArchivoOrigen"=:archivo AND "Periodo"=:periodo AND "UsuarioCarga"=:usuario
        """), {"archivo": nombre_archivo, "periodo": periodo, "usuario": usuario}).fetchone()
        load_batch_key = int(row[0])

        fact_rows = _read_table(wb, "FactMedicoInput", "tbl_FactMedicoInput")
        _insert_staging(db, load_batch_key, fact_rows, periodo=periodo)

        calcular_categorias_py(db, load_batch_key)  # jul-2026: motor en Python (antes cat.sp_CalcularCategoriaMedica)
        db.commit()

        batch = db.execute(text('SELECT "Estado","Mensaje" FROM "cat"."LoadBatch" WHERE "LoadBatchKey"=:key'), {"key": load_batch_key}).fetchone()
        conteo = db.execute(text("""
            SELECT COUNT(*),
                SUM(CASE WHEN "CategoriaCalculada"='A' THEN 1 ELSE 0 END),
                SUM(CASE WHEN "CategoriaCalculada"='B' THEN 1 ELSE 0 END),
                SUM(CASE WHEN "CategoriaCalculada"='C' THEN 1 ELSE 0 END),
                SUM(CASE WHEN "CategoriaCalculada"='D' THEN 1 ELSE 0 END),
                SUM(CASE WHEN "CategoriaCalculada"="CategoriaExcel" THEN 1 ELSE 0 END),
                SUM(CASE WHEN "CategoriaExcel" IS NOT NULL AND "CategoriaCalculada"<>"CategoriaExcel" THEN 1 ELSE 0 END)
            FROM "cat"."FactMedicoCategoriaSnapshot" WHERE "LoadBatchKey"=:key
        """), {"key": load_batch_key}).fetchone()

        return {
            "load_batch_key": load_batch_key, "estado": batch[0] if batch else "CALCULADO",
            "mensaje": batch[1] if batch else None, "filas_cargadas": len(fact_rows),
            "total_medicos": int(conteo[0]) if conteo and conteo[0] else 0,
            "categoria_a": int(conteo[1]) if conteo and conteo[1] else 0,
            "categoria_b": int(conteo[2]) if conteo and conteo[2] else 0,
            "categoria_c": int(conteo[3]) if conteo and conteo[3] else 0,
            "categoria_d": int(conteo[4]) if conteo and conteo[4] else 0,
            "conciliacion_ok": int(conteo[5]) if conteo and conteo[5] else 0,
            "conciliacion_diferencia": int(conteo[6]) if conteo and conteo[6] else 0,
        }
    except Exception as e:
        db.rollback()
        logger.error(f"[CAT] Error: {e}")
        raise


# ── Consultas ──────────────────────────────────────────────────────────────────

def get_lotes(db: Session, limit: int = 50) -> list:
    rows = db.execute(text("""
        SELECT lb."LoadBatchKey", lb."ArchivoOrigen", lb."Periodo", lb."CodigoPaisDefault",
            lb."FechaCargaUtc", lb."UsuarioCarga", lb."Estado", lb."Mensaje",
            COUNT(f."MedicoCategoriaKey") AS "TotalMedicos"
        FROM "cat"."LoadBatch" lb
        LEFT JOIN "cat"."FactMedicoCategoriaSnapshot" f ON f."LoadBatchKey"=lb."LoadBatchKey"
        GROUP BY lb."LoadBatchKey",lb."ArchivoOrigen",lb."Periodo",lb."CodigoPaisDefault",lb."FechaCargaUtc",lb."UsuarioCarga",lb."Estado",lb."Mensaje"
        ORDER BY lb."LoadBatchKey" DESC
        OFFSET 0 ROWS FETCH FIRST :lim ROWS ONLY
    """), {"lim": limit}).mappings().all()
    return [dict(r) for r in rows]


def get_resultados(db: Session, periodo=None, categoria=None, estado_conciliacion=None, representante=None, medico=None, especialidad=None, codigos_rep=None, skip=0, limit=100) -> dict:
    where_parts = ["1=1"]
    params = {"skip": skip, "limit": limit}
    # Scope de datos (Item 5 Mallén): None = ve todos; lista = solo esos representantes;
    # lista vacía = no ve nada (RM/GD sin vínculo).
    if codigos_rep is not None:
        if codigos_rep:
            keys = []
            for i, c in enumerate(codigos_rep):
                k = f"cr{i}"; params[k] = c; keys.append(f":{k}")
            where_parts.append(f'v."CodigoRepresentante" IN ({", ".join(keys)})')
        else:
            where_parts.append("1=0")
    if periodo:
        where_parts.append('v."Periodo"=:periodo'); params["periodo"] = periodo
    if categoria:
        where_parts.append('v."CategoriaCalculada"=:cat'); params["cat"] = categoria
    if estado_conciliacion:
        where_parts.append('v."EstadoConciliacion"=:ec'); params["ec"] = estado_conciliacion
    if representante:
        where_parts.append('(LOWER(v."NombreRepresentante") LIKE LOWER(:rep) OR LOWER(v."CodigoRepresentante") LIKE LOWER(:rep))'); params["rep"] = f"%{representante}%"
    if medico:
        where_parts.append('LOWER(v."NombreMedico") LIKE LOWER(:med)'); params["med"] = f"%{medico}%"
    if especialidad:
        where_parts.append('COALESCE(v."Especialidad", \'\') = :esp'); params["esp"] = especialidad
    wc = " AND ".join(where_parts)
    total = int(db.execute(text(f'SELECT COUNT(*) FROM "cat"."vwMedicoCategoriaConciliacion" v WHERE {wc}'), params).fetchone()[0])
    rows = db.execute(text(f"""
        SELECT * FROM (
            SELECT v.*, ROW_NUMBER() OVER (ORDER BY v."Periodo" DESC, v."NombreMedico") AS rn
            FROM "cat"."vwMedicoCategoriaConciliacion" v WHERE {wc}
        ) sub WHERE rn > :skip AND rn <= :skip + :limit
    """), params).mappings().all()
    return {"total": total, "items": [dict(r) for r in rows]}


def get_resumen(
    db: Session,
    periodo: str = None,
    pais: str = None,
    provincia: str = None,
    municipio: str = None,
    especialidad: str = None,
    representante: str = None,
    codigos_rep=None,
) -> dict:
    params: dict = {}
    filters = ["1=1"]

    if periodo:
        filters.append('f."Periodo" = :periodo')
        params["periodo"] = periodo
    if pais:
        filters.append('EXISTS (SELECT 1 FROM "cat"."DimPais" _p WHERE _p."PaisKey"=f."PaisKey" AND _p."CodigoPais"=:pais)')
        params["pais"] = pais
    if provincia:
        filters.append('f."Provincia" = :provincia')
        params["provincia"] = provincia
    if municipio:
        filters.append('f."Municipio" = :municipio')
        params["municipio"] = municipio
    if especialidad:
        filters.append('COALESCE(e."Especialidad",\'\') = :especialidad')
        params["especialidad"] = especialidad
    if representante:
        filters.append('COALESCE(r."NombreRepresentante",\'\') = :representante')
        params["representante"] = representante
    # Scope de datos (Item 5 Mallén): RM ve solo su código; GD los de su equipo.
    if codigos_rep is not None:
        if codigos_rep:
            keys = []
            for i, c in enumerate(codigos_rep):
                k = f"cr{i}"; params[k] = c; keys.append(f":{k}")
            filters.append(f'r."CodigoRepresentante" IN ({", ".join(keys)})')
        else:
            filters.append("1=0")

    where = "WHERE " + " AND ".join(filters)

    # Solo añadir JOINs si se filtran columnas de tablas relacionadas
    join_clause = ""
    if especialidad or representante or codigos_rep is not None:
        join_clause = """
        LEFT JOIN "cat"."DimMedico" m ON m."MedicoKey" = f."MedicoKey"
        LEFT JOIN "cat"."DimEspecialidad" e ON e."EspecialidadKey" = m."EspecialidadKey"
        LEFT JOIN "cat"."DimRepresentanteMedico" r ON r."RepresentanteKey" = f."RepresentanteKey"
        """

    row = db.execute(text(f"""
        SELECT COUNT(*),
            SUM(CASE WHEN f."CategoriaCalculada"='A' THEN 1 ELSE 0 END),
            SUM(CASE WHEN f."CategoriaCalculada"='B' THEN 1 ELSE 0 END),
            SUM(CASE WHEN f."CategoriaCalculada"='C' THEN 1 ELSE 0 END),
            SUM(CASE WHEN f."CategoriaCalculada"='D' THEN 1 ELSE 0 END),
            SUM(CASE WHEN f."CategoriaCalculada" IS NULL THEN 1 ELSE 0 END),
            SUM(CASE WHEN f."CategoriaCalculada"=f."CategoriaExcel" THEN 1 ELSE 0 END),
            SUM(CASE WHEN f."CategoriaExcel" IS NOT NULL AND f."CategoriaCalculada"<>f."CategoriaExcel" THEN 1 ELSE 0 END),
            AVG(f."PuntajeTotalPct")
        FROM "cat"."FactMedicoCategoriaSnapshot" f {join_clause} {where}
    """), params).fetchone()
    periodos = db.execute(text('SELECT DISTINCT "Periodo" FROM "cat"."FactMedicoCategoriaSnapshot" ORDER BY "Periodo" DESC')).scalars().all()
    return {
        "total_medicos": int(row[0]) if row and row[0] else 0,
        "categoria_a": int(row[1]) if row and row[1] else 0,
        "categoria_b": int(row[2]) if row and row[2] else 0,
        "categoria_c": int(row[3]) if row and row[3] else 0,
        "categoria_d": int(row[4]) if row and row[4] else 0,
        "sin_categoria": int(row[5]) if row and row[5] else 0,
        "conciliacion_ok": int(row[6]) if row and row[6] else 0,
        "conciliacion_diferencia": int(row[7]) if row and row[7] else 0,
        "puntaje_promedio": round(float(row[8]) * 100, 2) if row and row[8] else 0.0,
        "periodos_disponibles": list(periodos),
    }


def get_por_representante(db: Session, periodo=None) -> list:
    params: dict = {}
    where = "WHERE 1=1"
    if periodo:
        where += ' AND f."Periodo"=:periodo'
        params["periodo"] = periodo
    rows = db.execute(text(f"""
        SELECT
            COALESCE(f."Equipo",'Sin Equipo')            AS equipo,
            COALESCE(r."NombreRepresentante",'Sin Rep')  AS representante,
            SUM(CASE WHEN f."CategoriaCalculada"='A' THEN 1 ELSE 0 END) AS a,
            SUM(CASE WHEN f."CategoriaCalculada"='B' THEN 1 ELSE 0 END) AS b,
            SUM(CASE WHEN f."CategoriaCalculada"='C' THEN 1 ELSE 0 END) AS c,
            SUM(CASE WHEN f."CategoriaCalculada"='D' THEN 1 ELSE 0 END) AS d,
            COUNT(*) AS total
        FROM "cat"."FactMedicoCategoriaSnapshot" f
        LEFT JOIN "cat"."DimRepresentanteMedico" r ON r."RepresentanteKey" = f."RepresentanteKey"
        {where}
        GROUP BY COALESCE(f."Equipo",'Sin Equipo'), COALESCE(r."NombreRepresentante",'Sin Rep')
        ORDER BY equipo, representante
    """), params).mappings().all()
    return [dict(r) for r in rows]


def get_por_especialidad(db: Session, periodo=None) -> list:
    params: dict = {}
    where = "WHERE 1=1"
    if periodo:
        where += ' AND f."Periodo"=:periodo'
        params["periodo"] = periodo
    rows = db.execute(text(f"""
        SELECT
            COALESCE(e."Especialidad",'Sin Especialidad') AS especialidad,
            SUM(CASE WHEN f."CategoriaCalculada"='A' THEN 1 ELSE 0 END) AS a,
            SUM(CASE WHEN f."CategoriaCalculada"='B' THEN 1 ELSE 0 END) AS b,
            SUM(CASE WHEN f."CategoriaCalculada"='C' THEN 1 ELSE 0 END) AS c,
            SUM(CASE WHEN f."CategoriaCalculada"='D' THEN 1 ELSE 0 END) AS d,
            COUNT(*) AS total
        FROM "cat"."FactMedicoCategoriaSnapshot" f
        JOIN "cat"."DimMedico" m ON m."MedicoKey" = f."MedicoKey"
        LEFT JOIN "cat"."DimEspecialidad" e ON e."EspecialidadKey" = m."EspecialidadKey"
        {where}
        GROUP BY COALESCE(e."Especialidad",'Sin Especialidad')
        ORDER BY total DESC
    """), params).mappings().all()
    return [dict(r) for r in rows]


# ── Filtros en cascada para la pestaña Detalle Médicos ────────────────────────

def get_filtros_medicos(db: Session, pais: str = None, provincia: str = None) -> dict:
    params: dict = {}

    pais_where = ""
    if pais:
        pais_where = 'JOIN "cat"."DimPais" _p ON _p."PaisKey" = f."PaisKey" AND _p."CodigoPais" = :pais'
        params["pais"] = pais

    prov_where = pais_where
    if provincia:
        prov_where_extra = ' AND f."Provincia" = :provincia'
        params["provincia"] = provincia
    else:
        prov_where_extra = ""

    paises = db.execute(text(
        'SELECT "CodigoPais", "NombrePais" FROM "cat"."DimPais" WHERE "Activo"=TRUE ORDER BY "CodigoPais"'
    )).mappings().all()

    provincias = db.execute(text(f"""
        SELECT DISTINCT f."Provincia"
        FROM "cat"."FactMedicoCategoriaSnapshot" f {pais_where}
        WHERE f."Provincia" IS NOT NULL
        ORDER BY f."Provincia"
    """), params).scalars().all()

    municipios = db.execute(text(f"""
        SELECT DISTINCT f."Municipio"
        FROM "cat"."FactMedicoCategoriaSnapshot" f {pais_where}
        WHERE f."Municipio" IS NOT NULL {prov_where_extra}
        ORDER BY f."Municipio"
    """), params).scalars().all()

    esp_where = ""
    if pais:
        esp_where = 'JOIN "cat"."DimPais" _p2 ON _p2."PaisKey" = e."PaisKey" AND _p2."CodigoPais" = :pais'
    especialidades = db.execute(text(f"""
        SELECT DISTINCT e."Especialidad"
        FROM "cat"."DimEspecialidad" e {esp_where}
        WHERE e."Activo"=TRUE ORDER BY e."Especialidad"
    """), params).scalars().all()

    rep_where = ""
    if pais:
        rep_where = 'JOIN "cat"."DimPais" _p3 ON _p3."PaisKey" = r."PaisKey" AND _p3."CodigoPais" = :pais'
    representantes = db.execute(text(f"""
        SELECT DISTINCT r."NombreRepresentante", r."CodigoRepresentante"
        FROM "cat"."DimRepresentanteMedico" r {rep_where}
        WHERE r."Activo"=TRUE ORDER BY r."NombreRepresentante"
    """), params).mappings().all()

    return {
        "paises": [dict(r) for r in paises],
        "provincias": list(provincias),
        "municipios": list(municipios),
        "especialidades": list(especialidades),
        "representantes": [dict(r) for r in representantes],
    }


def get_detalle_medicos(
    db: Session,
    periodo: str = None,
    pais: str = None,
    provincia: str = None,
    municipio: str = None,
    especialidad: str = None,
    representante: str = None,
    categoria: str = None,
    skip: int = 0,
    limit: int = 50,
) -> dict:
    params: dict = {"skip": skip, "limit": limit}
    filters = ["1=1"]

    if periodo:
        filters.append('f."Periodo" = :periodo')
        params["periodo"] = periodo
    if pais:
        filters.append('EXISTS (SELECT 1 FROM "cat"."DimPais" _p WHERE _p."PaisKey"=f."PaisKey" AND _p."CodigoPais"=:pais)')
        params["pais"] = pais
    if provincia:
        filters.append('f."Provincia" = :provincia')
        params["provincia"] = provincia
    if municipio:
        filters.append('f."Municipio" = :municipio')
        params["municipio"] = municipio
    if especialidad:
        filters.append('COALESCE(e."Especialidad",\'\') = :especialidad')
        params["especialidad"] = especialidad
    if representante:
        filters.append('COALESCE(r."NombreRepresentante",\'\') = :representante')
        params["representante"] = representante
    if categoria:
        filters.append('f."CategoriaCalculada" = :categoria')
        params["categoria"] = categoria

    where = "WHERE " + " AND ".join(filters)

    total = db.execute(text(f"""
        SELECT COUNT(*)
        FROM "cat"."FactMedicoCategoriaSnapshot" f
        LEFT JOIN "cat"."DimMedico" m ON m."MedicoKey" = f."MedicoKey"
        LEFT JOIN "cat"."DimEspecialidad" e ON e."EspecialidadKey" = m."EspecialidadKey"
        LEFT JOIN "cat"."DimRepresentanteMedico" r ON r."RepresentanteKey" = f."RepresentanteKey"
        {where}
    """), params).scalar() or 0

    rows = db.execute(text(f"""
        SELECT
            f."MedicoCategoriaKey",
            m."NombreMedico",
            COALESCE(e."Especialidad",'—')          AS "Especialidad",
            COALESCE(f."Provincia",'—')              AS "Provincia",
            COALESCE(f."Municipio",'—')              AS "Municipio",
            COALESCE(f."Equipo",'—')                 AS "Equipo",
            COALESCE(r."NombreRepresentante",'—')    AS "Representante",
            f."Periodo",
            ROUND(COALESCE(f."PuntajeTotalPct",0)*100, 1) AS "PuntajeTotalPct",
            COALESCE(f."CategoriaCalculada",'?')     AS "CategoriaCalculada",
            COALESCE(f."CategoriaExcel",'—')         AS "CategoriaExcel",
            COALESCE(f."EstadoConciliacion",'—')     AS "EstadoConciliacion"
        FROM "cat"."FactMedicoCategoriaSnapshot" f
        LEFT JOIN "cat"."DimMedico" m ON m."MedicoKey" = f."MedicoKey"
        LEFT JOIN "cat"."DimEspecialidad" e ON e."EspecialidadKey" = m."EspecialidadKey"
        LEFT JOIN "cat"."DimRepresentanteMedico" r ON r."RepresentanteKey" = f."RepresentanteKey"
        {where}
        ORDER BY f."CategoriaCalculada", m."NombreMedico"
        OFFSET :skip ROWS FETCH NEXT :limit ROWS ONLY
    """), params).mappings().all()

    return {"total": total, "items": [dict(r) for r in rows]}


def get_componentes_medico(db: Session, medico_categoria_key: int) -> list:
    rows = db.execute(text("""
        SELECT
            c."NombreComponente",
            c."CodigoComponente",
            d."ValorEntradaTexto",
            d."ValorEntradaNumero",
            ROUND(COALESCE(d."PuntajePct",0)*100, 1) AS "PuntajePct",
            d."EstadoComponente"
        FROM "cat"."FactMedicoCategoriaDetalle" d
        JOIN "cat"."DimComponenteCategoria" c ON c."ComponenteKey" = d."ComponenteKey"
        WHERE d."MedicoCategoriaKey" = :key
        ORDER BY c."CodigoComponente"
    """), {"key": medico_categoria_key}).mappings().all()
    return [dict(r) for r in rows]


# ── Mantenimiento DIMs de clasificación / reglas ──────────────────────────────

def get_clasificaciones(db: Session) -> list:
    rows = db.execute(text("""
        SELECT c."ClasificacionKey", p."CodigoPais", c."Clase",
               c."PuntajeMinPct", c."PuntajeMaxPct", c."Activo",
               c."VigenteDesde", c."VigenteHasta"
        FROM "cat"."DimClasificacionMedica" c
        JOIN "cat"."DimPais" p ON p."PaisKey" = c."PaisKey"
        ORDER BY c."Clase" DESC
    """)).mappings().all()
    return [dict(r) for r in rows]


def upsert_clasificacion(db: Session, data: dict) -> dict:
    key = data.get("ClasificacionKey")
    if key:
        db.execute(text("""
            UPDATE "cat"."DimClasificacionMedica"
            SET "PuntajeMinPct"=:min, "PuntajeMaxPct"=:max, "Activo"=:activo,
                "VigenteDesde"=:desde, "VigenteHasta"=:hasta
            WHERE "ClasificacionKey"=:key
        """), {"key": key, "min": data["PuntajeMinPct"], "max": data["PuntajeMaxPct"],
               "activo": data.get("Activo", True),
               "desde": data.get("VigenteDesde"), "hasta": data.get("VigenteHasta")})
    else:
        pais_key = db.execute(
            text('SELECT "PaisKey" FROM "cat"."DimPais" WHERE "CodigoPais"=:c'), {"c": data["CodigoPais"]}
        ).scalar()
        db.execute(text("""
            INSERT INTO "cat"."DimClasificacionMedica"
              ("PaisKey", "Clase", "PuntajeMinPct", "PuntajeMaxPct", "Activo", "VigenteDesde", "VigenteHasta")
            VALUES (:pk, :clase, :min, :max, :activo, :desde, :hasta)
        """), {"pk": pais_key, "clase": data["Clase"],
               "min": data["PuntajeMinPct"], "max": data["PuntajeMaxPct"],
               "activo": data.get("Activo", True),
               "desde": data.get("VigenteDesde"), "hasta": data.get("VigenteHasta")})
    db.commit()
    return {"ok": True}


def update_componente(db: Session, componente_key: int, peso_pct: float) -> dict:
    db.execute(text("""
        UPDATE "cat"."DimComponenteCategoria"
        SET "PesoComponentePct"=:peso WHERE "ComponenteKey"=:key
    """), {"key": componente_key, "peso": peso_pct / 100})
    db.commit()
    return {"ok": True}


def get_componentes(db: Session) -> list:
    rows = db.execute(text("""
        SELECT "ComponenteKey", "CodigoComponente", "NombreComponente", "TipoEvaluacion",
               "PesoComponentePct", ROUND("PesoComponentePct"*100,1) AS "PesoPct", "Requerido", "Activo"
        FROM "cat"."DimComponenteCategoria" ORDER BY "CodigoComponente"
    """)).mappings().all()
    return [dict(r) for r in rows]


def get_reglas(db: Session, componente_codigo: str = None) -> list:
    where = ""
    params: dict = {}
    if componente_codigo:
        where = 'WHERE c."CodigoComponente"=:cod'
        params["cod"] = componente_codigo
    rows = db.execute(text(f"""
        SELECT r."ReglaKey", c."CodigoComponente", c."NombreComponente",
               r."CodigoRegla", r."ValorMinimo", r."ValorMaximo", r."ValorTexto",
               r."Criterio", r."PuntajePct",
               ROUND(r."PuntajePct"*100, 2) AS "PuntajePctDisplay",
               r."Activo", r."VigenteDesde", r."VigenteHasta",
               p."CodigoPais"
        FROM "cat"."DimReglaCategoriaMedica" r
        JOIN "cat"."DimComponenteCategoria" c ON c."ComponenteKey" = r."ComponenteKey"
        JOIN "cat"."DimPais" p ON p."PaisKey" = r."PaisKey"
        {where}
        ORDER BY c."CodigoComponente", r."Criterio" DESC
    """), params).mappings().all()
    return [dict(r) for r in rows]


def upsert_regla(db: Session, data: dict) -> dict:
    key = data.get("ReglaKey")
    if key:
        db.execute(text("""
            UPDATE "cat"."DimReglaCategoriaMedica"
            SET "ValorMinimo"=:min, "ValorMaximo"=:max, "ValorTexto"=:texto,
                "Criterio"=:criterio, "PuntajePct"=:puntaje, "Activo"=:activo,
                "VigenteDesde"=:desde, "VigenteHasta"=:hasta
            WHERE "ReglaKey"=:key
        """), {"key": key, "min": data.get("ValorMinimo"), "max": data.get("ValorMaximo"),
               "texto": data.get("ValorTexto"), "criterio": data["Criterio"],
               "puntaje": float(data["PuntajePct"]) / 100,
               "activo": data.get("Activo", True),
               "desde": data.get("VigenteDesde"), "hasta": data.get("VigenteHasta")})
    else:
        comp_key = db.execute(
            text('SELECT "ComponenteKey" FROM "cat"."DimComponenteCategoria" WHERE "CodigoComponente"=:c'),
            {"c": data["CodigoComponente"]}
        ).scalar()
        pais_key = db.execute(
            text('SELECT "PaisKey" FROM "cat"."DimPais" WHERE "CodigoPais"=:c'), {"c": data["CodigoPais"]}
        ).scalar()
        db.execute(text("""
            INSERT INTO "cat"."DimReglaCategoriaMedica"
              ("PaisKey", "ComponenteKey", "CodigoRegla", "ValorMinimo", "ValorMaximo", "ValorTexto",
               "Criterio", "PuntajePct", "Activo", "VigenteDesde", "VigenteHasta")
            VALUES (:pk, :comp, :cod, :min, :max, :texto, :criterio, :puntaje, :activo, :desde, :hasta)
        """), {"pk": pais_key, "comp": comp_key, "cod": data.get("CodigoRegla", ""),
               "min": data.get("ValorMinimo"), "max": data.get("ValorMaximo"),
               "texto": data.get("ValorTexto"), "criterio": data["Criterio"],
               "puntaje": float(data["PuntajePct"]) / 100,
               "activo": data.get("Activo", True),
               "desde": data.get("VigenteDesde"), "hasta": data.get("VigenteHasta")})
    db.commit()
    return {"ok": True}


def delete_regla(db: Session, regla_key: int) -> dict:
    db.execute(text('UPDATE "cat"."DimReglaCategoriaMedica" SET "Activo"=FALSE WHERE "ReglaKey"=:key'),
               {"key": regla_key})
    db.commit()
    return {"ok": True}


def get_paises_cat(db: Session) -> list:
    rows = db.execute(text('SELECT "CodigoPais" FROM "cat"."DimPais" ORDER BY "CodigoPais"')).scalars().all()
    return list(rows)


# ── Función de resolución de médico por código (Cobertura Predictiva/4DX) ─────
# Exportada porque dims.py la importa directamente.

def resolver_medico_por_codigo(
    db: Session,
    pais_codigo: str,
    *,
    codigo: str,
    nombre: str,
    especialidad_nombre: str | None = None,
    activo: bool = True,
) -> int:
    """Upsert en Config.DIM_Medico por (pais_codigo, codigo) — flujo Cobertura Predictiva.

    Distinto del dedup por (pais_codigo, nombre, centro_medico_id) que usa
    el flujo de Categorización Médica. Ver Medico.__doc__ en dimensiones.py.

    Retorna el id del registro (existente o recién creado).
    """
    # Intentar encontrar por (pais_codigo, codigo) con la llave filtrada
    row = db.execute(
        text("""
            SELECT id FROM "Config"."DIM_Medico"
            WHERE pais_codigo = :pc AND codigo = :cod
        """),
        {"pc": pais_codigo, "cod": codigo},
    ).first()

    if row:
        medico_id = row[0]
        db.execute(
            text("""
                UPDATE "Config"."DIM_Medico"
                SET nombre = :nombre, activo = :activo
                WHERE id = :id
            """),
            {"nombre": nombre, "activo": 1 if activo else 0, "id": medico_id},
        )
        return medico_id

    # No existe: insertar. centro_medico_id queda NULL (flujo Cobertura).
    # especialidad_nombre se ignora si no hay una DIM_Especialidad coincidente.
    especialidad_id = None
    if especialidad_nombre:
        esp_row = db.execute(
            text("""
                SELECT e.id FROM "Config"."DIM_Especialidad" e
                JOIN "Config"."DIM_Pais" p ON p.id = e.pais_id
                WHERE p.codigo = :pc AND e.nombre = :esp
            """),
            {"pc": pais_codigo, "esp": especialidad_nombre.strip()},
        ).first()
        if esp_row:
            especialidad_id = esp_row[0]

    db.execute(
        text("""
            INSERT INTO "Config"."DIM_Medico"
                (pais_codigo, codigo, nombre, activo, especialidad_id, centro_medico_id)
            VALUES
                (:pc, :cod, :nombre, :activo, :esp_id, NULL)
        """),
        {
            "pc": pais_codigo,
            "cod": codigo,
            "nombre": nombre,
            "activo": 1 if activo else 0,
            "esp_id": especialidad_id,
        },
    )
    # Se relee la clave
    # recién insertada por la clave natural (pais_codigo, codigo), portable en ambos motores.
    db.flush()
    new_id = db.execute(
        text("""
            SELECT id FROM "Config"."DIM_Medico"
            WHERE pais_codigo = :pc AND codigo = :cod
        """),
        {"pc": pais_codigo, "cod": codigo},
    ).scalar()
    return int(new_id) if new_id else 0


# ── Categorización del PANEL de un representante ──────────────────────────────

def get_panel_rm(db: Session, rm_id: int) -> dict:
    """Categorización del panel de UN representante (plano de asignación).

    Distinta de `get_resumen`, que resume el MOTOR de 5 criterios (`cat.*`, por período y
    alimentado por el Excel del administrador). Esta lee `Visita.DIM_MedicoVisita`: la
    categoría que el propio representante fijó para el médico DENTRO de su panel — el
    sistema le propone la calculada al darlo de alta y él la ajusta a su criterio.

    **No recibe ciclo, y es a propósito**: la categoría del panel es un dato estable entre
    ciclos (regla de negocio: solo cambia con una actualización manual o una carga masiva,
    2-3 veces al año). Sobre este panel el rep programa sus visitas y revisitas.

    Solo cuenta médicos activos: los dados de baja ya no se programan.
    """
    from app.models.dimensiones import Especialidad
    from app.models.visita import MedicoVisita

    esp_nom = dict(db.query(Especialidad.id, Especialidad.nombre).all())
    medicos = (db.query(MedicoVisita)
               .filter(MedicoVisita.vm_id == rm_id, MedicoVisita.activo == True)  # noqa: E712
               .order_by(MedicoVisita.categoria, MedicoVisita.nombre_completo).all())

    conteo = {"A": 0, "B": 0, "C": 0, "D": 0}
    sin_categoria = 0
    filas = []
    for m in medicos:
        cat = (m.categoria or "").strip().upper()
        if cat in conteo:
            conteo[cat] += 1
        else:
            sin_categoria += 1  # dato viejo o fuera de A/B/C/D: se reporta, no se inventa
        filas.append({
            "id": m.id,
            "nombre": m.nombre_completo,
            "especialidad": esp_nom.get(m.especialidad_id),
            "categoria": cat if cat in conteo else None,
            "centro_trabajo": m.centro_trabajo,
            "frecuencia_visita": m.frecuencia_visita,
            "estado_aprobacion": m.estado_aprobacion,
        })

    return {
        "total_medicos": len(filas),
        "categoria_a": conteo["A"], "categoria_b": conteo["B"],
        "categoria_c": conteo["C"], "categoria_d": conteo["D"],
        "sin_categoria": sin_categoria,
        "medicos": filas,
    }
