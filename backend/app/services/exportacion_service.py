"""
SCGCPR — Servicio de Exportación de Reportes (PDF / Excel)
CLAUDE.md §18 "Pendiente de Implementar": "Exportación PDF/Excel de reportes
(reportes exportables, ReportLab ya instalado)".

Genera reportes EN MEMORIA (BytesIO) — no se persisten en disco como los
certificados de `reconocimiento_service` — y se devuelven directamente como
StreamingResponse desde el router. Esto evita acumular archivos huérfanos en
`settings.REPORTS_DIR` para reportes que el usuario solo quiere descargar una
vez (a diferencia del certificado, que sí se reutiliza/cachea por diseño).

Reutiliza el mismo estilo visual ya establecido en
`reconocimiento_service.generar_certificado_pdf` (paleta de colores
"#1F4E79"/"#2E75B6", ReportLab Platypus con SimpleDocTemplate + Paragraph +
Table) para mantener identidad visual consistente entre certificados y
reportes.

Excel: usa openpyxl (ya en requirements.txt) con encabezados estilizados
(fondo azul corporativo, texto blanco en negrita) y autoancho de columnas.

Las funciones reciben `db: Session` (servicios = lógica pura, sin HTTP, según
convención CLAUDE.md §19) y devuelven `io.BytesIO` listo para enviar.
"""
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Optional

from loguru import logger
from sqlalchemy.orm import Session

from app.models.dimensiones import RepresentanteMedico, Pais, Ciclo, Premio
from app.models.hechos import RankingRM, ReconocimientoRM

# Paleta corporativa — coincide con reconocimiento_service.generar_certificado_pdf
_COLOR_TITULO = "#1F4E79"
_COLOR_ACENTO = "#2E75B6"


# ---------------------------------------------------------------------------
# Helpers comunes
# ---------------------------------------------------------------------------

def _nombre_pais(db: Session, pais_codigo: Optional[str]) -> str:
    if not pais_codigo:
        return "Todos los países"
    pais = db.query(Pais).filter(Pais.id == pais_codigo).first()
    return pais.nombre if pais else f"País #{pais_codigo}"


def _nombre_ciclo(db: Session, ciclo_id: Optional[int]) -> str:
    if not ciclo_id:
        return "Todos los ciclos"
    ciclo = db.query(Ciclo).filter(Ciclo.id == ciclo_id).first()
    return ciclo.nombre_canonico or ciclo.nombre if ciclo else f"Ciclo #{ciclo_id}"


def _query_ranking(db: Session, pais_codigo: Optional[str], ciclo_id: Optional[int], tipo_ranking: str):
    q = (
        db.query(
            RankingRM,
            RepresentanteMedico.nombre.label("rm_nombre"),
            RepresentanteMedico.codigo.label("rm_codigo"),
        )
        .join(RepresentanteMedico, RepresentanteMedico.id == RankingRM.rm_id)
        .filter(RankingRM.tipo_ranking == tipo_ranking.upper())
    )
    if pais_codigo:
        q = q.filter(RankingRM.pais_codigo == pais_codigo)
    if ciclo_id:
        q = q.filter(RankingRM.ciclo_id == ciclo_id)
    return q.order_by(RankingRM.posicion_global.asc()).all()


def _query_reconocimientos(db: Session, pais_codigo: Optional[str], ciclo_id: Optional[int]):
    q = (
        db.query(
            ReconocimientoRM,
            RepresentanteMedico.nombre.label("rm_nombre"),
            RepresentanteMedico.codigo.label("rm_codigo"),
            Premio.nombre.label("premio_nombre"),
        )
        .join(RepresentanteMedico, RepresentanteMedico.id == ReconocimientoRM.rm_id)
        .join(Premio, Premio.id == ReconocimientoRM.premio_id)
    )
    if pais_codigo:
        q = q.filter(ReconocimientoRM.pais_codigo == pais_codigo)
    if ciclo_id:
        q = q.filter(ReconocimientoRM.ciclo_id == ciclo_id)
    return q.order_by(ReconocimientoRM.fecha_calculo.desc()).all()


# ---------------------------------------------------------------------------
# Excel — openpyxl
# ---------------------------------------------------------------------------

def _construir_workbook(titulo_hoja: str, encabezados: list[str], filas: list[list]) -> BytesIO:
    """
    Arma un Workbook de una sola hoja con encabezados estilizados
    (fondo azul corporativo, texto blanco) y autoancho de columnas.
    Devuelve un BytesIO listo para StreamingResponse.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = titulo_hoja[:31]  # límite de Excel para nombres de hoja

    fill_encabezado = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_encabezado = Font(color="FFFFFF", bold=True)
    align_centro = Alignment(horizontal="center", vertical="center")

    ws.append(encabezados)
    for celda in ws[1]:
        celda.fill = fill_encabezado
        celda.font = font_encabezado
        celda.alignment = align_centro

    for fila in filas:
        ws.append(fila)

    # Autoancho aproximado según el contenido más largo por columna
    for idx, encabezado in enumerate(encabezados, start=1):
        col_letra = get_column_letter(idx)
        max_len = len(str(encabezado))
        for fila in filas:
            valor = fila[idx - 1]
            max_len = max(max_len, len(str(valor)) if valor is not None else 0)
        ws.column_dimensions[col_letra].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_ranking_excel(
    db: Session,
    pais_codigo: Optional[str] = None,
    ciclo_id: Optional[int] = None,
    tipo_ranking: str = "MENSUAL",
) -> BytesIO:
    """Exporta el ranking filtrado a un libro Excel (.xlsx) en memoria."""
    rows = _query_ranking(db, pais_codigo, ciclo_id, tipo_ranking)

    encabezados = [
        "Posición", "Posición anterior", "Variación", "Código RM", "Nombre RM",
        "Score total", "Elegible", "Tipo de ranking", "Fecha de generación",
    ]
    filas = []
    for r in rows:
        rk = r.RankingRM
        variacion = (rk.posicion_anterior or rk.posicion_global) - rk.posicion_global
        filas.append([
            rk.posicion_global,
            rk.posicion_anterior,
            variacion,
            r.rm_codigo,
            r.rm_nombre,
            float(rk.score_total or 0),
            "Sí" if rk.elegible else "No",
            rk.tipo_ranking,
            rk.fecha_generacion.strftime("%Y-%m-%d %H:%M") if rk.fecha_generacion else "",
        ])

    logger.info(f"Exportando ranking a Excel — {len(filas)} filas, tipo={tipo_ranking}, pais_codigo={pais_codigo}, ciclo_id={ciclo_id}")
    return _construir_workbook(f"Ranking {tipo_ranking.upper()}", encabezados, filas)


def exportar_reconocimientos_excel(
    db: Session,
    pais_codigo: Optional[str] = None,
    ciclo_id: Optional[int] = None,
) -> BytesIO:
    """Exporta los reconocimientos otorgados a un libro Excel (.xlsx) en memoria."""
    rows = _query_reconocimientos(db, pais_codigo, ciclo_id)

    encabezados = [
        "Código RM", "Nombre RM", "Premio", "Score al momento", "Posición en ranking",
        "Elegible", "Certificado generado", "Fecha de cálculo",
    ]
    filas = []
    for r in rows:
        rec = r.ReconocimientoRM
        filas.append([
            r.rm_codigo,
            r.rm_nombre,
            r.premio_nombre,
            float(rec.score_total or 0),
            rec.posicion_ranking,
            "Sí" if rec.elegible else "No",
            "Sí" if rec.certificado_generado else "No",
            rec.fecha_calculo.strftime("%Y-%m-%d %H:%M") if rec.fecha_calculo else "",
        ])

    logger.info(f"Exportando reconocimientos a Excel — {len(filas)} filas, pais_codigo={pais_codigo}, ciclo_id={ciclo_id}")
    return _construir_workbook("Reconocimientos", encabezados, filas)


# ---------------------------------------------------------------------------
# PDF — ReportLab (mismo estilo visual que generar_certificado_pdf)
# ---------------------------------------------------------------------------

def exportar_ranking_pdf(
    db: Session,
    pais_codigo: Optional[str] = None,
    ciclo_id: Optional[int] = None,
    tipo_ranking: str = "MENSUAL",
    top: Optional[int] = None,
) -> BytesIO:
    """
    Exporta el ranking filtrado a un PDF tabular en memoria (orientación
    horizontal — la tabla tiene varias columnas). Reutiliza la paleta y los
    estilos de párrafo establecidos en `generar_certificado_pdf`.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    rows = _query_ranking(db, pais_codigo, ciclo_id, tipo_ranking)
    if top:
        rows = rows[:top]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle("titulo", parent=styles["Title"],
                                  fontSize=20, textColor=colors.HexColor(_COLOR_TITULO),
                                  alignment=TA_CENTER, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=11, alignment=TA_CENTER, spaceAfter=14,
                               textColor=colors.HexColor(_COLOR_ACENTO))
    nota_style = ParagraphStyle("nota", parent=styles["Normal"],
                                fontSize=8, alignment=TA_LEFT, textColor=colors.grey)

    encabezados = ["Pos.", "Cód.", "Representante", "Score", "Categ.", "Elegible", "Variación"]
    datos = [encabezados]
    for r in rows:
        rk = r.RankingRM
        variacion = (rk.posicion_anterior or rk.posicion_global) - rk.posicion_global
        signo = "+" if variacion > 0 else ""
        datos.append([
            str(rk.posicion_global),
            r.rm_codigo,
            r.rm_nombre,
            f"{float(rk.score_total or 0):.2f}",
            str(rk.categoria_id or "-"),
            "Sí" if rk.elegible else "No",
            f"{signo}{variacion}" if rk.posicion_anterior else "—",
        ])

    tabla = Table(datos, repeatRows=1, colWidths=[1.6*cm, 2.4*cm, 7*cm, 2.4*cm, 2.2*cm, 2.2*cm, 2.6*cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(_COLOR_TITULO)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF1F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    story = [
        Paragraph("REPORTE DE RANKING", titulo_style),
        Paragraph(
            f"{tipo_ranking.upper()} · {_nombre_pais(db, pais_codigo)} · {_nombre_ciclo(db, ciclo_id)}",
            sub_style,
        ),
        tabla,
        Spacer(1, 0.6 * cm),
        Paragraph(
            f"Generado el {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')} · "
            f"{len(rows)} representante(s) · Sistema MIP (SCGCPR)",
            nota_style,
        ),
    ]

    doc.build(story)
    buffer.seek(0)
    logger.info(f"Exportando ranking a PDF — {len(rows)} filas, tipo={tipo_ranking}, pais_codigo={pais_codigo}, ciclo_id={ciclo_id}")
    return buffer


def exportar_reconocimientos_pdf(
    db: Session,
    pais_codigo: Optional[str] = None,
    ciclo_id: Optional[int] = None,
) -> BytesIO:
    """Exporta el listado de reconocimientos otorgados a un PDF tabular en memoria."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    rows = _query_reconocimientos(db, pais_codigo, ciclo_id)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle("titulo", parent=styles["Title"],
                                  fontSize=20, textColor=colors.HexColor(_COLOR_TITULO),
                                  alignment=TA_CENTER, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=11, alignment=TA_CENTER, spaceAfter=14,
                               textColor=colors.HexColor(_COLOR_ACENTO))
    nota_style = ParagraphStyle("nota", parent=styles["Normal"],
                                fontSize=8, alignment=TA_LEFT, textColor=colors.grey)

    encabezados = ["Cód.", "Representante", "Premio", "Score", "Posición", "Certificado"]
    datos = [encabezados]
    for r in rows:
        rec = r.ReconocimientoRM
        datos.append([
            r.rm_codigo,
            r.rm_nombre,
            r.premio_nombre,
            f"{float(rec.score_total or 0):.2f}",
            f"#{rec.posicion_ranking}" if rec.posicion_ranking else "—",
            "Generado" if rec.certificado_generado else "Pendiente",
        ])

    tabla = Table(datos, repeatRows=1, colWidths=[2.4*cm, 6.5*cm, 6.5*cm, 2.4*cm, 2.6*cm, 3*cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(_COLOR_TITULO)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (2, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF1F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    story = [
        Paragraph("REPORTE DE RECONOCIMIENTOS", titulo_style),
        Paragraph(f"{_nombre_pais(db, pais_codigo)} · {_nombre_ciclo(db, ciclo_id)}", sub_style),
        tabla,
        Spacer(1, 0.6 * cm),
        Paragraph(
            f"Generado el {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')} · "
            f"{len(rows)} reconocimiento(s) · Sistema MIP (SCGCPR)",
            nota_style,
        ),
    ]

    doc.build(story)
    buffer.seek(0)
    logger.info(f"Exportando reconocimientos a PDF — {len(rows)} filas, pais_codigo={pais_codigo}, ciclo_id={ciclo_id}")
    return buffer
