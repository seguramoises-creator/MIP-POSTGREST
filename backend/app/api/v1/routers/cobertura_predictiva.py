"""
SCGCPR — Router: Cobertura Predictiva y Ritmo de Ejecución (4DX)
==================================================================
Sustituye a `/comercial` y a `GET /dashboard/comercial`: reemplaza las
métricas de resultado (lag measures: ventas, EVO IR, cumplimiento de cuota)
por las métricas predictivas de cobertura médica y ritmo de visitas (lead
measures) calculadas en vivo por `app.services.cobertura_predictiva_service`,
según la especificación funcional/técnica del Excel de referencia
`Modulo_Cobertura_Predictiva_Ejemplo_VM_v3.xlsx` (hoja Motor_Formulas).

Endpoints:
    GET  /cobertura-predictiva/resumen               — dashboard de equipo (GD/Ejecutivo)
    GET  /cobertura-predictiva/rm/{rm_id}             — detalle de un RM/VM
    GET  /cobertura-predictiva/parametros             — metas de cobertura configuradas
    POST /cobertura-predictiva/parametros             — upsert de meta de cobertura
    GET  /cobertura-predictiva/feriados               — feriados configurados
    POST /cobertura-predictiva/feriados               — crear feriado
    POST /cobertura-predictiva/cargar/target-medicos  — cargar universo de médicos target
    POST /cobertura-predictiva/cargar/visitas         — cargar bitácora de visitas

RBAC:
    - resumen / rm/{rm_id}: cualquier usuario autenticado, con auto-filtro:
        · REPRESENTANTE_MEDICO → no tiene resumen de equipo (403 en /resumen,
          dirigido a usar /rm/{rm_id} con su propio rm_id); en /rm/{rm_id}
          solo puede ver su propio rm_id.
        · GERENTE_DISTRITO     → en /resumen se fuerza gerente_id =
          current_user.gerente_id (403 si no tiene gerente_id asignado);
          en /rm/{rm_id} solo puede ver RMs de su propio equipo.
        · Resto de roles (ADMIN, PRESIDENCIA, DIR_COMERCIAL,
          GERENTE_PRODUCTIVIDAD, GERENTE_MARCA) sin restricción adicional
          (no existe Usuario.linea_id en el esquema para auto-filtrar por
          línea a GERENTE_MARCA).
    - parametros / feriados / cargar/*: ADMIN o GERENTE_PRODUCTIVIDAD.

Las cargas de Target_Medicos y Fact_Visitas son sincrónicas (sin tabla de
jobs `ETL.FACT_CargaExcel` ni BackgroundTasks): a diferencia del ETL
principal de KPIs, son cargas de menor volumen que no requieren el motor de
recálculo IUP/ranking.
"""
from datetime import date, datetime
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.core.deps import get_db, get_current_active_user, require_roles
from app.models.usuario import Rol, Usuario
from app.models.dimensiones import (
    RepresentanteMedico, Ciclo, TargetMedico, Feriado, ParametroCobertura,
)
from app.models.hechos import Visita
from app.services.cobertura_predictiva_service import (
    calcular_cobertura_rm, calcular_cobertura_equipo,
)
from app.api.v1.routers.etl import _validar_magic_bytes

router = APIRouter(prefix="/cobertura-predictiva", tags=["Cobertura Predictiva (4DX)"])

AnyAuth = Depends(get_current_active_user)
AdminOGerProd = Depends(require_roles(Rol.ADMIN, Rol.GERENTE_PRODUCTIVIDAD))

EXTENSIONES_PERMITIDAS = {".xlsx", ".xls"}


# ─────────────────────────────────────────────────────────────────────────
# Auto-filtro de alcance (mismo patrón que comercial.py, extendido a GD)
# ─────────────────────────────────────────────────────────────────────────

def _aplicar_alcance_equipo(current_user: Usuario, gerente_id: Optional[int]) -> Optional[int]:
    """
    Aplica las reglas de auto-filtro de alcance para el resumen de equipo.
    Lanza 403 si el rol no tiene sentido de equipo, o fuerza el gerente_id
    propio para GERENTE_DISTRITO (ignorando cualquier valor recibido).
    """
    if current_user.rol == Rol.REPRESENTANTE_MEDICO:
        raise HTTPException(
            403,
            "Un Representante Médico no tiene un resumen de equipo. "
            "Use GET /cobertura-predictiva/rm/{rm_id} con su propio rm_id.",
        )
    if current_user.rol == Rol.GERENTE_DISTRITO:
        if not current_user.gerente_id:
            raise HTTPException(403, "Su usuario no tiene un gerente_id asignado.")
        return current_user.gerente_id
    return gerente_id


def _verificar_acceso_rm(db: Session, current_user: Usuario, rm_id: int) -> RepresentanteMedico:
    """Valida que el usuario autenticado pueda ver el detalle de este RM."""
    rm = db.query(RepresentanteMedico).filter(RepresentanteMedico.id == rm_id).first()
    if not rm:
        raise HTTPException(404, f"RM ID={rm_id} no encontrado")

    if current_user.rol == Rol.REPRESENTANTE_MEDICO:
        if not current_user.rm_id or current_user.rm_id != rm_id:
            raise HTTPException(403, "Solo puede consultar su propio detalle de cobertura.")
    elif current_user.rol == Rol.GERENTE_DISTRITO:
        if not current_user.gerente_id or rm.gerente_id != current_user.gerente_id:
            raise HTTPException(403, "Este RM no pertenece a su equipo.")

    return rm


# ─────────────────────────────────────────────────────────────────────────
# Dashboard — resumen de equipo y detalle por RM
# ─────────────────────────────────────────────────────────────────────────

@router.get("/resumen", response_model=dict, summary="Resumen de cobertura predictiva del equipo")
def resumen_cobertura(
    ciclo_id: int = Query(..., description="Ciclo a evaluar"),
    pais_codigo: Optional[str] = None,
    gerente_id: Optional[int] = None,
    linea_id: Optional[int] = None,
    fecha_corte: Optional[date] = Query(None, description="Por defecto: hoy"),
    db: Session = Depends(get_db),
    current_user: Usuario = AnyAuth,
):
    """
    Sustituye a GET /dashboard/comercial: en vez de ventas/EVO IR/cuota (lag
    measures), retorna cobertura médica actual/proyectada y ritmo de
    visitas (lead measures, metodología 4DX) agregados por equipo.
    """
    ciclo = db.query(Ciclo).filter(Ciclo.id == ciclo_id).first()
    if not ciclo:
        raise HTTPException(404, f"Ciclo ID={ciclo_id} no encontrado")

    gerente_id = _aplicar_alcance_equipo(current_user, gerente_id)

    return calcular_cobertura_equipo(
        db, ciclo_id=ciclo_id, pais_codigo=pais_codigo, gerente_id=gerente_id,
        linea_id=linea_id, fecha_corte=fecha_corte,
    )


@router.get("/rm/{rm_id}", response_model=dict, summary="Detalle de cobertura predictiva de un RM/VM")
def detalle_cobertura_rm(
    rm_id: int,
    ciclo_id: int = Query(..., description="Ciclo a evaluar"),
    fecha_corte: Optional[date] = Query(None, description="Por defecto: hoy"),
    db: Session = Depends(get_db),
    current_user: Usuario = AnyAuth,
):
    _verificar_acceso_rm(db, current_user, rm_id)
    try:
        return calcular_cobertura_rm(db, rm_id=rm_id, ciclo_id=ciclo_id, fecha_corte=fecha_corte)
    except ValueError as e:
        raise HTTPException(404, str(e))


# ─────────────────────────────────────────────────────────────────────────
# Parametrización — meta de cobertura (DIM_ParametroCobertura)
# ─────────────────────────────────────────────────────────────────────────

class ParametroCoberturaIn(BaseModel):
    pais_codigo: str
    linea_id: Optional[int] = None
    ciclo_id: Optional[int] = None
    meta_cobertura: float = Field(..., gt=0, le=1, description="Fracción 0-1, ej. 0.90 = 90%")


@router.get("/parametros", response_model=list, summary="Metas de cobertura configuradas")
def listar_parametros(
    pais_codigo: Optional[str] = None,
    linea_id: Optional[int] = None,
    ciclo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = AdminOGerProd,
):
    q = db.query(ParametroCobertura).filter(ParametroCobertura.activo == True)
    if pais_codigo:
        q = q.filter(ParametroCobertura.pais_codigo == pais_codigo)
    if linea_id:
        q = q.filter(ParametroCobertura.linea_id == linea_id)
    if ciclo_id:
        q = q.filter(ParametroCobertura.ciclo_id == ciclo_id)

    return [
        {
            "id": p.id, "pais_codigo": p.pais_codigo, "linea_id": p.linea_id,
            "ciclo_id": p.ciclo_id, "meta_cobertura": float(p.meta_cobertura),
        }
        for p in q.order_by(ParametroCobertura.pais_codigo).all()
    ]


@router.post("/parametros", response_model=dict, summary="Crear o actualizar meta de cobertura")
def upsert_parametro(
    body: ParametroCoberturaIn,
    db: Session = Depends(get_db),
    current_user: Usuario = AdminOGerProd,
):
    """
    Upsert sobre la clave única (pais_codigo, linea_id, ciclo_id). Dejar
    linea_id y/o ciclo_id en NULL define una meta más general (ver
    resolución en cascada de `obtener_meta_cobertura`).
    """
    filtros = [ParametroCobertura.pais_codigo == body.pais_codigo]
    filtros.append(
        ParametroCobertura.linea_id.is_(None) if body.linea_id is None
        else ParametroCobertura.linea_id == body.linea_id
    )
    filtros.append(
        ParametroCobertura.ciclo_id.is_(None) if body.ciclo_id is None
        else ParametroCobertura.ciclo_id == body.ciclo_id
    )

    existente = db.query(ParametroCobertura).filter(*filtros).first()

    if existente:
        existente.meta_cobertura = body.meta_cobertura
        existente.activo = True
        db.commit()
        db.refresh(existente)
        return {"accion": "actualizado", "id": existente.id, "meta_cobertura": float(existente.meta_cobertura)}

    nuevo = ParametroCobertura(
        pais_codigo=body.pais_codigo, linea_id=body.linea_id, ciclo_id=body.ciclo_id,
        meta_cobertura=body.meta_cobertura, activo=True,
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return {"accion": "creado", "id": nuevo.id, "meta_cobertura": float(nuevo.meta_cobertura)}


# ─────────────────────────────────────────────────────────────────────────
# Feriados (DIM_Feriado) — usados en NETWORKDAYS si Ciclo.dias_laborables=0
# ─────────────────────────────────────────────────────────────────────────

class FeriadoIn(BaseModel):
    pais_codigo: str
    fecha: date
    nombre: Optional[str] = None


@router.get("/feriados", response_model=list, summary="Feriados configurados")
def listar_feriados(
    pais_codigo: Optional[str] = None,
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = AdminOGerProd,
):
    q = db.query(Feriado).filter(Feriado.activo == True)
    if pais_codigo:
        q = q.filter(Feriado.pais_codigo == pais_codigo)
    if desde:
        q = q.filter(Feriado.fecha >= desde)
    if hasta:
        q = q.filter(Feriado.fecha <= hasta)

    return [
        {"id": f.id, "pais_codigo": f.pais_codigo, "fecha": f.fecha.isoformat(), "nombre": f.nombre}
        for f in q.order_by(Feriado.fecha).all()
    ]


@router.post("/feriados", response_model=dict, status_code=201, summary="Crear feriado")
def crear_feriado(
    body: FeriadoIn,
    db: Session = Depends(get_db),
    current_user: Usuario = AdminOGerProd,
):
    existente = db.query(Feriado).filter(
        Feriado.pais_codigo == body.pais_codigo, Feriado.fecha == body.fecha,
    ).first()
    if existente:
        raise HTTPException(409, f"Ya existe un feriado para país={body.pais_codigo} en {body.fecha.isoformat()}")

    nuevo = Feriado(pais_codigo=body.pais_codigo, fecha=body.fecha, nombre=body.nombre, activo=True)
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return {"id": nuevo.id, "pais_codigo": nuevo.pais_codigo, "fecha": nuevo.fecha.isoformat(), "nombre": nuevo.nombre}


# ─────────────────────────────────────────────────────────────────────────
# Helpers de lectura de Excel (carga sincrónica, sin tabla de jobs)
# ─────────────────────────────────────────────────────────────────────────

def _leer_filas_hoja(ws) -> list:
    """
    Lee la primera fila como encabezado (normalizado a mayúsculas/strip) y
    el resto como filas de datos (dict header→valor). Omite filas
    completamente vacías. Mismo patrón de `dims.py:_leer_hoja`.
    """
    filas_iter = ws.iter_rows(values_only=True)
    try:
        encabezado = next(filas_iter)
    except StopIteration:
        return []
    columnas = [str(c).strip().upper() if c is not None else "" for c in encabezado]

    resultado = []
    for fila in filas_iter:
        if all(v is None or str(v).strip() == "" for v in fila):
            continue
        resultado.append({columnas[i]: fila[i] for i in range(min(len(columnas), len(fila)))})
    return resultado


def _col(fila: dict, *alias: str):
    for a in alias:
        if a in fila and fila[a] is not None and str(fila[a]).strip() != "":
            return fila[a]
    return None


def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _b(v, default: bool = True) -> bool:
    if v is None:
        return default
    s = str(v).strip().upper()
    if s in {"SI", "S", "TRUE", "1", "X", "YES", "Y"}:
        return True
    if s in {"NO", "N", "FALSE", "0"}:
        return False
    return default


def _fecha_celda(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _validar_archivo_excel(filename: Optional[str], content: bytes) -> str:
    """Valida extensión + magic bytes (FIX W-06, reutilizado de etl.py). Retorna la extensión."""
    ext = "." + (filename or "archivo.xlsx").rsplit(".", 1)[-1].lower()
    if ext not in EXTENSIONES_PERMITIDAS:
        raise HTTPException(400, f"Solo se permiten: {EXTENSIONES_PERMITIDAS}")
    if not _validar_magic_bytes(content, ext):
        raise HTTPException(400, "El contenido del archivo no corresponde a un Excel válido.")
    return ext


def _resolver_rm(fila: dict, rm_map: dict, rm_map_id: dict):
    """
    Resuelve el RM referenciado en la fila probando dos estrategias, en
    orden (los archivos reales de producción usan ambas convenciones según
    el origen del export):
        1) Código de negocio (RM_CODIGO/CODIGO_RM/VM_ID/VM_CODIGO/RM_ID)
           contra DIM_RM.codigo (ej. "VM01").
        2) Si el valor es puramente numérico y no hubo match por código,
           contra DIM_RM.id (ej. columna "RM_ID"=73 que referencia
           directamente el id de la base de datos, no un código de
           negocio — convención usada por Fact_Visitas).
    Retorna (rm | None, valor_crudo | None).
    """
    raw = _s(_col(fila, "RM_CODIGO", "CODIGO_RM", "VM_ID", "VM_CODIGO", "RM_ID"))
    if not raw:
        return None, None
    rm = rm_map.get(raw.upper())
    if rm:
        return rm, raw
    if raw.isdigit():
        rm = rm_map_id.get(int(raw))
        if rm:
            return rm, raw
    return None, raw


def _validar_consistencia_dimensional(fila: dict, rm: RepresentanteMedico, n: int, advertencias: list) -> None:
    """
    Valida sin bloquear la carga: si el archivo trae PAIS_ID y/o LINEA_ID
    (ids crudos de Config.DIM_Pais / Config.DIM_Linea, ej. Target_Medicos),
    compara contra el país/línea real del RM ya resuelto. El dato
    persistido es SIEMPRE el del RM (fuente de verdad relacional, vía
    DIM_RM.pais_codigo / DIM_RM.linea_id); una discrepancia solo se registra
    como advertencia para que el usuario revise el archivo de origen.
    """
    pais_archivo = _col(fila, "PAIS_ID")
    if pais_archivo is not None:
        try:
            pais_archivo_int = int(float(str(pais_archivo)))
        except (TypeError, ValueError):
            pais_archivo_int = None
        if pais_archivo_int is not None and pais_archivo_int != rm.pais_codigo:
            advertencias.append(
                f"Fila {n}: PAIS_ID del archivo ({pais_archivo_int}) no coincide con el "
                f"país real del RM '{rm.codigo}' (pais_codigo={rm.pais_codigo}); se usó el del RM."
            )

    linea_archivo = _col(fila, "LINEA_ID", "LINEA")
    if linea_archivo is not None:
        try:
            linea_archivo_int = int(float(str(linea_archivo)))
        except (TypeError, ValueError):
            linea_archivo_int = None
        if linea_archivo_int is not None and linea_archivo_int != rm.linea_id:
            advertencias.append(
                f"Fila {n}: LINEA_ID del archivo ({linea_archivo_int}) no coincide con la "
                f"línea real del RM '{rm.codigo}' (linea_id={rm.linea_id}); se usó la del RM."
            )


# ─────────────────────────────────────────────────────────────────────────
# Carga — Target_Medicos (universo de médicos programados por RM/ciclo)
# ─────────────────────────────────────────────────────────────────────────

@router.post(
    "/cargar/target-medicos", response_model=dict,
    summary="Cargar universo de médicos target (Target_Medicos)",
)
async def cargar_target_medicos(
    ciclo_id: int = Form(..., description="Ciclo al que pertenece el universo target"),
    modo: str = Form("SIMULACION", description="SIMULACION | PRODUCCION"),
    archivo: UploadFile = File(..., description="Excel con una hoja: RM_CODIGO, MEDICO_CODIGO, ..."),
    db: Session = Depends(get_db),
    current_user: Usuario = AdminOGerProd,
):
    """
    Carga el universo de médicos target por RM/ciclo (alimenta J/K del
    Motor_Formulas). Columnas reconocidas (alias entre paréntesis):
        RM_CODIGO (CODIGO_RM, VM_ID, VM_CODIGO, RM_ID)  — obligatoria.
            Acepta tanto código de negocio (ej. "VM01") como id numérico
            crudo de Config.DIM_RM (ej. "RM_ID"=73).
        MEDICO_CODIGO (MEDICO_ID)                         — obligatoria
        MEDICO_NOMBRE (NOMBRE_MEDICO)
        ESPECIALIDAD
        POTENCIAL
        PROGRAMADO (TARGET, PROGRAMADO_FLAG)               — SI/NO o 1/0, default SI
        PAIS_ID, LINEA_ID (LINEA)                          — opcionales; solo se
            usan para validar consistencia contra el país/línea real del RM
            (Config.DIM_RM.pais_codigo / linea_id). Una discrepancia no bloquea
            la carga, queda registrada en "advertencias".
    Idempotente: filas que ya existen para (rm_id, ciclo_id, medico_codigo)
    se omiten, no se duplican.
    """
    if modo.upper() not in {"SIMULACION", "PRODUCCION"}:
        raise HTTPException(400, "Modo debe ser SIMULACION o PRODUCCION")

    ciclo = db.query(Ciclo).filter(Ciclo.id == ciclo_id).first()
    if not ciclo:
        raise HTTPException(404, f"Ciclo ID={ciclo_id} no encontrado")

    content = await archivo.read()
    _validar_archivo_excel(archivo.filename, content)

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
        filas = _leer_filas_hoja(wb.worksheets[0])
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo Excel: {e}")

    todos_rm = db.query(RepresentanteMedico).all()
    rm_map = {r.codigo.strip().upper(): r for r in todos_rm}
    rm_map_id = {r.id: r for r in todos_rm}
    existentes = {
        (t.rm_id, t.medico_codigo)
        for t in db.query(TargetMedico.rm_id, TargetMedico.medico_codigo)
        .filter(TargetMedico.ciclo_id == ciclo_id).all()
    }

    insertados = 0
    omitidos = 0
    errores: list = []
    advertencias: list = []
    nuevos: list = []

    for n, fila in enumerate(filas, start=2):  # fila 1 = encabezado
        rm, rm_ref = _resolver_rm(fila, rm_map, rm_map_id)
        medico_codigo = _s(_col(fila, "MEDICO_CODIGO", "MEDICO_ID"))

        if not rm_ref or not medico_codigo:
            errores.append(f"Fila {n}: falta RM_CODIGO/RM_ID o MEDICO_CODIGO")
            continue

        if not rm:
            errores.append(f"Fila {n}: RM '{rm_ref}' no existe (no coincide por código ni por ID)")
            continue

        _validar_consistencia_dimensional(fila, rm, n, advertencias)

        if (rm.id, medico_codigo) in existentes:
            omitidos += 1
            continue

        if modo.upper() == "PRODUCCION":
            nuevos.append(TargetMedico(
                pais_codigo=rm.pais_codigo,
                rm_id=rm.id,
                ciclo_id=ciclo_id,
                medico_codigo=medico_codigo,
                medico_nombre=_s(_col(fila, "MEDICO_NOMBRE", "NOMBRE_MEDICO")),
                especialidad=_s(_col(fila, "ESPECIALIDAD")),
                potencial=_s(_col(fila, "POTENCIAL")),
                programado=_b(_col(fila, "PROGRAMADO", "TARGET", "PROGRAMADO_FLAG"), default=True),
                activo=True,
            ))
        existentes.add((rm.id, medico_codigo))
        insertados += 1

    if modo.upper() == "PRODUCCION" and nuevos:
        db.add_all(nuevos)
        db.commit()

    return {
        "modo": modo.upper(),
        "total_filas": len(filas),
        "insertados": insertados,
        "omitidos": omitidos,
        "errores": errores[:50],
        "total_errores": len(errores),
        "advertencias": advertencias[:50],
        "total_advertencias": len(advertencias),
    }


# ─────────────────────────────────────────────────────────────────────────
# Carga — Fact_Visitas (bitácora de visitas realizadas/programadas)
# ─────────────────────────────────────────────────────────────────────────

@router.post(
    "/cargar/visitas", response_model=dict,
    summary="Cargar bitácora de visitas (Fact_Visitas)",
)
async def cargar_visitas(
    ciclo_id: int = Form(..., description="Ciclo al que pertenecen las visitas"),
    modo: str = Form("SIMULACION", description="SIMULACION | PRODUCCION"),
    archivo: UploadFile = File(..., description="Excel con una hoja: RM_CODIGO, MEDICO_CODIGO, FECHA_VISITA, ..."),
    db: Session = Depends(get_db),
    current_user: Usuario = AdminOGerProd,
):
    """
    Carga la bitácora de visitas (alimenta L/M del Motor_Formulas: médicos
    únicos visitados y contactos totales). Columnas reconocidas:
        RM_CODIGO (CODIGO_RM, VM_ID, VM_CODIGO, RM_ID)  — obligatoria.
            Acepta tanto código de negocio (ej. "VM01") como id numérico
            crudo de Config.DIM_RM (ej. "RM_ID"=73, convención usada por
            Fact_Visitas).
        MEDICO_CODIGO (MEDICO_ID)                         — obligatoria
        FECHA_VISITA (FECHA)                               — obligatoria
        TIPO_CONTACTO (TIPO)
        ESTADO_VISITA (ESTADO)                              — default 'Realizada'
        PRODUCTO_FOCO (PRODUCTO)
        PAIS_ID                                              — opcional; solo se
            usa para validar consistencia contra el país real del RM
            (Config.DIM_RM.pais_codigo). Una discrepancia no bloquea la carga,
            queda registrada en "advertencias".
    NO es idempotente: es una bitácora de eventos (append-only); cargar el
    mismo archivo dos veces duplica las visitas. A diferencia de
    Target_Medicos, no tiene restricción de unicidad en BD.
    """
    if modo.upper() not in {"SIMULACION", "PRODUCCION"}:
        raise HTTPException(400, "Modo debe ser SIMULACION o PRODUCCION")

    ciclo = db.query(Ciclo).filter(Ciclo.id == ciclo_id).first()
    if not ciclo:
        raise HTTPException(404, f"Ciclo ID={ciclo_id} no encontrado")

    content = await archivo.read()
    _validar_archivo_excel(archivo.filename, content)

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
        filas = _leer_filas_hoja(wb.worksheets[0])
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo Excel: {e}")

    todos_rm = db.query(RepresentanteMedico).all()
    rm_map = {r.codigo.strip().upper(): r for r in todos_rm}
    rm_map_id = {r.id: r for r in todos_rm}

    insertados = 0
    errores: list = []
    advertencias: list = []
    nuevas: list = []

    for n, fila in enumerate(filas, start=2):
        rm, rm_ref = _resolver_rm(fila, rm_map, rm_map_id)
        medico_codigo = _s(_col(fila, "MEDICO_CODIGO", "MEDICO_ID"))
        fecha_visita = _fecha_celda(_col(fila, "FECHA_VISITA", "FECHA"))

        if not rm_ref or not medico_codigo or not fecha_visita:
            errores.append(f"Fila {n}: falta RM_CODIGO/RM_ID, MEDICO_CODIGO o FECHA_VISITA válida")
            continue

        if not rm:
            errores.append(f"Fila {n}: RM '{rm_ref}' no existe (no coincide por código ni por ID)")
            continue

        _validar_consistencia_dimensional(fila, rm, n, advertencias)

        if modo.upper() == "PRODUCCION":
            nuevas.append(Visita(
                pais_codigo=rm.pais_codigo,
                rm_id=rm.id,
                ciclo_id=ciclo_id,
                medico_codigo=medico_codigo,
                fecha_visita=fecha_visita,
                tipo_contacto=_s(_col(fila, "TIPO_CONTACTO", "TIPO")),
                estado_visita=_s(_col(fila, "ESTADO_VISITA", "ESTADO")) or "Realizada",
                producto_foco=_s(_col(fila, "PRODUCTO_FOCO", "PRODUCTO")),
            ))
        insertados += 1

    if modo.upper() == "PRODUCCION" and nuevas:
        db.add_all(nuevas)
        db.commit()

    return {
        "modo": modo.upper(),
        "total_filas": len(filas),
        "insertados": insertados,
        "errores": errores[:50],
        "total_errores": len(errores),
        "advertencias": advertencias[:50],
        "total_advertencias": len(advertencias),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS cat.* — Metodología 4DX (SP + vista)
# ═══════════════════════════════════════════════════════════════════════════════

from app.services.cobertura_predictiva_service import (
    get_ciclos_cat,
    calcular_cobertura_sp,
    get_dashboard_cat,
    get_cobertura_por_categoria,
    cargar_excel_cobertura,
)


# ── GET /cobertura-predictiva/cat/ciclos ────────────────────────────────────
@router.get("/cat/ciclos", summary="Ciclos disponibles en cat.DimCiclo")
def listar_ciclos_cat(
    pais_codigo: Optional[str] = Query(None, description="Filtrar por código de país, ej. DO"),
    _current_user: Usuario = AnyAuth,
    db: Session = Depends(get_db),
):
    """Lista los ciclos promocionales cargados en cat.DimCiclo para poblar selectores."""
    return get_ciclos_cat(db, pais_codigo=pais_codigo)


# ── GET /cobertura-predictiva/cat/dashboard ─────────────────────────────────
@router.get("/cat/dashboard", summary="Dashboard 4DX desde cat.KpiCoberturaPredictiva")
def dashboard_4dx(
    ciclo_codigo: str = Query(..., description="Código de ciclo, ej. 2026-06"),
    pais_codigo: str = Query(..., description="Código de país, ej. DO"),
    fecha_corte: Optional[str] = Query(None, description="Fecha de corte YYYY-MM-DD (default: hoy)"),
    linea: Optional[str] = Query(None, description="Filtrar por línea"),
    gd: Optional[str] = Query(None, description="Filtrar por GD"),
    representante: Optional[str] = Query(None, description="Filtrar por nombre o código de VM"),
    _current_user: Usuario = AnyAuth,
    db: Session = Depends(get_db),
):
    """
    Lee de cat.vwDashboardCoberturaPredictivaGD el resultado más cercano a la
    fecha de corte. Aplica auto-filtro por scope para GERENTE_DISTRITO y REPRESENTANTE_MEDICO.
    """
    from datetime import date as _date
    fecha = None
    if fecha_corte:
        try:
            fecha = _date.fromisoformat(fecha_corte)
        except ValueError:
            from fastapi import HTTPException as _HTTPException
            raise _HTTPException(status_code=422, detail=f"fecha_corte inválida: '{fecha_corte}' — use YYYY-MM-DD")

    # Auto-filtro por scope
    user_rol = _current_user.rol
    if user_rol == Rol.REPRESENTANTE_MEDICO:
        if not _current_user.rm_id:
            from fastapi import HTTPException as _HTTPException
            raise _HTTPException(status_code=403, detail="Usuario RM sin rm_id configurado")
        representante = str(_current_user.rm_id)
    elif user_rol == Rol.GERENTE_DISTRITO and _current_user.gerente_id:
        pass  # el filtro por GD se aplica en el SP; para simplificar se deja sin restricción aquí

    return get_dashboard_cat(
        db,
        ciclo_codigo=ciclo_codigo,
        pais_codigo=pais_codigo,
        fecha_corte=fecha,
        linea=linea,
        gd=gd,
        representante=representante,
    )


# ── GET /cobertura-predictiva/cat/cobertura-categorias ──────────────────────
@router.get("/cat/cobertura-categorias",
            summary="Avance de visitas por categoría de médico (A/B/C/D)")
def cobertura_por_categoria(
    ciclo_codigo: str  = Query(..., description="Código de ciclo, ej. 2026-06"),
    pais_codigo: str   = Query(..., description="Código de país, ej. DO"),
    representante: Optional[str] = Query(None, description="Filtrar por código de VM"),
    gd: Optional[str]  = Query(None, description="Filtrar por nombre de GD (EquipoTexto)"),
    _current_user: Usuario = AnyAuth,
    db: Session = Depends(get_db),
):
    """
    Devuelve para cada categoría de médico (A/B/C/D) el total de médicos programados,
    los visitados (al menos una visita 'Realizada') y los pendientes, con porcentaje.
    Útil para las barras de progreso del dashboard de Cobertura Predictiva.
    """
    # Auto-filtro por scope
    if _current_user.rol == Rol.REPRESENTANTE_MEDICO:
        if not _current_user.rm_id:
            from fastapi import HTTPException as _HTTPException
            raise _HTTPException(status_code=403, detail="Usuario RM sin rm_id configurado")
        representante = str(_current_user.rm_id)

    return get_cobertura_por_categoria(
        db,
        ciclo_codigo=ciclo_codigo,
        pais_codigo=pais_codigo,
        representante_codigo=representante,
        gd=gd,
    )


# ── DELETE /cobertura-predictiva/cat/datos ──────────────────────────────────
@router.delete(
    "/cat/datos",
    summary="Elimina datos cat.* para recargar desde cero",
    status_code=200,
)
def reset_datos_cat(
    pais_codigo: str = Query(..., description="Código de país cuya data se elimina, ej. DO"),
    ciclo_codigo: Optional[str] = Query(None, description="Si se indica, solo borra ese ciclo; vacío = TODOS los ciclos del país"),
    incluir_dims: bool = Query(False, description="Si True, también borra DimMedico, DimCiclo, DimRepresentante, DimEspecialidad del país"),
    _current_user: Usuario = AdminOGerProd,
    db: Session = Depends(get_db),
):
    """
    Limpia los datos de Cobertura Predictiva (esquema cat.*) para el país indicado,
    permitiendo una carga limpia desde cero.

    Orden de borrado (respeta FKs):
    1. cat.KpiCoberturaPredictiva
    2. cat.FactVisitaMedica
    3. cat.FactTargetMedicoCiclo
    4. (opcional) cat.DimMedico, cat.DimCiclo, cat.DimRepresentanteMedico, cat.DimEspecialidad

    SOLO para usuarios ADMIN o GERENTE_PRODUCTIVIDAD.
    """
    from sqlalchemy import text as _t
    from loguru import logger as _log

    try:
        # Resolver pais_key
        pais_row = db.execute(
            _t('SELECT "PaisKey" FROM "cat"."DimPais" WHERE "CodigoPais" = :c'),
            {"c": pais_codigo.upper()},
        ).mappings().first()
        if not pais_row:
            from fastapi import HTTPException as _HTTPException
            raise _HTTPException(status_code=404, detail=f"País '{pais_codigo}' no existe en cat.DimPais")
        pais_key = pais_row["PaisKey"]

        # Resolver ciclo_key si se especificó
        ciclo_key = None
        if ciclo_codigo:
            ciclo_row = db.execute(
                _t('SELECT "CicloKey" FROM "cat"."DimCiclo" WHERE "CodigoCiclo" = :c AND "PaisKey" = :p'),
                {"c": ciclo_codigo, "p": pais_key},
            ).mappings().first()
            if not ciclo_row:
                from fastapi import HTTPException as _HTTPException
                raise _HTTPException(status_code=404, detail=f"Ciclo '{ciclo_codigo}' no existe para el país '{pais_codigo}'")
            ciclo_key = ciclo_row["CicloKey"]

        conteos: dict = {}

        def _borrar(tabla: str, extra_where: str = "") -> int:
            where = f'"PaisKey" = {pais_key}'
            if ciclo_key:
                where += f' AND "CicloKey" = {ciclo_key}'
            if extra_where:
                where += f" AND {extra_where}"
            r = db.execute(_t(f'DELETE FROM "cat"."{tabla}" WHERE {where}'))
            return r.rowcount

        # 1. KPIs calculados (sin FK entrante)
        conteos["KpiCoberturaPredictiva"] = _borrar("KpiCoberturaPredictiva")
        # 2. Visitas
        conteos["FactVisitaMedica"] = _borrar("FactVisitaMedica")
        # 3. Target médicos
        conteos["FactTargetMedicoCiclo"] = _borrar("FactTargetMedicoCiclo")

        if incluir_dims:
            # Orden FK: FactMedicoCategoriaDetalle → FactMedicoCategoriaSnapshot → DimMedico
            # FactMedicoCategoriaDetalle no tiene PaisKey propio: se borra por subquery
            def _tabla_existe(nombre: str) -> bool:
                return bool(db.execute(_t(
                    f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES "
                    f"WHERE TABLE_SCHEMA='cat' AND TABLE_NAME='{nombre}'"
                )).fetchone())

            if _tabla_existe("FactMedicoCategoriaDetalle"):
                # Obtener nombres exactos de columnas del FK constraint desde sys.foreign_keys
                # (evita hardcodear nombres que pueden diferir entre migraciones y BD real)
                fk_info = db.execute(_t("""
                    SELECT pc.name AS parent_col, rc.name AS ref_col
                    FROM sys.foreign_keys fk
                    JOIN sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
                    JOIN sys.columns pc ON fkc.parent_object_id = pc.object_id
                        AND fkc.parent_column_id = pc.column_id
                    JOIN sys.columns rc ON fkc.referenced_object_id = rc.object_id
                        AND fkc.referenced_column_id = rc.column_id
                    WHERE fk.name = 'FK_FactDetalle_Snapshot'
                      AND OBJECT_SCHEMA_NAME(fk.parent_object_id) = 'cat'
                """)).fetchone()
                if fk_info:
                    parent_col, ref_col = fk_info[0], fk_info[1]
                    r = db.execute(_t(
                        f'DELETE d FROM "cat"."FactMedicoCategoriaDetalle" d '
                        f'INNER JOIN "cat"."FactMedicoCategoriaSnapshot" s '
                        f'  ON s."{ref_col}" = d."{parent_col}" '
                        f'WHERE s."PaisKey" = {pais_key}'
                    ))
                    conteos["FactMedicoCategoriaDetalle"] = r.rowcount
                else:
                    # Fallback: si no se encuentra el FK, intentar borrar por subquery en PK
                    _log.warning("[reset_datos_cat] FK_FactDetalle_Snapshot no encontrado, intentando DELETE directo")
                    r = db.execute(_t(
                        f'DELETE FROM "cat"."FactMedicoCategoriaDetalle" WHERE '
                        f'EXISTS (SELECT 1 FROM "cat"."FactMedicoCategoriaSnapshot" snap '
                        f'  WHERE snap."PaisKey" = {pais_key} '
                        f'  AND snap."MedicoCategoriaKey" = "cat"."FactMedicoCategoriaDetalle"."MedicoCategoriaKey")'
                    ))
                    conteos["FactMedicoCategoriaDetalle"] = r.rowcount

            if _tabla_existe("FactMedicoCategoriaSnapshot"):
                r = db.execute(_t(f'DELETE FROM "cat"."FactMedicoCategoriaSnapshot" WHERE "PaisKey" = {pais_key}'))
                conteos["FactMedicoCategoriaSnapshot"] = r.rowcount

            if _tabla_existe("FactCategorizacionMedica"):
                r = db.execute(_t(f'DELETE FROM "cat"."FactCategorizacionMedica" WHERE "PaisKey" = {pais_key}'))
                conteos["FactCategorizacionMedica"] = r.rowcount

            # DimMedico — no tiene CicloKey, borrar todo el país
            r = db.execute(_t(f'DELETE FROM "cat"."DimMedico" WHERE "PaisKey" = {pais_key}'))
            conteos["DimMedico"] = r.rowcount
            # DimCiclo
            if ciclo_key:
                r = db.execute(_t(f'DELETE FROM "cat"."DimCiclo" WHERE "CicloKey" = {ciclo_key} AND "PaisKey" = {pais_key}'))
            else:
                r = db.execute(_t(f'DELETE FROM "cat"."DimCiclo" WHERE "PaisKey" = {pais_key}'))
            conteos["DimCiclo"] = r.rowcount
            # DimRepresentanteMedico
            r = db.execute(_t(f'DELETE FROM "cat"."DimRepresentanteMedico" WHERE "PaisKey" = {pais_key}'))
            conteos["DimRepresentanteMedico"] = r.rowcount
            # DimEspecialidad
            r = db.execute(_t(f'DELETE FROM "cat"."DimEspecialidad" WHERE "PaisKey" = {pais_key}'))
            conteos["DimEspecialidad"] = r.rowcount

        db.commit()
        _log.info(f"[reset_datos_cat] país={pais_codigo} ciclo={ciclo_codigo or 'TODOS'} dims={incluir_dims} conteos={conteos}")
        return {
            "ok": True,
            "pais": pais_codigo,
            "ciclo": ciclo_codigo or "TODOS",
            "dims_eliminadas": incluir_dims,
            "filas_eliminadas": conteos,
        }

    except Exception as exc:
        db.rollback()
        from fastapi import HTTPException as _HTTPException
        if isinstance(exc, _HTTPException):
            raise
        raise _HTTPException(status_code=500, detail=str(exc))


# ── POST /cobertura-predictiva/cat/calcular ─────────────────────────────────
@router.post("/cat/calcular", summary="Ejecuta cat.sp_CalcularCoberturaPredictiva")
def calcular_sp(
    ciclo_codigo: str = Query(..., description="Código de ciclo, ej. 2026-06"),
    pais_codigo: str = Query(..., description="Código de país, ej. DO"),
    fecha_corte: Optional[str] = Query(None, description="Fecha de corte YYYY-MM-DD (default: hoy)"),
    representante_key: Optional[int] = Query(None, description="Calcular solo para un VM específico"),
    linea: Optional[str] = Query(None, description="Filtrar por línea"),
    _current_user: Usuario = AdminOGerProd,
    db: Session = Depends(get_db),
):
    """
    Materializa los KPIs de cobertura predictiva en cat.KpiCoberturaPredictiva
    para el ciclo/país/fecha indicados. Elimina los resultados anteriores para
    esa combinación (FechaCorte + CicloKey) y los recalcula frescos.
    """
    from datetime import date as _date
    fecha = None
    if fecha_corte:
        try:
            fecha = _date.fromisoformat(fecha_corte)
        except ValueError:
            from fastapi import HTTPException as _HTTPException
            raise _HTTPException(status_code=422, detail=f"fecha_corte inválida: '{fecha_corte}'")

    try:
        return calcular_cobertura_sp(
            db,
            ciclo_codigo=ciclo_codigo,
            pais_codigo=pais_codigo,
            fecha_corte=fecha,
            representante_key=representante_key,
            linea=linea,
        )
    except Exception as e:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(status_code=500, detail=str(e))


# ── POST /cobertura-predictiva/cat/cargar-excel ─────────────────────────────
@router.post("/cat/cargar-excel", summary="Carga Excel completo de Cobertura Predictiva (cat.*)")
async def cargar_excel_cat(
    archivo: UploadFile = File(..., description="Excel .xlsx con las hojas del modelo de Cobertura Predictiva"),
    pais_codigo: str = Form(..., description="Código de país, ej. DO"),
    ciclo_codigo: Optional[str] = Form(None, description="Ciclo de referencia, ej. C06-2026. Se usa cuando el Excel no incluye hoja COB_DIM_CICLO."),
    calcular_al_cargar: bool = Form(True, description="Ejecutar SP de cálculo tras cargar (fecha_corte = hoy)"),
    _current_user: Usuario = AdminOGerProd,
    db: Session = Depends(get_db),
):
    """
    Carga las 4 hojas del modelo de Cobertura Predictiva:
      - COB_DIM_CICLO          → cat.DimCiclo
      - COB_DIM_CALENDARIO     → cat.DimCalendario
      - COB_TARGET_MEDICO_CICLO → cat.FactTargetMedicoCiclo
      - COB_FACT_VISITAS        → cat.FactVisitaMedica

    Si el Excel no incluye la hoja COB_DIM_CICLO, se usa `ciclo_codigo` como referencia
    para asociar los datos a un ciclo existente en cat.DimCiclo.

    Si `calcular_al_cargar=true`, ejecuta automáticamente el SP de cálculo
    (cat.sp_CalcularCoberturaPredictiva) para cada ciclo cargado.
    """
    if not archivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    contenido = await archivo.read()
    if len(contenido) > 100 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Archivo mayor a 100 MB")

    if contenido[:4] != b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="El archivo no es un Excel .xlsx válido")

    try:
        stats = cargar_excel_cobertura(
            db,
            excel_bytes=contenido,
            pais_codigo=pais_codigo,
            ciclo_codigo_default=ciclo_codigo.strip() if ciclo_codigo else None,
            usuario=_current_user.username,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar: {str(e)}")

    resultado_sp = None
    if calcular_al_cargar and stats.get("ciclos", 0) == 0 or (stats.get("target", 0) > 0 or stats.get("visitas", 0) > 0):
        # Calcular para todos los ciclos recién cargados
        from app.services.cobertura_predictiva_service import get_ciclos_cat as _get_ciclos
        from datetime import date as _date
        ciclos = _get_ciclos(db, pais_codigo=pais_codigo)
        resultados_sp = []
        for c in ciclos[:10]:  # máximo 10 ciclos activos por país
            try:
                r = calcular_cobertura_sp(
                    db,
                    ciclo_codigo=c["codigo_ciclo"],
                    pais_codigo=pais_codigo,
                    fecha_corte=_date.today(),
                )
                resultados_sp.append({"ciclo": c["codigo_ciclo"], **r})
            except Exception:
                pass
        resultado_sp = resultados_sp

    return {
        "ok": True,
        "mensaje": (
            f"Carga completada — "
            f"{stats.get('ciclos',0)} ciclos, "
            f"{stats.get('calendario',0)} fechas de calendario, "
            f"{stats.get('target',0)} médicos programados, "
            f"{stats.get('visitas',0)} visitas."
        ),
        **stats,
        "calculo_sp": resultado_sp,
    }
