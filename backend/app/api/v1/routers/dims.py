"""
SCGCPR — Router: Importación de Dimensiones (DIMs)
====================================================
Permite cargar masivamente las tablas de dimensiones desde un archivo Excel
(.xlsx) que contiene múltiples hojas, una por cada DIM.

Endpoints:
  POST /dims/preview  → Sube el Excel y retorna las hojas detectadas con
                        su mapeo propuesto a tabla del sistema
  POST /dims/importar → Sube el Excel + lista de hojas seleccionadas
                        y carga los datos en BD

Nombres de hojas soportados (del Excel DIM_MIP_FINAL.xlsx):
  DIM_PAIS, DIM_LINEA, DIM_INDICADOR,
  DIM_GERENTE / DIM_GERENTE_DISTRITO,
  DIM_RM / DIM_REPRESENTANTE_MEDICO,
  DIM_INDICADOR_TABLA / DIM_REGLA_PUNTUACION,
  DIM_CICLO / DIM_TIEMPO,
  DIM_MES / DIM_TIEMPO,
  DIM_RECONOCIMIENTO,
  DIM_META_INDICADOR,
  DIM_CATEGORIA_DESEMPENO,
  DIM_KPI_DASHBOARD,
  DIM_MEDICOCOBERTURA (universo médico de Cobertura Predictiva/4DX —
    se mapea a Config.DIM_Medico vía resolver_medico_por_codigo(), dedup
    por (pais_codigo, codigo); no se confunde con el universo de Categorización
    Médica, que dedup por (pais_codigo, nombre, centro_medico_id))
"""
import io
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.deps import get_db, require_roles
from app.models.usuario import Rol
from app.models.dimensiones import (
    Pais, Linea, Gerente, RepresentanteMedico,
    Indicador, IndicadorTabla, MetaIndicador, CategoriaDesempeno, KpiDashboard,
    Ciclo, Mes, Premio, Medico,
)
from app.services.categorizacion_service import resolver_medico_por_codigo

router = APIRouter(prefix="/dims", tags=["Importación DIMs"])

AdminOGerProd = Depends(require_roles(Rol.ADMIN, Rol.GERENTE_PRODUCTIVIDAD))

# ── Mapeo canónico (nombre normalizado → info) ─────────────────────────
HOJAS_CONOCIDAS: dict[str, dict] = {
    # Nombres originales del sistema
    "DIM_PAIS":            {"tabla": "DIM_Pais",           "label": "Países",                  "orden": 1},
    "DIM_LINEA":           {"tabla": "DIM_Linea",          "label": "Líneas de Productos",      "orden": 2},
    "DIM_GERENTE":         {"tabla": "DIM_Gerente",        "label": "Gerentes",                 "orden": 3},
    "DIM_RM":              {"tabla": "DIM_RM",             "label": "Representantes Médicos",    "orden": 4},
    "DIM_INDICADOR":       {"tabla": "DIM_Indicador",      "label": "Indicadores de Desempeño", "orden": 5},
    "DIM_INDICADOR_TABLA": {"tabla": "DIM_IndicadorTabla", "label": "Rangos de Puntuación",     "orden": 6},
    "DIM_CICLO":           {"tabla": "DIM_Ciclo",          "label": "Ciclos de Trabajo",        "orden": 7},
    "DIM_MES":             {"tabla": "DIM_Mes",            "label": "Meses del Año",            "orden": 8},
    "DIM_RECONOCIMIENTO":  {"tabla": "DIM_Premio",         "label": "Reconocimientos",          "orden": 9},
    "DIM_META_INDICADOR":  {"tabla": "DIM_MetaIndicador",  "label": "Metas por Indicador",      "orden": 10},
    "DIM_CATEGORIA_DESEMPENO": {"tabla": "DIM_CategoriaDesempeno", "label": "Categorías de Desempeño", "orden": 11},
    "DIM_KPI_DASHBOARD":   {"tabla": "DIM_KpiDashboard",   "label": "KPIs de Dashboard",        "orden": 12},
    "DIM_MEDICOCOBERTURA": {"tabla": "DIM_Medico",         "label": "Médicos (Cobertura Predictiva)", "orden": 13},

    # Alias del Excel DIM_MIP_FINAL.xlsx
    "DIM_GERENTE_DISTRITO":    {"tabla": "DIM_Gerente",        "label": "Gerentes de Distrito",     "orden": 3},
    "DIM_REPRESENTANTE_MEDICO":{"tabla": "DIM_RM",             "label": "Representantes Médicos",    "orden": 4},
    "DIM_REGLA_PUNTUACION":    {"tabla": "DIM_IndicadorTabla", "label": "Rangos de Puntuación",     "orden": 6},
    "DIM_TIEMPO":              {"tabla": "DIM_Ciclo/DIM_Mes",  "label": "Ciclos y Meses",           "orden": 7},
    # Hoja de factores RM (formato multi-bloque del Excel TABLA_RM.xlsx)
    "TABLAS RM":               {"tabla": "DIM_IndicadorTabla", "label": "Tablas de Factores RM",    "orden": 6},
    "TABLAS_RM":               {"tabla": "DIM_IndicadorTabla", "label": "Tablas de Factores RM",    "orden": 6},
    "TABLA_RM":                {"tabla": "DIM_IndicadorTabla", "label": "Tablas de Factores RM",    "orden": 6},
    "DIM_TABLA_RM":            {"tabla": "DIM_IndicadorTabla", "label": "Tablas de Factores RM",    "orden": 6},
}

# Alias: nombre del Excel → nombre canónico interno para _importar_hoja
ALIAS: dict[str, str] = {
    "DIM_GERENTE_DISTRITO":     "DIM_GERENTE",
    "DIM_REPRESENTANTE_MEDICO": "DIM_RM",
    "DIM_REGLA_PUNTUACION":     "DIM_INDICADOR_TABLA",
    "DIM_TIEMPO":               "DIM_TIEMPO",  # handler propio
    # Alias para la hoja de factores (multi-bloque, formato TABLA_RM.xlsx)
    "TABLAS RM":                "TABLAS_RM",
    "TABLA_RM":                 "TABLAS_RM",
    "DIM_TABLA_RM":             "TABLAS_RM",
}


# ── Schemas ────────────────────────────────────────────────────────────

class HojaInfo(BaseModel):
    nombre_hoja: str
    tabla_sistema: str
    label: str
    filas: int
    columnas: List[str]
    reconocida: bool
    orden: int

class PreviewResponse(BaseModel):
    hojas: List[HojaInfo]
    total_hojas: int
    hojas_reconocidas: int

class ResultadoHoja(BaseModel):
    nombre_hoja: str
    label: str
    exitoso: bool
    insertados: int
    omitidos: int
    errores: int
    mensaje: str

class ImportarResponse(BaseModel):
    resultados: List[ResultadoHoja]
    total_insertados: int
    total_omitidos: int
    total_errores: int


# ── Helpers de conversión ──────────────────────────────────────────────

def _leer_hoja(ws) -> list[dict]:
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip().upper() if cell.value is not None else "")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows

def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.upper() == "NULL":
        return None
    return s

def _i(v) -> Optional[int]:
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None

def _d(v) -> Optional[Decimal]:
    try: return Decimal(str(v))
    except: return None

def _fecha(v) -> Optional[date]:
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except: return None

# Códigos de país alternativos: las hojas DIM_LINEA/DIM_GERENTE_DISTRITO/
# DIM_REPRESENTANTE_MEDICO/DIM_INDICADOR/DIM_META_INDICADOR del Excel actual
# usan el código ISO 3166-1 alpha-2 estándar para República Dominicana ("DO"),
# mientras que la hoja DIM_PAIS usa la abreviatura local en español ("RD").
# Es el mismo país con dos convenciones de código distintas dentro del mismo
# archivo — se normaliza aquí en vez de exigir que el Excel cambie su
# estructura (instrucción explícita del usuario: adaptar el código, no las
# tablas DIMS/FACT).
PAIS_CODIGO_ALIAS: dict[str, str] = {
    "DO": "RD",
}

def _pais_c(pais_map: dict, valor) -> Optional[str]:
    """
    Resuelve el código de país (la llave string usada en pais_map/linea_map/
    gerente_map/indicador_map) a partir de un valor de Excel que puede venir
    como código string (ej. "RD", o su alias "DO" — ver PAIS_CODIGO_ALIAS) o
    como el id numérico crudo de Config.DIM_Pais (ej. 5) — formato real de los
    archivos DIMS/FACT actuales del usuario, donde columnas como
    PAIS_CODIGO/PAIS_ID llevan el id numérico de DIM_Pais en vez del código de
    2 letras.

    No modifica pais_map — solo lo consulta — para no romper el código que
    itera pais_map.items() esperando exactamente una entrada por país
    (DIM_TIEMPO, TABLAS_RM, fallback "todos los países" de DIM_INDICADOR/
    DIM_INDICADOR_TABLA/DIM_META_INDICADOR).
    """
    v = _s(valor)
    if not v:
        return None
    v = v.upper()
    if v in pais_map:
        return v
    alias = PAIS_CODIGO_ALIAS.get(v)
    if alias and alias in pais_map:
        return alias
    vi = _i(valor)
    if vi is not None:
        for cod, pid in pais_map.items():
            if pid == vi:
                return cod
    return None


def _leer_tablas_rm(ws) -> list[dict]:
    """
    Lector para la hoja 'Tablas RM' del Excel DIM_MIP_FINAL.xlsx.

    Formato: bloques de 5 columnas lado a lado:
      INDICADOR_ID | CODIGO_INDICADOR | PESO | RESULTADO | FACTOR

    - RESULTADO contiene el valor de lookup (85, 86, … o texto 'menos 85').
      Las filas con texto en RESULTADO se omiten automáticamente.
    - FACTOR viene en formato porcentaje en Excel (75% → almacenado como 0.75).
    - Cada fila incluye su propio INDICADOR_ID y PESO → no es necesario
      rastrear el último ID por bloque.

    Retorna lista de dicts:
      {indicador_id, codigo_indicador, peso, valor, factor}
    """
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0
    if max_col == 0 or max_row == 0:
        return []

    # Detectar fila de headers: primera fila donde col-1 = "INDICADOR_ID"
    header_row = None
    for r in range(1, min(10, max_row + 1)):
        v = ws.cell(r, 1).value
        if v and str(v).strip().upper() == "INDICADOR_ID":
            header_row = r
            break
    if header_row is None:
        return []

    data_rows: list[dict] = []

    for r in range(header_row + 1, max_row + 1):
        # Bloques de 5 columnas: INDICADOR_ID | CODIGO_INDICADOR | PESO | RESULTADO | FACTOR
        for col_start in range(1, max_col, 5):
            col_ind  = col_start        # INDICADOR_ID
            col_cod  = col_start + 1   # CODIGO_INDICADOR
            col_peso = col_start + 2   # PESO
            col_res  = col_start + 3   # RESULTADO
            col_fac  = col_start + 4   # FACTOR

            if col_res > max_col:
                break

            v_id   = ws.cell(r, col_ind).value
            v_cod  = ws.cell(r, col_cod).value
            v_peso = ws.cell(r, col_peso).value
            v_res  = ws.cell(r, col_res).value
            v_fac  = ws.cell(r, col_fac).value if col_fac <= max_col else None

            # Sin RESULTADO → fila vacía para este bloque
            if v_res is None:
                continue

            # RESULTADO debe ser numérico (omite 'menos 85' y similares)
            try:
                val_f = float(str(v_res))
            except (ValueError, TypeError):
                continue

            # INDICADOR_ID obligatorio y numérico
            if v_id is None:
                continue
            try:
                ind_id = int(float(str(v_id)))
            except (ValueError, TypeError):
                continue

            # PESO desde la fila (ya no se consulta la BD)
            try:
                peso_f = int(float(str(v_peso))) if v_peso is not None else 0
            except (ValueError, TypeError):
                peso_f = 0

            # FACTOR: Excel almacena 75% como 0.75
            try:
                fac_f = float(str(v_fac)) if v_fac is not None else 0.0
            except (ValueError, TypeError):
                continue

            data_rows.append({
                "indicador_id":    ind_id,
                "codigo_indicador": str(v_cod).strip().upper() if v_cod else "",
                "peso":            peso_f,
                "valor":           val_f,
                "factor":          fac_f,
            })

    return data_rows


# ── Endpoint 1: Preview ────────────────────────────────────────────────

@router.post("/preview", response_model=PreviewResponse,
             summary="Leer Excel y detectar hojas disponibles")
async def preview_dims(
    file: UploadFile = File(..., description="Archivo Excel (.xlsx) con hojas DIM_*"),
    _=AdminOGerProd,
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    contenido = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

    hojas = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        nombre_upper = nombre_hoja.strip().upper()
        info = HOJAS_CONOCIDAS.get(nombre_upper, {})

        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers.append(str(cell.value).strip())

        filas = max(0, ws.max_row - 1)

        hojas.append(HojaInfo(
            nombre_hoja=nombre_hoja,
            tabla_sistema=info.get("tabla", nombre_hoja),
            label=info.get("label", nombre_hoja),
            filas=filas,
            columnas=headers,
            reconocida=bool(info),
            orden=info.get("orden", 99),
        ))

    wb.close()
    hojas.sort(key=lambda h: (0 if h.reconocida else 1, h.orden))

    return PreviewResponse(
        hojas=hojas,
        total_hojas=len(hojas),
        hojas_reconocidas=sum(1 for h in hojas if h.reconocida),
    )


# ── Endpoint 2: Importar ───────────────────────────────────────────────

@router.post("/importar", response_model=ImportarResponse,
             summary="Importar hojas seleccionadas del Excel a la BD")
async def importar_dims(
    file: UploadFile = File(..., description="Archivo Excel (.xlsx) con hojas DIM_*"),
    hojas: str = Form(..., description="JSON con lista de nombres de hojas a importar"),
    db: Session = Depends(get_db),
    _=AdminOGerProd,
):
    import json
    hojas_seleccionadas: List[str] = []
    _h = hojas.strip()
    try:
        parsed = json.loads(_h)
        if isinstance(parsed, list):
            hojas_seleccionadas = [str(x).strip() for x in parsed if x]
        elif isinstance(parsed, str):
            hojas_seleccionadas = [parsed.strip()]
        else:
            raise ValueError()
    except Exception:
        # Fallback: texto plano o CSV  →  ["TABLA_RM"] o "TABLA_RM,DIM_PAIS"
        hojas_seleccionadas = [x.strip() for x in _h.split(",") if x.strip()]
    if not hojas_seleccionadas:
        raise HTTPException(400, "El campo 'hojas' debe ser un JSON array de strings")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    contenido = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)

    resultados = []
    total_ins = total_om = total_err = 0

    try:
        def orden_hoja(nombre):
            info = HOJAS_CONOCIDAS.get(nombre.strip().upper(), {})
            return info.get("orden", 99)

        hojas_ordenadas = sorted(hojas_seleccionadas, key=orden_hoja)

        # Mapas de lookup para FKs
        pais_map: dict[str, str] = {p.codigo: p.codigo for p in db.query(Pais).filter(Pais.activo == True).all()}
        linea_map: dict = {}   # (pais_codigo, linea_id_excel) → id BD
        gerente_map: dict = {} # (pais_codigo, gerente_id_excel) → id BD
        indicador_id_map: dict = {}  # indicador_id_excel → id BD (por pais)

        # Pre-cargar mapas con datos ya en BD
        for lin in db.query(Linea).all():
            pc = next((k for k, v in pais_map.items() if v == lin.pais_codigo), None)
            if pc:
                linea_map[(pc, lin.codigo)] = lin.id

        for ger in db.query(Gerente).all():
            pc = next((k for k, v in pais_map.items() if v == ger.pais_codigo), None)
            if pc:
                gerente_map[(pc, ger.codigo)] = ger.id

        indicador_map: dict = {}  # (pais_codigo, codigo_indicador) → id BD
        for ind in db.query(Indicador).all():
            pc = next((k for k, v in pais_map.items() if v == ind.pais_codigo), None)
            if pc:
                indicador_map[(pc, ind.codigo)] = ind.id

        # Pre-build indicador_id_map desde la hoja DIM_INDICADOR del mismo Excel
        # (permite importar DIM_REGLA_PUNTUACION sin tener que importar DIM_INDICADOR primero)
        excel_ind_codigo_map: dict[int, str] = {}
        for sname in wb.sheetnames:
            if sname.strip().upper() in ("DIM_INDICADOR",):
                ws_ind = wb[sname]
                for row_ind in _leer_hoja(ws_ind):
                    ind_id_ex = _i(row_ind.get("INDICADOR_ID") or row_ind.get("ID"))
                    cod = _s(row_ind.get("CODIGO_INDICADOR") or row_ind.get("CODIGO"))
                    if ind_id_ex and cod:
                        excel_ind_codigo_map[ind_id_ex] = cod.upper()
        # Poblar indicador_id_map con los IDs de la BD usando el mapa de códigos
        for pc, pid in pais_map.items():
            for ind_id_ex, cod in excel_ind_codigo_map.items():
                db_id = indicador_map.get((pc, cod))
                if db_id:
                    indicador_id_map[(pc, ind_id_ex)] = db_id

        def _refrescar_mapas():
            """
            Reconstruye completamente todos los mapas de lookup desde BD + Excel.
            Se llama antes de cada hoja para garantizar que los datos insertados
            por hojas anteriores estén disponibles como FKs.
            """
            # 1. Refrescar pais_map desde BD
            pais_map.clear()
            for p in db.query(Pais).all():
                pais_map[p.codigo] = p.codigo

            # 2. Reconstruir linea_map: cruza hoja Excel DIM_LINEA con BD
            #    Guarda tanto (pais_c, linea_id_excel) como (pais_c, codigo)
            linea_map.clear()
            for sname in wb.sheetnames:
                if sname.strip().upper() in ("DIM_LINEA",):
                    for row_l in _leer_hoja(wb[sname]):
                        # Soporta: PAIS_CODIGO (sistema) y PAIS_ID (Excel, id numérico crudo)
                        pc     = _pais_c(pais_map, row_l.get("PAIS_CODIGO") or row_l.get("PAIS_ID"))
                        cod    = _s(row_l.get("CODIGO"))
                        # Soporta: LINEA_ID (sistema) e ID (Excel, PK propio de la hoja)
                        lex_id = _i(row_l.get("LINEA_ID") or row_l.get("ID"))
                        pid    = pais_map.get(pc)
                        if not pid or not cod:
                            continue
                        db_lin = db.query(Linea).filter(
                            Linea.pais_codigo == pid, Linea.codigo == cod
                        ).first()
                        if db_lin:
                            linea_map[(pc, cod)]    = db_lin.id
                            if lex_id is not None:
                                linea_map[(pc, lex_id)] = db_lin.id

            # 3. Reconstruir gerente_map: cruza DIM_GERENTE(_DISTRITO) con BD
            gerente_map.clear()
            for sname in wb.sheetnames:
                if sname.strip().upper() in ("DIM_GERENTE", "DIM_GERENTE_DISTRITO"):
                    for row_g in _leer_hoja(wb[sname]):
                        # Soporta: PAIS_CODIGO (sistema) y PAIS_ID (Excel, id numérico crudo)
                        pc     = _pais_c(pais_map, row_g.get("PAIS_CODIGO") or row_g.get("PAIS_ID"))
                        cod    = _s(row_g.get("CODIGO"))
                        # Soporta: GERENTE_ID (sistema) e ID (Excel, PK propio de la hoja)
                        gex_id = _i(row_g.get("GERENTE_ID") or row_g.get("ID"))
                        pid    = pais_map.get(pc)
                        if not pid or not cod:
                            continue
                        db_ger = db.query(Gerente).filter(
                            Gerente.pais_codigo == pid, Gerente.codigo == cod
                        ).first()
                        if db_ger:
                            gerente_map[(pc, cod)]    = db_ger.id
                            if gex_id is not None:
                                gerente_map[(pc, gex_id)] = db_ger.id

            # 4. Refrescar indicador_map desde BD
            indicador_map.clear()
            for ind in db.query(Indicador).all():
                pc = next((k for k, v in pais_map.items() if v == ind.pais_codigo), None)
                if pc:
                    indicador_map[(pc, ind.codigo)] = ind.id

            # 5. Reconstruir indicador_id_map (Excel ID → BD id)
            indicador_id_map.clear()
            for pc in list(pais_map.keys()):
                for ind_id_ex, cod in excel_ind_codigo_map.items():
                    db_id = indicador_map.get((pc, cod))
                    if db_id:
                        indicador_id_map[(pc, ind_id_ex)] = db_id

        for nombre_hoja in hojas_ordenadas:
            ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.strip().upper() == nombre_hoja.strip().upper():
                    ws = wb[sheet_name]
                    nombre_hoja = sheet_name
                    break

            nombre_upper = nombre_hoja.strip().upper()
            # Resolver alias → nombre canónico
            nombre_canonico = ALIAS.get(nombre_upper, nombre_upper)

            info = HOJAS_CONOCIDAS.get(nombre_upper, {})
            label = info.get("label", nombre_hoja)

            if ws is None:
                resultados.append(ResultadoHoja(
                    nombre_hoja=nombre_hoja, label=label,
                    exitoso=False, insertados=0, omitidos=0, errores=1,
                    mensaje=f"Hoja '{nombre_hoja}' no encontrada en el archivo",
                ))
                total_err += 1
                continue

            # Refrescar mapas ANTES de cada hoja para capturar lo insertado en hojas anteriores
            _refrescar_mapas()

            try:
                ins, om, err, msg = _importar_hoja(
                    db, nombre_canonico, ws,
                    pais_map, linea_map, gerente_map, indicador_map, indicador_id_map
                )
                resultados.append(ResultadoHoja(
                    nombre_hoja=nombre_hoja, label=label,
                    exitoso=err == 0, insertados=ins, omitidos=om, errores=err,
                    mensaje=msg,
                ))
                total_ins += ins; total_om += om; total_err += err
            except Exception as e:
                db.rollback()
                resultados.append(ResultadoHoja(
                    nombre_hoja=nombre_hoja, label=label,
                    exitoso=False, insertados=0, omitidos=0, errores=1,
                    mensaje=str(e)[:400],
                ))
                total_err += 1

    finally:
        wb.close()

    return ImportarResponse(
        resultados=resultados,
        total_insertados=total_ins,
        total_omitidos=total_om,
        total_errores=total_err,
    )


# ── Función interna: importar cada hoja ───────────────────────────────

def _importar_hoja(
    db: Session,
    nombre_canonico: str,
    ws,
    pais_map: dict,
    linea_map: dict,
    gerente_map: dict,
    indicador_map: dict,
    indicador_id_map: dict,
):
    """
    Importa una hoja. Retorna (insertados, omitidos, errores, mensaje).
    Maneja tanto los nombres originales del sistema como los alias del Excel.
    """
    rows = _leer_hoja(ws)
    ins = om = err = 0

    # ── DIM_PAIS ──────────────────────────────────────────────────────
    if nombre_canonico == "DIM_PAIS":
        for row in rows:
            # Soporta: PAIS_CODIGO (sistema) y CODIGO (Excel DIM_MIP_FINAL)
            codigo = _s(row.get("PAIS_CODIGO") or row.get("CODIGO"))
            if not codigo:
                err += 1; continue
            if db.query(Pais).filter(Pais.codigo == codigo).first():
                om += 1; continue
            obj = Pais(
                codigo=codigo,
                nombre=_s(row.get("NOMBRE")) or codigo,
                moneda=_s(row.get("MONEDA")),
                zona_horaria=_s(row.get("ZONA_HORARIA")),
                activo=True,
            )
            db.add(obj); db.flush()
            pais_map[codigo] = obj.id
            ins += 1
        db.commit()

    # ── DIM_LINEA ─────────────────────────────────────────────────────
    elif nombre_canonico == "DIM_LINEA":
        from sqlalchemy import text as _text

        # Si LINEA_ID del Excel es numérico, preservarlo como id de BD (IDENTITY_INSERT)
        usar_id_natural = bool(rows and _i(rows[0].get("LINEA_ID")))
        if usar_id_natural:
            db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_Linea] ON"))

        try:
            ids_usados: set = set()
            for row in rows:
                # Soporta: PAIS_CODIGO (sistema) y PAIS_ID (Excel, id numérico crudo)
                pais_c = _pais_c(pais_map, row.get("PAIS_CODIGO") or row.get("PAIS_ID"))
                codigo = _s(row.get("CODIGO"))
                linea_id_orig = _i(row.get("LINEA_ID"))
                pais_codigo = pais_map.get(pais_c)
                if not pais_codigo or not codigo:
                    err += 1; continue
                existing = db.query(Linea).filter(Linea.pais_codigo == pais_codigo, Linea.codigo == codigo).first()
                if existing:
                    if linea_id_orig is not None:
                        linea_map[(pais_c, linea_id_orig)] = existing.id
                    linea_map[(pais_c, codigo)] = existing.id
                    om += 1; continue
                obj = Linea(pais_codigo=pais_codigo, codigo=codigo, nombre=_s(row.get("NOMBRE")) or codigo, activo=True)
                # Usar LINEA_ID del Excel como id de BD si está libre (no ya ocupado en BD ni en este lote)
                if usar_id_natural and linea_id_orig is not None and linea_id_orig not in ids_usados:
                    if not db.query(Linea).filter(Linea.id == linea_id_orig).first():
                        obj.id = linea_id_orig
                        ids_usados.add(linea_id_orig)
                db.add(obj); db.flush()
                if linea_id_orig is not None:
                    linea_map[(pais_c, linea_id_orig)] = obj.id
                linea_map[(pais_c, codigo)] = obj.id
                ins += 1
        finally:
            if usar_id_natural:
                db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_Linea] OFF"))

        db.commit()

    # ── DIM_GERENTE (+ alias DIM_GERENTE_DISTRITO) ────────────────────
    elif nombre_canonico == "DIM_GERENTE":
        from sqlalchemy import text as _text

        # Si GERENTE_ID del Excel es numérico, preservarlo como id de BD (IDENTITY_INSERT)
        usar_id_natural = bool(rows and _i(rows[0].get("GERENTE_ID")))
        if usar_id_natural:
            db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_Gerente] ON"))

        try:
            ids_usados: set = set()
            for row in rows:
                codigo = _s(row.get("CODIGO"))
                # Soporta: PAIS_CODIGO (sistema) y PAIS_ID (Excel, id numérico crudo)
                pais_c = _pais_c(pais_map, row.get("PAIS_CODIGO") or row.get("PAIS_ID"))
                ger_id_orig = _i(row.get("GERENTE_ID"))
                pais_codigo = pais_map.get(pais_c)
                if not pais_codigo or not codigo:
                    err += 1; continue
                existing = db.query(Gerente).filter(Gerente.pais_codigo == pais_codigo, Gerente.codigo == codigo).first()
                if existing:
                    if ger_id_orig is not None:
                        gerente_map[(pais_c, ger_id_orig)] = existing.id
                    gerente_map[(pais_c, codigo)] = existing.id
                    om += 1; continue
                obj = Gerente(
                    pais_codigo=pais_codigo, codigo=codigo,
                    nombre=_s(row.get("NOMBRE")) or codigo,
                    tipo=_s(row.get("TIPO")) or "DISTRITO",
                    activo=True,
                )
                # Usar GERENTE_ID del Excel como id de BD si está libre
                if usar_id_natural and ger_id_orig is not None and ger_id_orig not in ids_usados:
                    if not db.query(Gerente).filter(Gerente.id == ger_id_orig).first():
                        obj.id = ger_id_orig
                        ids_usados.add(ger_id_orig)
                db.add(obj); db.flush()
                if ger_id_orig is not None:
                    gerente_map[(pais_c, ger_id_orig)] = obj.id
                gerente_map[(pais_c, codigo)] = obj.id
                ins += 1
        finally:
            if usar_id_natural:
                db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_Gerente] OFF"))

        db.commit()

    # ── DIM_RM (+ alias DIM_REPRESENTANTE_MEDICO) ─────────────────────
    elif nombre_canonico == "DIM_RM":
        from sqlalchemy import text as _text

        # Si RM_CODIGO es numérico, usarlo como id de BD (IDENTITY_INSERT)
        # para que FACT_ResultadoIndicador.rm_id coincida con el código del Excel.
        usar_id_natural = bool(rows and _i(
            _s(rows[0].get("RM_CODIGO") or rows[0].get("CODIGO_RM")
               or rows[0].get("CODIGO_ID") or rows[0].get("CODIGO_ORIGEN_EXCEL"))
        ))

        if usar_id_natural:
            db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_RM] ON"))

        try:
            for row in rows:
                # Soporta: RM_CODIGO/CODIGO_RM (sistema/Excel histórico) y
                # CODIGO_ID/CODIGO_ORIGEN_EXCEL (Excel DIM_MIP_FINAL actual del usuario)
                rm_codigo = _s(row.get("RM_CODIGO") or row.get("CODIGO_RM")
                               or row.get("CODIGO_ID") or row.get("CODIGO_ORIGEN_EXCEL"))
                # Soporta: PAIS_CODIGO con valor de código string o id numérico crudo
                pais_c    = _pais_c(pais_map, row.get("PAIS_CODIGO"))
                linea_id_orig  = _i(row.get("LINEA_ID"))
                ger_id_orig    = _i(row.get("GERENTE_ID"))
                # Soporta: NOMBRE (sistema) y NOMBRE_RM (Excel)
                nombre    = _s(row.get("NOMBRE") or row.get("NOMBRE_RM"))
                pais_codigo   = pais_map.get(pais_c)
                linea_id  = linea_map.get((pais_c, linea_id_orig))
                gerente_id = gerente_map.get((pais_c, ger_id_orig))

                if not pais_codigo or not linea_id or not rm_codigo:
                    err += 1; continue
                if db.query(RepresentanteMedico).filter(
                    RepresentanteMedico.codigo == str(rm_codigo)
                ).first():
                    om += 1; continue

                rm_id_natural = _i(rm_codigo) if usar_id_natural else None

                obj = RepresentanteMedico(
                    pais_codigo=pais_codigo, linea_id=linea_id, gerente_id=gerente_id,
                    codigo=str(rm_codigo),
                    nombre=nombre or str(rm_codigo),
                    cedula=_s(row.get("CEDULA")),
                    email=_s(row.get("EMAIL")),
                    fecha_ingreso=_fecha(row.get("FECHA_INGRESO")),
                    activo=True,
                )
                if rm_id_natural is not None:
                    obj.id = rm_id_natural
                db.add(obj)
                ins += 1
        finally:
            if usar_id_natural:
                db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_RM] OFF"))

        db.commit()

    # ── DIM_INDICADOR ─────────────────────────────────────────────────
    elif nombre_canonico == "DIM_INDICADOR":
        from sqlalchemy import text as _text

        # Si INDICADOR_ID del Excel es numérico, preservarlo como id de BD (IDENTITY_INSERT).
        # Para multi-país: el mismo INDICADOR_ID solo se puede usar una vez (primer país);
        # los países subsiguientes con el mismo id obtendrán un id auto-generado.
        usar_id_natural = bool(rows and _i(rows[0].get("INDICADOR_ID")))
        if usar_id_natural:
            db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_Indicador] ON"))

        try:
            ids_usados: set = set()  # IDs ya asignados en este lote o en BD
            for row in rows:
                # Soporta: CODIGO (sistema) y CODIGO_INDICADOR (Excel)
                codigo  = _s(row.get("CODIGO") or row.get("CODIGO_INDICADOR"))
                # Soporta: NOMBRE (sistema) y NOMBRE_INDICADOR (Excel)
                nombre  = _s(row.get("NOMBRE") or row.get("NOMBRE_INDICADOR"))
                # Soporta: MODULO (sistema) y BLOQUE_MODELO (Excel)
                modulo  = _s(row.get("MODULO") or row.get("BLOQUE_MODELO")) or "GESTION"
                # Soporta: PONDERACION_PCT (sistema) y PESO (Excel)
                pond    = _i(row.get("PONDERACION_PCT") or row.get("PESO")) or 0
                # INDICADOR_ID del Excel → se usa como id de BD
                ind_id_excel = _i(row.get("INDICADOR_ID"))

                if not codigo:
                    err += 1; continue

                pais_c_from_row = _pais_c(pais_map, row.get("PAIS_CODIGO"))

                # Si la fila trae PAIS_CODIGO, usar solo ese país; si no, crear para todos los países
                target_paises = (
                    {pais_c_from_row: pais_map[pais_c_from_row]}
                    if pais_c_from_row and pais_c_from_row in pais_map
                    else pais_map
                )

                for pais_c, pais_codigo in target_paises.items():
                    existing = db.query(Indicador).filter(
                        Indicador.pais_codigo == pais_codigo,
                        Indicador.codigo == codigo,
                    ).first()
                    if existing:
                        # Actualizar orden/ponderacion si INDICADOR_ID está disponible
                        if ind_id_excel is not None:
                            existing.orden = ind_id_excel
                        if pond:
                            existing.ponderacion_pct = pond
                            existing.peso_iup = Decimal(str(pond)) / Decimal("100")
                        db.flush()
                        indicador_map[(pais_c, codigo)] = existing.id
                        if ind_id_excel is not None:
                            indicador_id_map[(pais_c, ind_id_excel)] = existing.id
                        om += 1; continue

                    obj = Indicador(
                        pais_codigo=pais_codigo,
                        codigo=codigo,
                        nombre=nombre or codigo,
                        rol=_s(row.get("ROL")) or "RM",
                        modulo=modulo,
                        tipo_periodo=_s(row.get("TIPO_PERIODO")) or "CICLO",
                        ponderacion_pct=pond,
                        escala=_i(row.get("ESCALA")) or 100,
                        valor_min=_d(row.get("VALOR_MIN")),
                        valor_max=_d(row.get("VALOR_MAX")),
                        peso_iup=Decimal(str(pond)) / Decimal("100"),
                        orden=ind_id_excel or 0,
                        activo=True,
                    )
                    # Usar INDICADOR_ID del Excel como id de BD si está libre
                    if usar_id_natural and ind_id_excel is not None and ind_id_excel not in ids_usados:
                        if not db.query(Indicador).filter(Indicador.id == ind_id_excel).first():
                            obj.id = ind_id_excel
                            ids_usados.add(ind_id_excel)
                    db.add(obj); db.flush()
                    indicador_map[(pais_c, codigo)] = obj.id
                    if ind_id_excel is not None:
                        indicador_id_map[(pais_c, ind_id_excel)] = obj.id
                    ins += 1
        finally:
            if usar_id_natural:
                db.execute(_text("SET IDENTITY_INSERT [Config].[DIM_Indicador] OFF"))

        db.commit()

    # ── DIM_INDICADOR_TABLA (+ alias DIM_REGLA_PUNTUACION) ───────────
    elif nombre_canonico == "DIM_INDICADOR_TABLA":
        # Primero: recolectar todos los (ind_id_bd, pais_codigo) únicos a importar
        # y borrar sus registros existentes (para reemplazar datos incorrectos)
        combinaciones_a_limpiar: set = set()
        for row in rows:
            ind_codigo   = _s(row.get("INDICADOR_CODIGO"))
            ind_id_excel = _i(row.get("INDICADOR_ID"))
            pais_c_from_row = _pais_c(pais_map, row.get("PAIS_CODIGO"))
            target_paises = (
                {pais_c_from_row: pais_map[pais_c_from_row]}
                if pais_c_from_row and pais_c_from_row in pais_map
                else pais_map
            )
            for pais_c, pais_codigo in target_paises.items():
                ind_id_bd = None
                if ind_codigo:
                    ind_id_bd = indicador_map.get((pais_c, ind_codigo.upper()))
                if not ind_id_bd and ind_id_excel is not None:
                    ind_id_bd = indicador_id_map.get((pais_c, ind_id_excel))
                if ind_id_bd:
                    combinaciones_a_limpiar.add((ind_id_bd, pais_codigo))

        for ind_id_bd, pais_codigo in combinaciones_a_limpiar:
            deleted = db.query(IndicadorTabla).filter(
                IndicadorTabla.indicador_id == ind_id_bd,
                IndicadorTabla.pais_codigo == pais_codigo,
            ).delete(synchronize_session=False)
            om += deleted  # contar como omitidos/reemplazados

        # Segundo: insertar los nuevos registros
        for row in rows:
            ind_codigo   = _s(row.get("INDICADOR_CODIGO"))
            ind_id_excel = _i(row.get("INDICADOR_ID"))
            pais_c_from_row = _pais_c(pais_map, row.get("PAIS_CODIGO"))

            desde = _d(row.get("VALOR_DESDE") or row.get("RESULTADO_MIN"))
            hasta_raw = row.get("VALOR_HASTA")
            if hasta_raw in (None, ""):
                hasta_raw = row.get("RESULTADO_MAX")
            hasta = _d(hasta_raw)
            if hasta is None and desde is not None:
                hasta = Decimal("999999")

            # Soporta: PUNTOS (sistema) y PUNTOS_REGLA (Excel)
            puntos = _d(row.get("PUNTOS") or row.get("PUNTOS_REGLA"))

            if desde is None:
                err += 1; continue

            target_paises = (
                {pais_c_from_row: pais_map[pais_c_from_row]}
                if pais_c_from_row and pais_c_from_row in pais_map
                else pais_map
            )

            for pais_c, pais_codigo in target_paises.items():
                ind_id_bd = None
                if ind_codigo:
                    ind_id_bd = indicador_map.get((pais_c, ind_codigo.upper()))
                if not ind_id_bd and ind_id_excel is not None:
                    ind_id_bd = indicador_id_map.get((pais_c, ind_id_excel))

                if not ind_id_bd:
                    err += 1; continue

                db.add(IndicadorTabla(
                    indicador_id=ind_id_bd, pais_codigo=pais_codigo,
                    rango_desde=desde, rango_hasta=hasta,
                    puntos=puntos, activo=True,
                ))
                ins += 1
        db.commit()

    # ── DIM_TIEMPO → popula DIM_Ciclo + DIM_Mes ──────────────────────
    elif nombre_canonico == "DIM_TIEMPO":
        trimestre_map = {1:1,2:1,3:1,4:2,5:2,6:2,7:3,8:3,9:3,10:4,11:4,12:4}
        semestre_map  = {1:1,2:1,3:1,4:1,5:1,6:1,7:2,8:2,9:2,10:2,11:2,12:2}
        for row in rows:
            anio       = _i(row.get("ANIO")) or 2026
            ciclo_num  = _i(row.get("CICLO"))
            mes_num    = _i(row.get("MES_NUMERO") or row.get("MESID"))
            mes_nombre = _s(row.get("MES_NOMBRE") or row.get("MES"))
            trimestre  = _i(row.get("TRIMESTRE")) or trimestre_map.get(mes_num or 1, 1)

            # Insertar en DIM_Mes
            if mes_num and not db.query(Mes).filter(Mes.mes == mes_num, Mes.anio == anio).first():
                abrev = mes_nombre[:3] if mes_nombre else str(mes_num)
                db.add(Mes(
                    anio=anio, mes=mes_num,
                    nombre=mes_nombre,
                    abrev=abrev,
                    ciclo_mes=ciclo_num,
                    trimestre=trimestre,
                    semestre=semestre_map.get(mes_num, 1),
                ))
                ins += 1
            else:
                om += 1

            # Insertar en DIM_Ciclo (uno por país)
            if ciclo_num:
                for pais_c, pais_codigo in pais_map.items():
                    if not db.query(Ciclo).filter(
                        Ciclo.pais_codigo == pais_codigo,
                        Ciclo.numero == ciclo_num,
                        Ciclo.anio == anio,
                    ).first():
                        db.add(Ciclo(
                            pais_codigo=pais_codigo, anio=anio, numero=ciclo_num,
                            nombre=f"C{ciclo_num:02d}-{anio}",
                            nombre_canonico=f"Ciclo {ciclo_num} {anio}",
                            fecha_inicio=date(anio, mes_num or 1, 1),
                            fecha_fin=date(anio, mes_num or 12, 28),
                            dias_laborables=21, cerrado=False, activo=True,
                        ))
        db.commit()

    # ── DIM_CICLO (formato original del sistema) ──────────────────────
    elif nombre_canonico == "DIM_CICLO":
        for row in rows:
            numero = _i(row.get("CICLO_ID"))
            nombre = _s(row.get("CICLO_NOMBRE_CORTO"))
            canon  = _s(row.get("NOMBRE_CANONICO"))
            if not numero: continue
            for pais_c, pais_codigo in pais_map.items():
                if db.query(Ciclo).filter(Ciclo.pais_codigo == pais_codigo, Ciclo.numero == numero, Ciclo.anio == 2026).first():
                    om += 1; continue
                db.add(Ciclo(
                    pais_codigo=pais_codigo, anio=2026, numero=numero,
                    nombre=nombre, nombre_canonico=canon,
                    fecha_inicio=date(2026, 1, 1), fecha_fin=date(2026, 12, 31),
                    dias_laborables=21, cerrado=False, activo=True,
                ))
                ins += 1
        db.commit()

    # ── DIM_MES (formato original del sistema) ────────────────────────
    elif nombre_canonico == "DIM_MES":
        trimestre_map = {1:1,2:1,3:1,4:2,5:2,6:2,7:3,8:3,9:3,10:4,11:4,12:4}
        semestre_map  = {1:1,2:1,3:1,4:1,5:1,6:1,7:2,8:2,9:2,10:2,11:2,12:2}
        for row in rows:
            mes_num = _i(row.get("MESID"))
            if not mes_num: continue
            if db.query(Mes).filter(Mes.mes == mes_num, Mes.anio == 2026).first():
                om += 1; continue
            db.add(Mes(
                anio=2026, mes=mes_num,
                nombre=_s(row.get("MES")),
                abrev=_s(row.get("MESABREV")),
                ciclo_mes=_i(row.get("CICLOMES")),
                trimestre=trimestre_map.get(mes_num, 1),
                semestre=semestre_map.get(mes_num, 1),
            ))
            ins += 1
        db.commit()

    # ── DIM_RECONOCIMIENTO → DIM_Premio ──────────────────────────────
    elif nombre_canonico == "DIM_RECONOCIMIENTO":
        for row in rows:
            # Soporta: CODIGO (sistema) y CODIGO_RECONOCIMIENTO (Excel)
            codigo = _s(row.get("CODIGO") or row.get("CODIGO_RECONOCIMIENTO"))
            nombre = _s(row.get("NOMBRE") or row.get("NOMBRE_RECONOCIMIENTO"))
            if not codigo: continue
            if db.query(Premio).filter(Premio.codigo == codigo).first():
                om += 1; continue
            db.add(Premio(
                codigo=codigo,
                nombre=nombre or codigo,
                categoria="RM",
                frecuencia="MENSUAL",
                activo=True,
            ))
            ins += 1
        db.commit()

    # ── DIM_META_INDICADOR → DIM_MetaIndicador ───────────────────────
    elif nombre_canonico == "DIM_META_INDICADOR":
        for row in rows:
            # Soporta: INDICADOR_CODIGO (sistema) e INDICADOR_ID numérico (Excel)
            ind_codigo      = _s(row.get("INDICADOR_CODIGO"))
            ind_id_excel    = _i(row.get("INDICADOR_ID"))
            pais_c_from_row = _pais_c(pais_map, row.get("PAIS_CODIGO"))

            peso     = _d(row.get("PESO")) or Decimal("0")
            minimo   = _d(row.get("MINIMO"))
            objetivo = _d(row.get("OBJETIVO"))
            maximo   = _d(row.get("MAXIMO"))
            pmax     = _d(row.get("PUNTAJE_MAXIMO"))
            meta100  = _d(row.get("META_100"))
            tipo_calc = _s(row.get("TIPO_CALCULO"))
            orden    = _i(row.get("ORDEN_DASHBOARD"))

            if ind_codigo is None and ind_id_excel is None:
                err += 1; continue

            # La hoja no trae PAIS_CODIGO → la meta aplica al indicador
            # homólogo de cada país (igual criterio que DIM_INDICADOR_TABLA)
            target_paises = (
                {pais_c_from_row: pais_map[pais_c_from_row]}
                if pais_c_from_row and pais_c_from_row in pais_map
                else pais_map
            )

            for pais_c, pais_codigo in target_paises.items():
                ind_id_bd = None
                if ind_codigo:
                    ind_id_bd = indicador_map.get((pais_c, ind_codigo))
                elif ind_id_excel is not None:
                    ind_id_bd = indicador_id_map.get((pais_c, ind_id_excel))

                if not ind_id_bd:
                    err += 1; continue

                existing = db.query(MetaIndicador).filter(
                    MetaIndicador.indicador_id == ind_id_bd
                ).first()
                if existing:
                    om += 1; continue

                db.add(MetaIndicador(
                    indicador_id=ind_id_bd,
                    peso=peso,
                    minimo=minimo, objetivo=objetivo, maximo=maximo,
                    puntaje_maximo=pmax, meta_100=meta100,
                    tipo_calculo=tipo_calc, orden_dashboard=orden,
                    activo=True,
                ))
                ins += 1
        db.commit()

    # ── DIM_CATEGORIA_DESEMPENO → DIM_CategoriaDesempeno ──────────────
    # Tabla global de catálogo (sin PAIS_CODIGO, sin FK): inserta si el
    # CODIGO_CATEGORIA no existe todavía.
    elif nombre_canonico == "DIM_CATEGORIA_DESEMPENO":
        for row in rows:
            codigo = _s(row.get("CODIGO_CATEGORIA"))
            nombre = _s(row.get("NOMBRE_CATEGORIA"))
            if not codigo:
                err += 1; continue

            existing = db.query(CategoriaDesempeno).filter(
                CategoriaDesempeno.codigo == codigo
            ).first()
            if existing:
                om += 1; continue

            db.add(CategoriaDesempeno(
                codigo=codigo,
                nombre=nombre or codigo,
                score_min=_d(row.get("SCORE_MIN")),
                score_max=_d(row.get("SCORE_MAX")),
                color_dashboard=_s(row.get("COLOR_DASHBOARD")),
                activo=True,
            ))
            ins += 1
        db.commit()

    # ── DIM_KPI_DASHBOARD → DIM_KpiDashboard ──────────────────────────
    # Tabla global de catálogo (sin PAIS_CODIGO, sin FK): inserta si el
    # CODIGO_KPI_DASHBOARD no existe todavía.
    elif nombre_canonico == "DIM_KPI_DASHBOARD":
        for row in rows:
            codigo = _s(row.get("CODIGO_KPI_DASHBOARD"))
            nombre = _s(row.get("NOMBRE_KPI_DASHBOARD"))
            if not codigo:
                err += 1; continue

            existing = db.query(KpiDashboard).filter(
                KpiDashboard.codigo == codigo
            ).first()
            if existing:
                om += 1; continue

            db.add(KpiDashboard(
                codigo=codigo,
                nombre=nombre or codigo,
                pagina_dashboard=_s(row.get("PAGINA_DASHBOARD")),
                tipo_calculo=_s(row.get("TIPO_CALCULO")),
                descripcion=_s(row.get("DESCRIPCION")),
                activo=True,
            ))
            ins += 1
        db.commit()

    # ── TABLAS_RM — Tablas de factores en formato multi-bloque ───────────
    # Hoja 'Tablas RM' del Excel TABLA_RM.xlsx.
    # Estructura: bloques de 3 cols [INDICADOR_ID | valor | factor].
    # Cada celda puntos = ROUND(factor × peso_indicador, 4) para que el SP
    # de puntajes haga solo el floor-match y devuelva puntos ya calculados.
    elif nombre_canonico == "TABLAS_RM":
        from collections import defaultdict

        tablas_data = _leer_tablas_rm(ws)
        if not tablas_data:
            return 0, 0, 1, (
                "No se pudieron leer datos de 'Tablas RM' — "
                "verifique que la fila de headers diga 'INDICADOR_ID' en la columna A"
            )

        # Mapa id_BD → (codigo, nombre) para poblar los campos descriptivos
        ind_info_map: dict[int, tuple[str, str]] = {
            ind.id: (ind.codigo or "", ind.nombre or "")
            for ind in db.query(Indicador).all()
        }

        # Agrupar entradas por indicador_id del Excel
        bloques: dict[int, list[dict]] = defaultdict(list)
        for entry in tablas_data:
            bloques[entry["indicador_id"]].append(entry)

        for ind_id_excel, entradas in bloques.items():
            # PESO viene directamente de la columna PESO del Excel (no se consulta BD)
            peso_bloque = entradas[0]["peso"]

            for pais_c, pais_codigo in pais_map.items():
                # Resolver indicador_id de BD
                ind_id_bd = indicador_id_map.get((pais_c, ind_id_excel))
                if not ind_id_bd:
                    continue  # indicador no importado para este país → saltar

                # Reemplazar: borrar registros existentes para este (indicador, país)
                deleted = db.query(IndicadorTabla).filter(
                    IndicadorTabla.indicador_id == ind_id_bd,
                    IndicadorTabla.pais_codigo == pais_codigo,
                ).delete(synchronize_session=False)
                om += deleted

                # Insertar nuevos registros
                for entrada in entradas:
                    # Usar PESO de la fila; si es 0, usar el del bloque como fallback
                    valor_d = Decimal(str(round(entrada["valor"], 4)))
                    # puntos almacena solo el FACTOR puro (ej: 1.08).
                    # El SP multiplica factor × peso en runtime:
                    #   puntos_obtenidos = factor × DIM_Indicador.ponderacion_pct
                    puntos = Decimal(str(round(entrada["factor"], 4)))
                    # Código y nombre real desde DIM_Indicador (BD)
                    cod_bd, nom_bd = ind_info_map.get(ind_id_bd, ("", ""))
                    # Fallback al código que viene del Excel si la BD no lo tiene
                    cod = cod_bd or entrada.get("codigo_indicador", "")
                    nom = nom_bd or entrada.get("codigo_indicador", "")
                    val_label = int(entrada["valor"]) if entrada["valor"] == int(entrada["valor"]) else round(entrada["valor"], 2)
                    fac_pct = f"{round(entrada['factor']*100)}%"
                    desc = f"{cod} | {val_label} | {fac_pct}"
                    db.add(IndicadorTabla(
                        indicador_id=ind_id_bd,
                        pais_codigo=pais_codigo,
                        codigo_indicador=cod,
                        nombre_indicador=nom,
                        rango_desde=valor_d,
                        rango_hasta=valor_d,   # floor-match en el SP
                        puntos=puntos,
                        descripcion=desc,
                        activo=True,
                    ))
                    ins += 1

        db.commit()

    # ── DIM_MEDICOCOBERTURA (universo médico de Cobertura Predictiva/4DX) ──
    # Mapea a Config.DIM_Medico vía resolver_medico_por_codigo() — dedup por
    # (pais_codigo, codigo), distinto del dedup por (pais_codigo, nombre, centro)
    # usado por el Excel de Categorización Médica (ver resolver_medico()).
    elif nombre_canonico == "DIM_MEDICOCOBERTURA":
        for row in rows:
            pais_c = _pais_c(pais_map, row.get("PAIS_CODIGO") or row.get("PAIS_ID"))
            pais_codigo = pais_map.get(pais_c)
            codigo = _s(row.get("MEDICO_ID"))
            nombre = _s(row.get("MEDICO"))
            if not pais_codigo or not codigo or not nombre:
                err += 1; continue

            activo_raw = _s(row.get("ACTIVO"))
            activo = (
                True if activo_raw is None
                else activo_raw.upper() in ("SI", "S", "TRUE", "1", "X", "YES", "Y")
            )

            existing = db.query(Medico).filter(
                Medico.pais_codigo == pais_codigo, Medico.codigo == codigo,
            ).first()

            try:
                resolver_medico_por_codigo(
                    db, pais_codigo, codigo=codigo, nombre=nombre,
                    especialidad_nombre=_s(row.get("ESPECIALIDAD")),
                    activo=activo,
                )
            except Exception:
                db.rollback()
                err += 1; continue

            if existing:
                om += 1
            else:
                ins += 1

        db.commit()

    else:
        return 0, 0, 0, f"Hoja '{nombre_canonico}' no tiene handler — omitida"

    return ins, om, err, f"{ins} insertados, {om} ya existían, {err} errores"
