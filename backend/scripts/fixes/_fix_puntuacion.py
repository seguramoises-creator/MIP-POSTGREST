# --- credenciales desde backend/.env (parametrizado, no hardcodear) ---
import os as _os, pathlib as _pl
try:
    from dotenv import load_dotenv as _ld
    _ld(_pl.Path(__file__).resolve().parents[2] / '.env')
except Exception:
    pass
os = _os
# ----------------------------------------------------------------------
"""
_fix_puntuacion.py  (v2)
========================
Corrige Config.DIM_IndicadorTabla y recalcula todos los ciclos abiertos.

    cd C:\\Users\\Lenovo\\Proyecto\\MSM\\backend
    .\\venv\\Scripts\\activate
    python _fix_puntuacion.py
"""

import sys
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent / "DIMS_FACTS_V2" / "DIM_MIP_FINAL.xlsx"

DB_SERVER   = "127.0.0.1"
DB_PORT     = 1433
DB_NAME     = "SCGCPR"
DB_USER     = "segura"
DB_PASSWORD = os.environ.get('DB_PASSWORD', '')


# ── helpers ──────────────────────────────────────────────────────────────

def _leer_hoja(ws) -> list[dict]:
    hdrs = [str(c.value).strip().upper() if c.value is not None else f"_COL{i}"
            for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({hdrs[i]: row[i] for i in range(len(hdrs))})
    return rows


def _i(v):
    try:
        return int(v)
    except Exception:
        return None


def _f(v):
    try:
        return float(v)
    except Exception:
        return None


# ── main ─────────────────────────────────────────────────────────────────

def main():
    import pymssql
    import openpyxl

    print("=" * 60)
    print("FIX PUNTUACION v2")
    print("=" * 60)

    # ── 1. Excel ─────────────────────────────────────────────────────────
    if not EXCEL_PATH.exists():
        print(f"ERROR: no existe {EXCEL_PATH}")
        sys.exit(1)

    print(f"\n[1/4] Leyendo Excel (sin read_only) ...")
    # SIN read_only para evitar problemas de stream
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    print(f"  Hojas: {wb.sheetnames}")

    # DIM_INDICADOR → {ind_id_excel: codigo_upper}
    ind_codigo_map: dict[int, str] = {}
    for sname in wb.sheetnames:
        if sname.strip().upper() == "DIM_INDICADOR":
            rows_ind = _leer_hoja(wb[sname])
            print(f"  DIM_INDICADOR: {len(rows_ind)} filas leídas")
            for r in rows_ind:
                ind_id = _i(r.get("INDICADOR_ID"))
                cod    = r.get("CODIGO_INDICADOR") or r.get("CODIGO")
                if ind_id is not None and cod:
                    ind_codigo_map[ind_id] = str(cod).strip().upper()
            break

    print(f"  Mapa ind_id→codigo: {len(ind_codigo_map)} entradas")
    for k, v in sorted(ind_codigo_map.items()):
        print(f"    {k} → {v}")

    # DIM_REGLA_PUNTUACION → lista de rangos
    reglas: list[dict] = []
    for sname in wb.sheetnames:
        if sname.strip().upper() == "DIM_REGLA_PUNTUACION":
            rows_reg = _leer_hoja(wb[sname])
            print(f"\n  DIM_REGLA_PUNTUACION: {len(rows_reg)} filas leídas")
            if rows_reg:
                print(f"  Cabeceras detectadas: {list(rows_reg[0].keys())}")
                print(f"  Primera fila: {rows_reg[0]}")
            for r in rows_reg:
                ind_id = _i(r.get("INDICADOR_ID"))
                r_min  = _f(r.get("RESULTADO_MIN"))
                r_max  = _f(r.get("RESULTADO_MAX"))
                puntos = _f(r.get("PUNTOS_REGLA")) or _f(r.get("PUNTOS"))
                if ind_id is None or r_min is None or puntos is None:
                    continue
                hasta = r_max if r_max is not None else 999999.0
                reglas.append({
                    "ind_id_excel": ind_id,
                    "desde": r_min,
                    "hasta": hasta,
                    "puntos": puntos,
                })
            break

    print(f"  Reglas cargadas: {len(reglas)}")
    if not reglas:
        print("  ERROR: no se pudieron leer las reglas del Excel. Abortando.")
        sys.exit(1)

    from collections import Counter
    cnt = Counter(r["ind_id_excel"] for r in reglas)
    for k, v in sorted(cnt.items()):
        print(f"    ID={k} ({ind_codigo_map.get(k,'?')}): {v} rangos, "
              f"pts: {[r['puntos'] for r in reglas if r['ind_id_excel']==k][:3]}...")

    # ── 2. Conectar BD ────────────────────────────────────────────────────
    print(f"\n[2/4] Conectando SQL Server {DB_SERVER}/{DB_NAME} ...")
    conn = pymssql.connect(
        server=DB_SERVER, port=DB_PORT,
        database=DB_NAME, user=DB_USER, password=DB_PASSWORD,
        as_dict=True,
    )
    cur = conn.cursor()
    print("  OK")

    # ── 3. Mapa código → (pais_id, ind_id_bd) en BD ───────────────────
    cur.execute("SELECT id, pais_id, codigo FROM Config.DIM_Indicador WHERE activo = 1")
    ind_en_bd = cur.fetchall()
    print(f"\n[3/4] Indicadores en BD: {len(ind_en_bd)}")

    from collections import defaultdict
    cod_to_bd: dict[str, list] = defaultdict(list)
    for row in ind_en_bd:
        cod = (row["codigo"] or "").strip().upper()
        if cod:
            cod_to_bd[cod].append((row["pais_id"], row["id"]))

    print("  Primeros 5 códigos en BD:", list(cod_to_bd.keys())[:5])

    total_borrados = 0
    total_insertados = 0
    total_errores = 0

    for ind_id_ex, cod in sorted(ind_codigo_map.items()):
        pares = cod_to_bd.get(cod)
        if not pares:
            print(f"  AVISO: '{cod}' no encontrado en BD — saltado")
            continue

        rangos_ind = [r for r in reglas if r["ind_id_excel"] == ind_id_ex]
        if not rangos_ind:
            print(f"  AVISO: sin rangos para '{cod}' (ID={ind_id_ex}) — saltado")
            continue

        print(f"  {cod}: {len(pares)} país(es) × {len(rangos_ind)} rangos ...", end=" ")
        ins_cod = 0
        for pais_id, ind_id_bd in pares:
            # Borrar
            cur.execute(
                "DELETE FROM Config.DIM_IndicadorTabla "
                "WHERE indicador_id = %d AND pais_id = %d",
                (ind_id_bd, pais_id),
            )
            total_borrados += cur.rowcount if cur.rowcount > 0 else 0

            # Insertar
            for rg in rangos_ind:
                try:
                    cur.execute(
                        "INSERT INTO Config.DIM_IndicadorTabla "
                        "(indicador_id, pais_id, rango_desde, rango_hasta, puntos, activo) "
                        "VALUES (%d, %d, %s, %s, %s, 1)",
                        (ind_id_bd, pais_id,
                         str(rg["desde"]), str(rg["hasta"]), str(rg["puntos"])),
                    )
                    ins_cod += 1
                except Exception as e:
                    print(f"\n    ERR INSERT ({cod} pais={pais_id} desde={rg['desde']}): {e}")
                    total_errores += 1

        total_insertados += ins_cod
        print(f"insertados {ins_cod}")

    conn.commit()
    print(f"\n  Borradas: {total_borrados} filas antiguas")
    print(f"  Insertadas: {total_insertados} filas nuevas")
    if total_errores:
        print(f"  ERRORES de INSERT: {total_errores}")

    # ── 4. Encontrar qué ciclos tienen datos KPI ─────────────────────
    print("\n[4/4] Buscando ciclos con datos en FACT_ResultadoIndicador ...")
    cur.execute("""
        SELECT TOP 10
            ri.ciclo_id,
            c.nombre AS ciclo_nombre,
            c.cerrado,
            COUNT(*) AS total_filas,
            SUM(CASE WHEN ri.resultado_porcentaje IS NOT NULL THEN 1 ELSE 0 END) AS con_pct,
            SUM(CASE WHEN ri.puntos_obtenidos IS NOT NULL THEN 1 ELSE 0 END) AS con_pts
        FROM DW.FACT_ResultadoIndicador ri
        LEFT JOIN Config.DIM_Ciclo c ON c.id = ri.ciclo_id
        WHERE ri.activo = 1
        GROUP BY ri.ciclo_id, c.nombre, c.cerrado
        ORDER BY ri.ciclo_id DESC
    """)
    ciclos_con_datos = cur.fetchall()

    if not ciclos_con_datos:
        print("  No hay datos en FACT_ResultadoIndicador — nada que recalcular")
        ciclos = []
    else:
        print(f"  {'CicloID':>8}  {'Nombre':<15}  {'Cerrado':>7}  {'Filas':>6}  {'ConPct':>6}  {'ConPts':>6}")
        print("  " + "-" * 60)
        for r in ciclos_con_datos:
            print(f"  {r['ciclo_id']:>8}  {str(r['ciclo_nombre'] or ''):<15}  "
                  f"{str(r['cerrado']):>7}  {r['total_filas']:>6}  "
                  f"{r['con_pct']:>6}  {r['con_pts']:>6}")

        # ── UPDATE directo: recalcular puntos_obtenidos para TODOS los
        #    ciclos con datos, sin importar si están cerrados o abiertos.
        print()
        print("  Actualizando puntos_obtenidos directamente (bypasa ciclo cerrado) ...")
        cur.execute("""
            UPDATE ri
            SET ri.puntos_obtenidos =
                CASE
                    WHEN (SELECT COUNT(*) FROM Config.DIM_IndicadorTabla t2
                          WHERE t2.indicador_id = ri.indicador_id
                            AND t2.pais_id = ri.pais_id AND t2.activo = 1) = 0
                        THEN CASE WHEN ri.resultado_porcentaje > 100 THEN 100
                                  WHEN ri.resultado_porcentaje < 0   THEN 0
                                  ELSE ri.resultado_porcentaje END
                    WHEN (SELECT TOP 1 t.puntos
                          FROM Config.DIM_IndicadorTabla t
                          WHERE t.indicador_id = ri.indicador_id
                            AND t.pais_id = ri.pais_id AND t.activo = 1
                            AND ri.resultado_porcentaje
                                BETWEEN t.rango_desde AND t.rango_hasta
                          ORDER BY t.rango_desde ASC) IS NOT NULL
                        THEN (SELECT TOP 1 t.puntos
                              FROM Config.DIM_IndicadorTabla t
                              WHERE t.indicador_id = ri.indicador_id
                                AND t.pais_id = ri.pais_id AND t.activo = 1
                                AND ri.resultado_porcentaje
                                    BETWEEN t.rango_desde AND t.rango_hasta
                              ORDER BY t.rango_desde ASC)
                    ELSE (SELECT TOP 1 t.puntos
                          FROM Config.DIM_IndicadorTabla t
                          WHERE t.indicador_id = ri.indicador_id
                            AND t.pais_id = ri.pais_id AND t.activo = 1
                          ORDER BY t.rango_desde DESC)
                END,
                ri.fecha_calculo = SYSUTCDATETIME()
            FROM DW.FACT_ResultadoIndicador ri
            WHERE ri.activo = 1
              AND ri.resultado_porcentaje IS NOT NULL
        """)
        filas_act = cur.rowcount
        conn.commit()
        print(f"  puntos_obtenidos actualizados: {filas_act} filas")

        # ── Intentar SP para ciclos abiertos (para regenerar rankings) ───
        print()
        ciclos_abiertos = [r for r in ciclos_con_datos if not r["cerrado"]]
        if ciclos_abiertos:
            print("  Regenerando rankings para ciclos abiertos ...")
            for r in ciclos_abiertos:
                cid   = r["ciclo_id"]
                cnomb = r["ciclo_nombre"] or f"ID={cid}"
                print(f"    Ciclo {cid} '{cnomb}' ...", end=" ")
                try:
                    cur.execute(
                        "EXEC DW.sp_RecalcularCiclo @ciclo_id = %d, @pais_id = NULL",
                        (cid,),
                    )
                    rows_sp = cur.fetchall()
                    conn.commit()
                    if rows_sp:
                        rs = rows_sp[0]
                        if rs.get("abortado"):
                            print(f"ABORTADO: {rs.get('motivo')}")
                        else:
                            print(f"OK — {rs.get('filas_kpi_actualizadas','?')} KPI, "
                                  f"{rs.get('rankings_generados','?')} rankings")
                    else:
                        print("OK")
                except Exception as e:
                    conn.rollback()
                    print(f"ERR: {e}")
        else:
            print("  No hay ciclos abiertos con datos — rankings no regenerados")
            print("  (los puntos_obtenidos ya fueron actualizados directamente)")
        ciclos = ciclos_con_datos

    cur.close()
    conn.close()

    print()
    print("=" * 60)
    print("COMPLETADO")
    print(f"  Rangos insertados: {total_insertados}")
    print(f"  Ciclos con datos encontrados: {len(ciclos)}")
    if total_insertados > 0:
        print()
        print("  Recarga Productividad — Ptos deben mostrar 5-26,")
        print("  TOTAL ~80-110.")
    else:
        print()
        print("  PROBLEMA: 0 rangos insertados. Ver mensajes AVISO/ERR arriba.")
    print("=" * 60)


if __name__ == "__main__":
    main()
